import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

# Column map (1-based):
# F=6 PowerBI | G=7 Tableau | H=8 PBI Proof | I=9 TAB Proof | J=10 Findings/Conclusion | K=11 Status

# ─── S.No 15 | Cross-filtering and drill-down/drill-through | Row 16 ───────

ws.cell(16, 6).value = (
    'Automatic cross-filtering across all visuals on the same page with zero setup. '
    'Drill-down replaces chart content level-by-level using a hierarchy (e.g. Brand > Unit). '
    'Drill-through via right-click on any data point; filter context passed automatically; '
    'Back button auto-generated on the target page.'
)

ws.cell(16, 7).value = (
    'Cross-filtering requires manual Filter Actions per source-target visual pair — not automatic. '
    'Drill-down expands hierarchy rows in place; the chart does not replace itself with the next level. '
    'Drill-through needs an explicit Navigate Dashboard Action; no automatic back navigation.'
)

ws.cell(16, 10).value = (
    'Power BI wins. '
    'All three behaviors — cross-filtering, drill-down, and drill-through — work out of the box '
    'with zero configuration in Power BI. '
    'Tableau requires a manual Dashboard Action for every source-target pair, hierarchy expansion '
    'behaves differently (rows grow rather than the chart replacing itself), and there is no automatic '
    'back navigation after drill-through. '
    'For GTF FBC dashboards where drilling Brand > Unit > Unit Detail is a daily workflow, '
    'Power BI reduces both authoring time and end-user friction significantly.'
)

# ─── S.No 16 | Tooltip customization and rich interactivity | Row 17 ───────

ws.cell(17, 6).value = (
    'Report page tooltips: a dedicated tooltip page is authored, toggled as tooltip-type, '
    'and attached to the host visual. Fixed canvas size (320x240 px). Hover-only — cannot interact. '
    'Cannot mix free text and a viz in the same tooltip. Approx 20 min to author.'
)

ws.cell(17, 7).value = (
    'Viz-in-Tooltip: embeds any existing worksheet into a tooltip via a single tag in the tooltip editor. '
    'Reuses sheets already built — no rebuild needed. Supports mixed text and viz in one editor. '
    'Canvas size is configurable. Approx 5 min if the source sheet already exists.'
)

ws.cell(17, 10).value = (
    'No decisive winner — this is not a differentiating factor for GTF. '
    'Both tools deliver comparable rich tooltips that filter to the hovered data point. '
    'Tableau is faster to author (reuses existing sheets, no dedicated page needed) and more flexible '
    '(configurable canvas size, text + viz in same editor). '
    'Power BI requires a separately authored tooltip page and has a fixed canvas size, '
    'but produces the same visual outcome for the end user. '
    'Neither tool has a meaningful advantage that would influence the selection decision.'
)

wb.save(FILEPATH)
print('Saved.')

# Verify
wb2 = openpyxl.load_workbook(FILEPATH)
ws2 = wb2['Evaluation Matrix - 1']
for rn, label in [(16, 'S.No 15'), (17, 'S.No 16')]:
    r = ws2[rn]
    print(f'\n{label}')
    print(f'  F (PowerBI)   : {str(r[5].value)[:110]}')
    print(f'  G (Tableau)   : {str(r[6].value)[:110]}')
    print(f'  J (Conclusion): {str(r[9].value)[:120]}')
    print(f'  K (Status)    : {r[10].value}')
