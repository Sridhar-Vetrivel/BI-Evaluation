import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 171: Git integration / version control ──
u(171,
  f="""Native Git integration — GA as of 2024, built into Fabric/Power BI Service.
Workspace Git connection: connect any Premium/Fabric workspace to an Azure DevOps or GitHub repository branch. Semantic model files (.bim / TMDL format) and report definition (report.json) are committed automatically on publish. Pull from Git updates the workspace.
Branch-based development: developers work on feature branches; PRs merge to main; workspace syncs from main. Standard GitOps practices now apply to BI content.
Deployment pipelines: Dev → Test → Prod promotion with per-stage data source override rules — the 'release' mechanism complementing Git branching.
Fabric workspace: Fabric items (notebooks, lakehouses, semantic models) all stored as files in OneLake and committable to Git — unified version control across the data platform.""",
  g="""No native Git integration — revision history only; third-party tooling for full Git.
Revision history: Tableau Server/Cloud stores the last N published versions of each workbook and data source (N configurable by admin). Authors can revert to any previous version from the web UI. Not version control — no branching, no diff, no PR review.
Content Migration Tool (CMT): migrate workbooks/data sources between projects, sites, or servers — used as a manual promotion workflow (dev site → prod site).
Third-party Git: Broadcom Catalyst Tableau Extension, ZAP BI, community scripts commit .twb/.twbx (XML files) to Git. Requires process discipline; not platform-native.
Limitation: .twb/.twbx files are XML — technically diff-able but Tableau Desktop does not show diffs natively; reviewing changes requires external diff tooling.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/developer/projects/projects-git
https://learn.microsoft.com/en-us/power-bi/developer/projects/projects-overview
https://learn.microsoft.com/en-us/power-bi/create-reports/deployment-pipelines-overview""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/revision_history_admin.htm
https://help.tableau.com/current/server/en-us/cmt-intro.htm
https://github.com/tableau/tableau-migration-sdk (Tableau Migration SDK — open source)""",
  j="""Power BI wins clearly on native version control.
Power BI's native Git integration (GA 2024) brings real software engineering practices to BI — branching, PRs, merge reviews, and commit history for semantic models and reports. This enables multiple developers to work simultaneously without content conflicts.
Tableau's revision history is a point-in-time rollback tool, not version control. Native Git for Tableau requires third-party tools and manual process enforcement — not a platform guarantee.
For GTF's multi-developer BI team (brand analytics, franchise performance, catering modules): Git-based branch isolation prevents accidental overwrites and enables code-reviewed DAX changes.
Advantage: POWER BI — native Git integration + deployment pipelines; Tableau relies on revision history only.""",
  k="COMPLETED"
)

# ── S.No 183: DirectQuery mode / live connection performance ──
u(183,
  f="""DirectQuery and Live Connection are mature and central to the Power BI architecture.
DirectQuery: available for 100+ connector types. Every visual interaction generates a SQL/MDX query sent to the source — no data cached locally. Appropriate when data freshness or volume precludes Import.
Live Connection: to SSAS Tabular, AAS, Power BI semantic models, and Fabric semantic models — the model lives at the source; full DAX is supported (unlike some DirectQuery limitations).
Composite models: mix DirectQuery + Import sources in one model — tables from different sources join in the model layer. Unique Power BI capability. DirectQuery on Power BI semantic models: query another certified semantic model live without copying.
Performance: governed by the source engine. For AtScale: autonomous aggregates deliver near-Import speed for most patterns. Power BI Performance Analyzer + DAX Studio expose query timings.""",
  g="""Live Connection is Tableau's primary mode for semantic layer consumption.
Live Connection: Tableau queries the source at view time. Supported for all major relational, cloud DW, and semantic layer sources. No extract required for AtScale — connect live, Tableau sends SQL, AtScale responds.
No concept of 'Import mode' for a semantic layer — Tableau does not cache semantic model data locally (Hyper extracts are data snapshots, not semantic model caches).
No Composite model equivalent: Tableau can join tables within one data source or blend across two data sources — but blending across a live AtScale connection and a local Excel file requires a separate blend (less seamless than Power BI's Composite model).
Performance: governed by AtScale for live connections. Performance Recording + query logs expose query timings per sheet.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-use-directquery
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-composite-models
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-directquery-datasets-azure-analysis-services""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/live_connection.htm
https://help.tableau.com/current/pro/desktop/en-us/extracting_data.htm
https://help.tableau.com/current/pro/desktop/en-us/perf_record_create_desktop.htm""",
  j="""Power BI edge for connection mode flexibility.
Both tools connect live to AtScale and benefit equally from AtScale's autonomous aggregates — query performance on the GTF production workload will be comparable. Power BI's Composite model (mixing live AtScale + local Excel or SQL tables) is a unique differentiator for hybrid data scenarios Tableau cannot match natively.
Power BI's 100+ DirectQuery connectors also give a broader connectivity story for non-AtScale sources.
Advantage: POWER BI for connection mode flexibility (Composite model + Live Connection + DirectQuery mix); TIE for AtScale live query performance.""",
  k="BLOCKER"
)

# ── S.No 184: Configurable auto-refresh intervals ──
u(184,
  f="""Most sophisticated near-real-time refresh capability in the comparison.
Automatic page refresh (DirectQuery / Live): configure a report page to auto-query the source every N seconds/minutes — as low as 1 second in Premium. Fixed interval or change detection trigger.
Streaming datasets: push real-time data to a Power BI streaming dataset via REST API; tiles on dashboards update within seconds. No query polling needed.
Hybrid table (Premium): a regular Import table with an additional 'real-time' DirectQuery partition covering the most recent period — combines historical performance with real-time freshness.
Standard scheduled refresh: 8/day (Pro), 48/day (Premium), or unlimited via XMLA programmatic refresh.""",
  g="""Limited near-real-time options compared to Power BI.
Extract refresh: minimum 15-minute interval on Tableau Cloud (no sub-15-min extract refresh). On Tableau Server, intervals are admin-configurable but still not second-level.
Live connection: refreshes on user interaction (page load, filter change) — not time-based auto-refresh. Users see current data each time they interact, but the view does not auto-update while they are watching.
No streaming dataset equivalent: Tableau has no native push-streaming ingestion path for real-time tiles.
No automatic page refresh (clock-driven, sub-minute polling) equivalent.
Closest approach: reload browser tab or auto-refresh browser tab via external tooling (not a Tableau feature).""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-automatic-page-refresh
https://learn.microsoft.com/en-us/power-bi/connect-data/service-real-time-streaming
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-incremental-refresh#streaming-dataflows""",
  i="""Official documentation:
https://help.tableau.com/current/online/en-us/to_keep_data_fresh.htm
https://help.tableau.com/current/server/en-us/schedule_manage_extract.htm
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_jobs_tasks_and_schedules.htm""",
  j="""Power BI wins for near-real-time dashboard scenarios.
Power BI's automatic page refresh (down to 1 second in Premium) and streaming datasets support operational dashboards where data changes frequently — franchise order count, live catering board status. Tableau has no equivalent; live connections refresh only on user interaction.
For GTF's typical franchise reporting (daily/weekly financial metrics), neither tool's real-time capability is a differentiator — scheduled refresh is adequate.
If GTF requires live kitchen/catering order monitoring (truly real-time), Power BI is the only viable option from this comparison.
Advantage: POWER BI for near-real-time and sub-minute auto-refresh; TIE for standard daily/weekly scheduled data.""",
  k="COMPLETED"
)

# ── S.No 185: Incremental refresh ──
u(185,
  f="""Most sophisticated incremental refresh in the BI market.
Policy-based incremental refresh: define a historical window (e.g., 2 years frozen) and a rolling refresh window (e.g., last 10 days re-processed). Power BI automatically partitions the table and refreshes only the rolling window. Zero partition management code needed — defined via UI in Desktop with RangeStart/RangeEnd parameters.
Near-real-time hybrid table (Premium): adds a DirectQuery partition covering the most recent minutes/hours on top of the historical Import partition — same table combines historical performance with live freshness.
XMLA partition management: advanced pattern using Tabular Editor to programmatically add/delete/merge partitions — enables custom partition strategies beyond the UI-defined policy.""",
  g="""Incremental refresh is available but less sophisticated than Power BI.
Append-based incremental refresh: Tableau processes only rows newer than the last refresh timestamp (keyed on a date/integer column). Full refresh replaces all rows; append adds only new rows.
No partition management: Tableau does not partition the extract — the incremental append is applied to the monolithic .hyper file. No frozen historical partition equivalent.
No near-real-time hybrid table: cannot mix a historical extract partition with a live DirectQuery partition in one table.
Minimum increment: a single timestamp-keyed column; no support for multiple partition keys.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-incremental-refresh
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-incremental-refresh#hybrid-tables-with-real-time-data
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-incremental-refresh#xmla-endpoint""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/extracting_data.htm#refresh-extracts
https://help.tableau.com/current/online/en-us/to_keep_data_fresh.htm
https://help.tableau.com/current/server/en-us/schedule_manage_extract.htm""",
  j="""Power BI wins on incremental refresh sophistication.
Power BI's policy-based partition management, hybrid real-time table, and XMLA partition control are significantly more advanced than Tableau's append-based approach. For GTF's large datasets (118M row synthetic dataset + production scale), Power BI's partition strategy means full refreshes never occur — only the recent window is processed.
Tableau's incremental append is adequate for smaller, append-only datasets but does not scale to complex partition management requirements.
For GTF production (multi-year historical catering/sales data + daily new data): Power BI incremental refresh is the more operationally reliable solution.
Advantage: POWER BI — policy-based partition management + hybrid real-time table vs Tableau's append-only incremental.""",
  k="COMPLETED"
)

# ── S.No 195: Time zone management ──
u(195,
  f="""Time zone management is capacity-level in Power BI — limited per-workspace granularity.
Capacity-level time zone: in the Power BI Admin Portal, capacity admins can set the time zone for the entire Premium/Fabric capacity. All scheduled refreshes on that capacity use this time zone.
Scheduled refresh UI: users select their desired refresh time in local time — Power BI converts to UTC for scheduling; displayed back in local time.
Report subscriptions: sender selects delivery time; recipients receive based on the subscription schedule, not their own time zone.
Limitation: no per-workspace or per-report time zone override — all workspaces on the same capacity share the same time zone setting. For global franchises in multiple time zones, this can cause refresh/delivery timing mismatches.""",
  g="""More granular time zone control — per-site and per-schedule configurable.
Per-site time zone: Tableau Server and Tableau Cloud allow setting a time zone per site. All extract refreshes and subscriptions on the site respect this setting.
Per-schedule time zone: Tableau Server admin can configure individual refresh schedules in specific time zones — enabling a UK extract to refresh at 2:00 AM UK time while a US extract refreshes at 2:00 AM ET on the same server.
Datetime data handling: Tableau supports 'Extract timezone' for timezone-aware datetime columns — stores timestamps with timezone offset, converts to the viewer's local time zone at render time.
Subscription delivery: configurable per subscription for both time and time zone.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/admin/service-admin-portal-capacity-settings
https://learn.microsoft.com/en-us/power-bi/connect-data/refresh-scheduled-refresh
https://learn.microsoft.com/en-us/power-bi/collaborate-share/end-user-subscribe""",
  i="""Official documentation:
https://help.tableau.com/current/online/en-us/to_site_general.htm
https://help.tableau.com/current/server/en-us/schedule_add.htm
https://help.tableau.com/current/pro/desktop/en-us/dates_timezones.htm""",
  j="""Advantage Tableau for multi-timezone franchise operations.
Tableau's per-site and per-schedule time zone control is meaningfully more granular than Power BI's capacity-level setting. For GTF with franchise operators across US time zones (ET, CT, MT, PT), Tableau can schedule each region's nightly report delivery in their local business hours without affecting other regions.
Power BI's capacity-level time zone requires all workspaces to share the same setting — a meaningful operational constraint for a multi-timezone franchise business.
Advantage: TABLEAU — per-site and per-schedule time zone granularity; Power BI limited to capacity-level time zone.""",
  k="COMPLETED"
)

wb.save(FILEPATH)
print("Batch 8 saved: S.No 171, 183, 184, 185, 195")
