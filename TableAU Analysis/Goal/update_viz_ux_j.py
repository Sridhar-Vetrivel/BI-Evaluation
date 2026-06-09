import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

# Column J = col 10 (Findings/Conclusion)
# Only writing J for rows where J is currently empty. Rows 15, 16, 17 already filled — skip.

# ─── S.No 11 | Row 12 | Standard chart types ───────────────────────────────
ws.cell(12, 10).value = (
    'No decisive winner for GTF. '
    'Both tools cover bar, line, scatter, pie, waterfall, and funnel natively. '
    'Tableau has a slight edge: bullet chart is a first-class native type, '
    'while Power BI requires a third-party AppSource visual for bullet charts. '
    'For GTF\'s core reporting needs, both tools are equivalent on standard chart coverage.'
)

# ─── S.No 12 | Row 13 | Advanced/specialized chart types ───────────────────
ws.cell(13, 10).value = (
    'No decisive winner. '
    'Both tools handle treemaps, heatmaps, and small multiples natively. '
    'Sankey and chord diagrams require workarounds in both tools — '
    'AppSource custom visuals in Power BI and data densification templates in Tableau. '
    'Neither tool has a meaningful native advantage for the advanced chart types listed. '
    'Not a differentiating factor for GTF\'s evaluation.'
)

# ─── S.No 13 | Row 14 | Custom visualization support ───────────────────────
ws.cell(14, 10).value = (
    'No decisive winner. '
    'Power BI supports D3.js directly via the Custom Visual SDK (TypeScript-based) '
    'and offers a large AppSource marketplace of certified visuals. '
    'Tableau supports custom visuals via Dashboard Extensions API (JavaScript/D3) '
    'and Viz Extensions (newer, mark-level customization). '
    'Neither tool supports native Vega-Lite. '
    'Both are capable for custom chart needs — not a differentiating factor.'
)

# ─── S.No 17 | Row 18 | Dashboard layout flexibility ───────────────────────
ws.cell(18, 10).value = (
    'Tableau wins for dashboard layout and visual design quality. '
    'Tableau\'s floating layout mode provides pixel-exact X/Y/W/H positioning via the Layout panel, '
    'enabling branded header bands, overlapping text on maps, and polished executive scorecards '
    'that are impractical to author in Power BI. '
    'Device Designer (desktop/tablet/phone) is maintained in one workbook — '
    'Power BI requires a separate mobile canvas that must be updated independently. '
    'For GTF\'s executive brand scorecards and iPad-facing FBC dashboards, '
    'Tableau delivers a significantly higher visual quality output for the same authoring effort.'
)

# ─── S.No 18 | Row 19 | Theming and branding consistency ───────────────────
ws.cell(19, 10).value = (
    'Power BI wins for enterprise-wide branding. '
    'Power BI\'s JSON theme files define color palette, fonts, and default visual styles '
    'and can be published org-wide so all reports inherit GTF\'s brand automatically. '
    'Tableau relies on template workbooks and style guides — branding is workbook-level, '
    'not centrally deployable. '
    'For a franchise brand like GTF with consistent brand standards across 20+ brands, '
    'Power BI\'s centralized theming is the stronger approach.'
)

# ─── S.No 19 | Row 20 | Animation and transitions for storytelling ─────────
ws.cell(20, 10).value = (
    'Tableau wins for storytelling and animation. '
    'Tableau\'s Story Points feature provides a dedicated narrative authoring surface — '
    'a sequence of slides with captions and annotations — with no equivalent in Power BI. '
    'Tableau\'s Pages Shelf and Mark Animations (since 2020.1) animate transitions between states. '
    'Power BI offers bookmarks for guided navigation and a Play Axis for time-animated scatter/bubble charts, '
    'but lacks a structured storytelling surface. '
    'For GTF QBR presentations or brand performance storytelling, Tableau has a clear advantage.'
)

# ─── S.No 20 | Row 21 | Accessibility ──────────────────────────────────────
ws.cell(21, 10).value = (
    'No decisive winner — both tools meet enterprise accessibility standards. '
    'Both support screen readers (Narrator/NVDA/JAWS for Power BI, server-side for Tableau), '
    'keyboard navigation, and WCAG 2.1 AA conformance with published documentation. '
    'Power BI has a slight edge: built-in Accessibility Insights checker and Microsoft\'s broad '
    'accessibility investment give it more consistent coverage across native visuals. '
    'Not a differentiating factor for GTF\'s selection decision.'
)

wb.save(FILEPATH)
print('Saved.')

# Verify
wb2 = openpyxl.load_workbook(FILEPATH)
ws2 = wb2['Evaluation Matrix - 1']
for rnum, sno in [(12,'11'), (13,'12'), (14,'13'), (18,'17'), (19,'18'), (20,'19'), (21,'20')]:
    j = ws2.cell(rnum, 10).value or ''
    k = ws2.cell(rnum, 11).value or ''
    print(f"S.No {sno} (row {rnum}) | K={k} | J={j[:100]}")
