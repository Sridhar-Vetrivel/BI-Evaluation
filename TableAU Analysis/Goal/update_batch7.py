import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 152: Scheduled email subscriptions ──
u(152,
  f="""Most capable subscription system in the comparison — direct GTF requirement.
Standard subscriptions: email delivery of a report page as PDF/PPTX/PNG on a schedule (hourly, daily, weekly, monthly). Can send 'only when data changes.' Supports multiple recipients per subscription.
Dynamic per-recipient subscriptions (Premium/Fabric): one subscription + a mapping semantic model (columns: email, filter value, optional subject) → Power BI reads the mapping at send time and delivers a differently filtered snapshot to each recipient. Designed specifically for franchise/territory bursting.
Paginated report subscriptions (Premium): pixel-perfect PDF or Excel attachment per schedule. Parameters can be per-recipient rows in the mapping dataset — true data-driven subscription.
Recipient cap: 1,000 rows in the dynamic mapping dataset per subscription.""",
  g="""Basic subscriptions — one content set per schedule; no native per-recipient filtering.
View/workbook subscriptions: email delivery of a view or workbook as PDF or PNG on a configured schedule. Recipients all receive the same content (with their individual RLS applied if configured).
'Only when data changes': available on Tableau Cloud (checks at send time; skips if view unchanged).
Workbook-level vs view-level: subscribe to a full workbook (all sheets) or a specific view.
Per-recipient bursting: not a native feature — requires REST API scripting (loop over recipients, apply URL filters, send email) or third-party tools (ChristianSteven, Rollstack, Informer).
Data-Driven Alerts: threshold-based notification (not a report snapshot) — complementary to subscriptions.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-subscribe
https://learn.microsoft.com/en-us/power-bi/collaborate-share/end-user-subscribe#subscribe-with-dynamic-per-recipient-subscriptions
https://learn.microsoft.com/en-us/power-bi/paginated-reports/subscriptions/paginated-reports-subscriptions""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/subscribe_user.htm
https://help.tableau.com/current/server/en-us/subscribe.htm
https://help.tableau.com/current/server/en-us/data_alerts.htm""",
  j="""Power BI wins decisively — this is a core GTF client must-have.
The 'Robots' requirement (automated per-franchise filtered email delivery) is a native Power BI Premium feature via dynamic subscriptions. Configure once with a franchise-email mapping dataset; Power BI handles the rest with no ongoing engineering.
Tableau has no equivalent native capability — per-recipient bursting requires building a REST API loop or purchasing a third-party bursting tool. For GTF with potentially hundreds of franchise recipients, that is significant ongoing maintenance.
Advantage: POWER BI — dynamic per-recipient subscriptions are a native platform feature (Premium); Tableau requires custom engineering or third-party software.""",
  k="COMPLETED"
)

# ── S.No 153: Personalized filter state persistence ──
u(153,
  f="""Personal bookmarks for user-saved report state.
Personal bookmarks: up to 20 personal bookmarks per report per user, saved in the Power BI Service. Captures filter state, slicer selections, sort order, and visible/hidden objects — full state snapshot.
'Persist filters' setting (configurable per report by author): when enabled, Power BI automatically restores a user's last filter state when they revisit the report. Users can reset to author-defined defaults via 'Reset to default.'
Limitation: 20-bookmark cap per report per user is restrictive for power users who manage many filter states (territory managers, FBCs).""",
  g="""Custom Views — more flexible, no cap, URL-shareable.
Custom Views: users save their current filter/sort/parameter state as a named view on any published report in Tableau Server/Cloud. Set as 'My Default View' — Tableau loads that state automatically on next visit.
No limit on Custom Views per user per workbook.
Shareable: a Custom View can be shared via URL — recipient opens the workbook in that exact saved state. Useful for escalation paths ('see the issue I flagged').
URL parameters: filter state is embedded in the view URL — bookmarkable in the browser natively without using the Custom View feature.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-bookmarks
https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-bookmarks
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-report-filter""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/customview.htm
https://help.tableau.com/current/server/en-us/custom_views.htm
https://help.tableau.com/current/server/en-us/customview_manage.htm""",
  j="""Slight Tableau advantage for consumer state persistence.
Tableau Custom Views have no bookmark cap, are URL-shareable, and function as a default view automatically. Power BI personal bookmarks are capped at 20 — a real limitation for territory managers who track many franchise combinations.
Power BI's 'Persist filters' automatic state restoration is simpler for casual users; Tableau requires explicit Custom View creation.
For GTF's territory managers (complex multi-filter states per region/brand): Tableau's uncapped, shareable Custom Views are more operationally flexible.
Advantage: TABLEAU — uncapped custom views with URL sharing; Power BI's 20-bookmark limit is constraining for power users.""",
  k="COMPLETED"
)

# ── S.No 155: Print and export options for consumers ──
u(155,
  f="""Comprehensive consumer-facing export — especially strong with paginated reports.
Standard reports: consumers can Export to PDF (full report), PPTX (PowerPoint), or PNG. Underlying data: CSV or XLSX (30K / 150K row caps respectively). 'Download as PDF' respects report theme and visual layout.
Paginated reports (Premium): consumer exports to PDF (pixel-perfect, headers/footers), Excel (formatted with formulas preserved), CSV, Word, PowerPoint, or image. Accessible PDF (tagged) for screen-reader compliance.
Print: standard browser print dialog from Power BI Service for quick printing.
Author-controlled export: report author can disable export options per report if data sensitivity requires.""",
  g="""Good consumer export options for interactive reports.
View/Dashboard: Download PDF (view-level or full dashboard), PNG (image), CSV (underlying data as crosstab), Excel (crosstab, removes formatting).
Workbook: Download TWBX, PDF (multi-sheet), or Crosstab (single sheet to Excel).
No native PowerPoint export — PDF is the closest; conversion to PPTX requires external tooling.
No Word/docx or accessible PDF (tagged) export.
Print: browser print from the Service for quick printing.
Author control: permissions model controls whether Viewer role can download data (Download capability can be revoked per project).""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-export
https://learn.microsoft.com/en-us/power-bi/paginated-reports/report-builder/export-reports
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-print""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/save_export.htm
https://help.tableau.com/current/pro/desktop/en-us/export_printpdf.htm
https://help.tableau.com/current/server/en-us/permissions_capabilities.htm""",
  j="""Power BI wins on export format richness.
Power BI paginated reports deliver PDF, Excel (formatted), Word, PowerPoint, and accessible PDF from a single report template — covers GTF's catering detail report, executive summary PPTX, and Excel data extract use cases in one tool.
Tableau's consumer export covers PDF, PNG, CSV, Excel — adequate for dashboard snapshots but lacks native PowerPoint and Word formats, and has no accessible PDF for compliance.
For GTF franchise operators who print/export catering order details: Power BI paginated reports are purpose-built; Tableau requires workarounds.
Advantage: POWER BI — broader format set (PPTX, Word, accessible PDF) via paginated reports.""",
  k="COMPLETED"
)

# ── S.No 156: Bookmark and personal view saving ──
u(156,
  f="""Two-tier bookmark system — author and personal bookmarks coexist.
Author bookmarks: designed into the report by the author; capture full page state (filters, slicers, sort, visibility of every object). Presented as a navigation list or triggered by buttons/images. Support interactive UX patterns (show/hide panels, toggle views, guided storytelling).
Personal bookmarks (consumer): up to 20 per report per user; override the author's published state with the user's own filter/sort preferences. Show alongside or instead of author bookmarks.
Selection pane: author uses this to precisely control which objects each bookmark shows/hides — enables sophisticated modal-popup UX within a report page.
Limitation: 20 personal bookmark cap; author must use 'Update' to re-capture state after any layout change.""",
  g="""Custom Views for consumer state saving; no author bookmark equivalent for visibility control.
Custom Views: consumer-created named saves of filter/sort/parameter state; uncapped; set as default; URL-shareable. Functionally equivalent to Power BI personal bookmarks with no cap.
Author-created navigation: Dashboard buttons (Go to Sheet), Show/Hide button containers (manually triggered), Dynamic Zone Visibility (programmatically show/hide sheet zones via a boolean field), Sheet swapping via parameter — assembles bookmark-like UX from multiple features. More complex to build than Power BI's bookmark + Selection pane workflow.
No single 'author bookmark' that captures object visibility + filters in one click — must orchestrate Dynamic Zone Visibility + parameters separately.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-bookmarks
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-bookmarks
https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-selection-pane""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/customview.htm
https://help.tableau.com/current/pro/desktop/en-us/actions_filter.htm
https://help.tableau.com/current/pro/desktop/en-us/zone_visibility.htm""",
  j="""Power BI wins for author-designed interactive UX; Tableau wins for consumer flexibility.
Power BI bookmarks (author-created) are a first-class feature enabling sophisticated one-click UX patterns — show/hide panels, toggle metric views, guided navigation — all built in 10 minutes with Bookmark + Selection pane. No Tableau equivalent matches this simplicity for authors.
Tableau Custom Views are more flexible for consumers (no cap, URL-shareable) and better for users who save many filter states. But assembling a comparable author-UX requires coordinating Dynamic Zone Visibility + parameters + actions — significantly more effort.
Advantage: POWER BI for author-designed bookmark UX; TABLEAU for consumer view flexibility (no cap).""",
  k="COMPLETED"
)

# ── S.No 157: Paginated report support ──
u(157,
  f="""Clear Power BI strength — and a GTF client must-have already identified.
Paginated Reports (Power BI Report Builder): RDL/SSRS-based engine purpose-built for pixel-perfect, multi-page, printable, large-row operational reports. True pagination: rows flow across pages automatically with repeating column headers, group headers, and page footers.
Unlimited rows: no row-count ceiling — the tablix renders and exports as many rows as the query returns. Tested at millions of rows.
Export: PDF (pixel-perfect), Excel (formatted), CSV, Word, PowerPoint, image, accessible PDF.
Data source: can point directly at AtScale (via XMLA or SQL), a Power BI semantic model, Databricks, or any other ODBC/OLE DB source.
License gate: requires Premium Per User (PPU) or Fabric/Premium capacity for consumers to view paginated reports. No Pro-tier support.""",
  g="""Weakest area — no SSRS-style paginated report engine exists in Tableau.
Tableau is built for visualization, not large tabular operational reporting. Text tables can display detail data but render slowly at high row counts and have no true pagination.
Workaround: INDEX() calculated field + 'Page Number' parameter to simulate pagination by filtering to N rows per 'page' — fragile, manually maintained, not print-ready.
Export: Crosstab to Excel (up to Excel's 1M row limit) and CSV for data extraction. PDF export of the view — not a pixel-perfect paginated document.
Third-party: some teams use Logi Analytics, SSRS, or Crystal Reports alongside Tableau for paginated operational reports. Adds another tool to the stack.
For the GTF catering detail report (franchisee-level line-item detail, printable, hundreds of rows per page): Tableau has no adequate native solution.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/paginated-reports/paginated-reports-report-builder-power-bi
https://learn.microsoft.com/en-us/power-bi/paginated-reports/report-design/report-parameters-paginated-reports
https://learn.microsoft.com/en-us/power-bi/paginated-reports/report-builder/export-reports""",
  i="""Official documentation (workarounds — no native paginated report feature):
https://help.tableau.com/current/pro/desktop/en-us/calculations_calculatedfields_lod.htm
https://help.tableau.com/current/pro/desktop/en-us/export_printpdf.htm
https://community.tableau.com/s/question/0D54T00000C5WobSAF/paginated-reports (community workaround thread)""",
  j="""Power BI wins decisively — this is GTF's most critical differentiator.
Power BI Paginated Reports directly address the catering detail report requirement: unlimited rows, true pagination, print-ready PDF with headers/footers, and pixel-perfect layout. This is the SSRS capability GTF currently relies on in the old SSAS architecture.
Tableau has no comparable feature — the closest approximation (INDEX() pagination workaround + Crosstab export) does not meet a professional operational report standard.
If the catering detail report must be produced from the BI tool (not a separate reporting system), Power BI is the only viable choice.
Advantage: POWER BI — decisive, platform-level capability; Tableau gap is fundamental (not a configuration issue).""",
  k="COMPLETED"
)

wb.save(FILEPATH)
print("Batch 7 saved: S.No 152, 153, 155, 156, 157")
