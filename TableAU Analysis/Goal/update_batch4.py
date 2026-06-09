import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 71: Deployment modes ──
u(71,
  f="""Three deployment modes available.
SaaS (Power BI Service / Fabric): Microsoft-managed cloud, Azure-hosted, automatic updates, SLA-backed. Most GTF-relevant path — no infrastructure to manage.
On-premises (Power BI Report Server): annual license, Windows Server, fully self-hosted. Suitable for air-gapped or data-residency-constrained environments. Receives quarterly feature updates vs monthly cloud.
Hybrid: Power BI Service (cloud) + On-Premises Data Gateway for data connectivity to on-prem/VNet sources. Most common enterprise pattern — cloud rendering, on-prem data.
Fabric: extends the SaaS model with a full analytics platform (data engineering, warehousing, science, BI) on Azure. Private links and VNet data gateways for network isolation without on-prem infrastructure.""",
  g="""Three deployment modes available.
SaaS (Tableau Cloud): Salesforce-managed, hosted on AWS (multiple regions), automatic updates, 99.9% SLA. Recommended for new deployments.
Self-hosted (Tableau Server): full control over infrastructure, OS, upgrades, backups. Deploy on-prem (Windows/Linux), Azure IaaS, AWS EC2, or GCP Compute. Requires sizing, patching, HA planning.
Hybrid: Tableau Cloud + Tableau Bridge for private-network data sources. Bridge runs on-prem; Cloud renders and serves dashboards.
No equivalent to Fabric's unified analytics platform — Tableau is BI-only; data engineering requires separate tooling (Databricks, dbt, etc.).""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/report-server/get-started
https://learn.microsoft.com/en-us/fabric/fundamentals/fabric-private-links-overview
https://powerbi.microsoft.com/en-us/power-bi-service-sla/""",
  i="""Official documentation:
https://help.tableau.com/current/online/en-us/to_get_started_online.htm
https://help.tableau.com/current/server/en-us/install.htm
https://help.tableau.com/current/online/en-us/to_bridge_install.htm""",
  j="""Tie — both offer SaaS, self-hosted, and hybrid options.
Power BI SaaS (Fabric) is Azure-native — natural fit for GTF's Azure-first architecture with private links for VNet isolation. Tableau Cloud runs on AWS; connecting to GTF's Azure-hosted AtScale/Databricks adds cross-cloud network complexity.
For cloud-first GTF: Power BI/Fabric on Azure eliminates cross-cloud data egress concerns and simplifies network topology.
Tableau Server (self-hosted) gives maximum control but adds significant IT overhead.
Advantage: POWER BI for Azure-native cloud deployment; TIE for hybrid/self-hosted.""",
  k="COMPLETED"
)

# ── S.No 72: Cloud platform support ──
u(72,
  f="""Azure-native; multi-cloud data connectivity via connectors.
Power BI Service / Fabric runs exclusively on Microsoft Azure (global regions). Data sources: native connectors to Azure SQL, Synapse, Databricks, ADLS, Cosmos DB, and 150+ external sources including AWS Redshift, S3, GCP BigQuery.
Fabric: unified platform on Azure — compute, storage, BI in one Azure region/tenant.
On-prem or cross-cloud data: on-premises data gateway bridges cloud rendering to any network-accessible source (AWS RDS, GCP BigQuery, on-prem SQL Server).
Power BI Report Server: deployable on any Windows Server — bare metal, Hyper-V, Azure IaaS, AWS EC2, or GCP Compute.""",
  g="""Tableau Cloud hosted on AWS; Tableau Server deployable anywhere.
Tableau Cloud: hosted on AWS (East US, EU West, APAC, etc.) — data sources in Azure require cross-cloud connectivity (Tableau Bridge or public endpoints). Native connectors to Databricks, Snowflake, BigQuery, Redshift, Azure SQL, Synapse, ADLS.
Tableau Server: deployable on Azure IaaS (Azure VM), AWS EC2, GCP Compute, on-prem — full infrastructure control. When hosted on Azure VM, network latency to AtScale/Databricks is minimised.
No preferred-cloud dependency at connector level — Tableau supports all major cloud databases equally.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/connect-data/power-bi-data-sources
https://learn.microsoft.com/en-us/fabric/fundamentals/microsoft-fabric-overview
https://azure.microsoft.com/en-us/products/power-bi/""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/exampleconnections_overview.htm
https://help.tableau.com/current/online/en-us/to_site_storage.htm
https://www.tableau.com/cloud-hosting""",
  j="""Power BI wins for Azure-first organisations.
Power BI/Fabric is built on Azure — same network fabric as GTF's Databricks, ADF, and AtScale. No cross-cloud latency, no egress costs, private link available end-to-end.
Tableau Cloud on AWS introduces Azure-to-AWS network hops for every query against GTF's Azure-based AtScale — adds latency and potential egress charges. Mitigated by deploying Tableau Server on Azure IaaS instead, but that adds self-hosting overhead.
Advantage: POWER BI for Azure-native deployment; TIE if Tableau Server is deployed on Azure IaaS.""",
  k="COMPLETED"
)

# ── S.No 76: Admin portal capabilities ──
u(76,
  f="""Comprehensive — Fabric Admin Portal is one of the richest in the market.
Tenant settings: 200+ configurable governance toggles (who can publish, export, use Q&A, embed, etc.).
Workspace and capacity management: view all workspaces, reassign, monitor usage.
Usage metrics: per-workspace report and dataset usage (views, unique users, refresh success/failures).
Fabric Capacity Metrics app: real-time and 14-day historical CPU throttling, operation-level breakdown — essential for capacity right-sizing.
Scanner API (Admin REST): full tenant metadata scan — all workbooks, datasets, data sources, lineage — for governance tooling.
Audit logs: Power BI activity log exported to Azure Monitor / Log Analytics for SIEM integration.""",
  g="""Capable — Tableau Admin areas are well-developed for enterprise governance.
Server Admin / Site Admin: usage stats, user activity, workbook performance, background task queue, storage usage per site.
Tableau Cloud Management: capacity (VizQL sessions), job scheduler, site security policies, auth configuration.
Admin Views (built-in): prebuilt dashboards for space used, traffic to views, traffic to data sources, background tasks for extracts, background tasks for flows.
REST API + Metadata API (GraphQL): programmatic governance — query all content, lineage, certification status, usage counts. More developer-friendly for custom governance tooling than Power BI's Scanner API.
Activity Log (Cloud, 2023+): detailed security-relevant event log for compliance audit.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/admin/service-admin-portal
https://learn.microsoft.com/en-us/fabric/enterprise/metrics-app
https://learn.microsoft.com/en-us/power-bi/admin/service-admin-auditing""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/adminview.htm
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_metadata.htm
https://help.tableau.com/current/online/en-us/activity_log_overview.htm""",
  j="""Tie — both have enterprise-grade admin portals.
Power BI: Fabric Capacity Metrics App is exceptional for capacity planning and throttling diagnosis — more detailed than Tableau's session-level monitoring. 200+ tenant settings give fine governance control.
Tableau: REST + Metadata GraphQL APIs are more developer-friendly for building custom governance dashboards and lineage tools. Admin Views are pre-built and useful out-of-the-box.
For GTF (Fabric capacity purchase expected): Power BI Capacity Metrics App is a valuable day-to-day operational tool.
Advantage: TIE — Power BI better for capacity metrics; Tableau better for programmatic governance via API.""",
  k="COMPLETED"
)

# ── S.No 78: Backup and restore ──
u(78,
  f="""Backup and restore is architecture-dependent — no single command for full workspace.
Semantic models: XMLA endpoint (Premium) allows Tabular Editor / ALM Toolkit to export model as .bim (Tabular Model BIM file) — a full schema + measure backup. Restore by re-deploying via XMLA.
Git integration (Fabric/Premium): workspace connected to Git branch — every publish is a commit; rollback = git revert + sync. Most robust version-control-based restore.
Desktop .pbix export: available for Import-mode reports (Live Connection reports cannot export with model). Limited — not all Fabric items exportable as .pbix.
Fabric items: metadata stored in Git; data (Lakehouse, Warehouse) backed up separately per Azure storage policies.""",
  g="""Tableau Server provides an administrator-initiated, explicit full-backup command.
tabcmd: `tabcmd backup` — creates a full site backup (.tsbak) including all workbooks, data sources, users, schedules, permissions. Restores with `tabcmd restore`. Scriptable for scheduled nightly backups.
Tableau Cloud: Salesforce manages infrastructure backup; site content exportable via REST API (workbooks, data sources) or Content Migration Tool. No admin-initiated .tsbak equivalent on Cloud.
Revision history: Tableau Server/Cloud keeps last N versions of each workbook (configurable) — lightweight per-item rollback without full backup/restore.
Content Migration Tool: move content between sites/servers for DR or environment promotion.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/developer/projects/projects-git
https://learn.microsoft.com/en-us/analysis-services/tom/introduction-to-the-tabular-object-model-tom-in-analysis-services-amo
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-enhanced-dataset-metadata""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/db_backup.htm
https://help.tableau.com/current/server/en-us/db_restore.htm
https://help.tableau.com/current/server/en-us/revision_history_admin.htm
https://help.tableau.com/current/server/en-us/cmt-intro.htm""",
  j="""Advantage Tableau for self-managed server backup; Power BI stronger for Git-based CI/CD restore.
Tableau Server's `tabcmd backup` is a single explicit command that captures a full site snapshot — IT-friendly, scriptable, and testable. Tableau Cloud has no equivalent admin backup (Salesforce manages infrastructure).
Power BI has no single-command full-workspace backup; recovery relies on Git (excellent if configured) or .pbix exports (incomplete for Fabric items). Without Git, restoring a Power BI workspace after accidental deletion is complex.
For GTF: if using Fabric (SaaS), Git integration must be configured deliberately as the backup strategy.
Advantage: TABLEAU (Server) for explicit backup/restore; POWER BI with Git for version-controlled content recovery.""",
  k="COMPLETED"
)

wb.save(FILEPATH)
print("Batch 4 saved: S.No 71, 72, 76, 78")
