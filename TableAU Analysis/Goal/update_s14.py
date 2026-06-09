import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

# S.No 14 = spreadsheet row 15
# F=6 PowerBI | G=7 Tableau | J=10 Findings/Conclusion | K=11 Status

ROW = 15

ws.cell(ROW, 6).value = (
    'Best-in-class for matrix conditional formatting. '
    'Background color, data bars, and icons applied independently per value column '
    'via native Format panel — no DAX or calculated fields needed for basic rules. '
    'Each value column (Variance $, Actual, Variance %) gets its own independent formatting. '
    'Field Parameters enable measure toggle natively with no formula wiring. '
    'Authoring time for all 4 capabilities: ~10 min.'
)

ws.cell(ROW, 7).value = (
    'Capable but more manual setup required. '
    'Background color via calculated color field — applies row-level, not independently per column. '
    'No native data bars inside text table cells — bar chart workaround floated alongside is the closest approximation. '
    'Icons via Unicode calculated field; icon and number share the same cell color, cannot be independently styled. '
    'Measure toggle requires a Parameter plus a CASE calculated field. '
    'Authoring time for equivalent output: ~25-30 min.'
)

ws.cell(ROW, 10).value = (
    'Power BI wins clearly. '
    'Power BI applies background color, data bars, and icons per-column per-cell '
    'through a native panel in approximately 10 minutes with no formula work. '
    'Tableau achieves comparable visual output but requires calculated fields for color (row-level only), '
    'a separate worksheet floated alongside for data bars, '
    'and Unicode characters for icons that cannot be independently colored from the number in the same cell. '
    'Tableau takes 2-3x the authoring time for the same result. '
    'For GTF P&L and catering variance reports (client requirement #6d), Power BI is the significantly stronger choice.'
)

ws.cell(ROW, 11).value = 'COMPLETED'

wb.save(FILEPATH)
print('Saved.')

wb2 = openpyxl.load_workbook(FILEPATH)
ws2 = wb2['Evaluation Matrix - 1']
r = ws2[ROW]
print(f'\nS.No 14 verified:')
print(f'  F: {str(r[5].value)[:100]}')
print(f'  G: {str(r[6].value)[:100]}')
print(f'  J: {str(r[9].value)[:120]}')
print(f'  K: {r[10].value}')
