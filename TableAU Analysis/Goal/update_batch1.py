import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

# col indices (1-based): F=6 PowerBI, G=7 Tableau, H=8 PBI Proof, I=9 Tab Proof, J=10 Findings, K=11 Status
# spreadsheet row = S.No + 1

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 7: Live connection vs import mode (only needs Findings; Status stays BLOCKER) ──
u(7,
  j="""Live vs Import is the foundational data-freshness decision.
Power BI offers Import (VertiPaq in-memory, fastest queries), DirectQuery (live pushdown), and Composite/Dual (mix both per table) — the most flexible trio in the market. Tableau offers Live Connection or Extract (Hyper engine), with no composite-model equivalent.
For the GTF AtScale architecture: both tools use Live Connection to AtScale; import/extract mode of AtScale data is impractical at franchise scale. The three-mode flexibility of Power BI is valuable for hybrid sources (e.g., mixing a live AtScale model with a local Excel budget table). Tableau does not support this without a separate data source.
Advantage: POWER BI — richer mode mix; Composite models enable hybrid live+import patterns Tableau cannot match natively."""
)

# ── S.No 9: Gateway (only needs Findings; Status already COMPLETED) ──
u(9,
  j="""Both tools have production-grade gateway/bridge solutions.
Power BI On-Premises Data Gateway (Standard): Windows-only, HA clusters up to 10 nodes, Kerberos SSO delegation, load balancing, Microsoft-managed monthly updates. Tableau Bridge: Windows-only, pool-based HA, live + extract over private networks — functionally equivalent but newer and less mature on clustering/monitoring tooling.
For GTF/AtScale: the gateway routes Power BI Service queries to the on-prem/VNet-hosted AtScale instance. Kerberos constrained delegation is required for per-user SSO — tested and documented by AtScale for Power BI. Tableau's AtScale Cloud Connector (OAuth SSO, no Bridge needed for Tableau Cloud) makes the connectivity story marginally simpler on the Tableau side.
Advantage: TIE — both are capable. Power BI gateway more mature on clustering/monitoring; Tableau AtScale Cloud Connector sidesteps Bridge for Cloud deployments."""
)

# ── S.No 21: RBAC ──
u(21,
  f="""Capable. Four fixed workspace roles: Admin (full control), Member (edit + publish app), Contributor (edit, no app publish), Viewer (read-only). Roles assigned to Entra users/groups.
App Audiences: one App can show different content tabs to different Entra groups — functional RBAC at the consumer layer.
Object Level Security (OLS): restricts tables/columns from appearing in reports; configured via Tabular Editor. Tenant Admin Settings: global toggles (e.g., who can publish to web, export data, use Q&A) managed in the Admin Portal.
Limitation: only 4 fixed roles per workspace — no custom-permission granularity at the workspace level. Governance hierarchy is flat (workspaces are peers, not nested).""",
  g="""Capable. Capability-based permission model at site/project/workbook/view/data-source level. Site roles (Creator, Explorer, ExplorerCanPublish, Viewer, SiteAdminCreator, ServerAdmin) set the ceiling; project-level capabilities further restrict or grant individual actions (View, Filter, Download, Web Edit, Overwrite, etc.).
Project Leader role: delegated admin over a project and all nested sub-projects — enables distributed governance without full Server/Site Admin access.
Permission inheritance: projects can be locked (enforce inheritance) or unlocked (per-item overrides). Finer-grained than Power BI's four fixed roles.
Limitation: more complex to set up and audit than Power BI's simpler role model; no native OLS equivalent for column/table hiding.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-roles-new-workspaces
https://learn.microsoft.com/en-us/power-bi/enterprise/service-admin-object-level-security
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-create-distribute-apps#create-and-manage-multiple-audiences""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/permissions.htm
https://help.tableau.com/current/server/en-us/permissions_capabilities.htm
https://help.tableau.com/current/server/en-us/projects.htm#lock""",
  j="""Tie with context-dependent advantage.
Power BI: simpler 4-role model, easier to administer at scale; OLS for column/table restriction is a genuine extra Tableau lacks. Entra group assignment is seamless for Microsoft orgs.
Tableau: more granular capability-based permissions + nested project hierarchy = better for complex, multi-team governance structures.
For GTF (franchise hierarchy, territory managers, corporate): Tableau's nested project structure maps more naturally to the org chart, but Power BI's OLS + Entra groups cover the GTF security requirements more simply.
Advantage: TIE — Power BI simpler; Tableau more granular.""",
  k="COMPLETED"
)

# ── S.No 22: RLS and CLS ──
u(22,
  f="""Best-in-class for enterprise RLS without add-on cost.
Dynamic RLS: DAX filter using USERPRINCIPALNAME() against a security/bridge table — scales to thousands of franchise users with one role definition. Static roles for fixed group mappings.
Object Level Security (OLS): restricts specific tables or columns — unavailable users see an error on access attempt. OLS configured via Tabular Editor (no native Desktop UI).
Test with 'View as role' or 'View as other user' in Power BI Service.
AtScale path: if RLS is enforced in AtScale (recommended), Power BI passes user identity via SSO/Kerberos delegation; no double-filter needed. RLS in the semantic model inherits by all connecting reports automatically.
Included in Pro license — no add-on required.""",
  g="""Capable but centralized RLS requires a paid add-on.
Four approaches: (1) User filters (manual UI, small lists only); (2) Calculated-field RLS with USERNAME()/ISMEMBEROF() as a data source filter; (3) Entitlement-table RLS (username → franchise value join); (4) Virtual Connections + Data Policies — centralized RLS defined once, enforced on all downstream content.
Data Policies (approach 4) require the Data Management add-on (separate license cost).
Column-level security: no clean CLS equivalent to Power BI's OLS. Can hide columns in published data source, but no error-on-access enforcement.
AtScale path: push RLS to AtScale's own policy engine; Tableau queries AtScale with user context. Avoid double-filtering.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/enterprise/service-admin-rls
https://learn.microsoft.com/en-us/power-bi/enterprise/service-admin-object-level-security
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-dynamic-m-query-parameters""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/rls_options_overview.htm
https://help.tableau.com/current/server/en-us/dm_vconn_overview.htm
https://help.tableau.com/current/server/en-us/data_policy.htm""",
  j="""Power BI wins for GTF use case.
Power BI dynamic RLS is included in Pro, highly tested, and works end-to-end with Entra SSO — franchise filtering via USERPRINCIPALNAME() against a franchise-user bridge table is a standard enterprise pattern. OLS adds column-level security Tableau cannot match natively.
Tableau's strongest centralized RLS (Virtual Connections + Data Policies) requires purchasing the Data Management add-on, adding cost for what Power BI delivers out-of-the-box.
For GTF with 750+ franchise users needing per-territory row filtering: Power BI is more cost-efficient and simpler to maintain.
Advantage: POWER BI — native RLS + OLS included in Pro; no additional license needed.""",
  k="COMPLETED"
)

# ── S.No 23: SSO / Azure AD ──
u(23,
  f="""Best-in-class for Azure AD / Entra ID shops.
Power BI Service is built on Microsoft Entra ID — SSO is intrinsic, not a bolt-on. Any user signed into Microsoft 365 is automatically authenticated to Power BI with no additional setup. MFA and Conditional Access policies (device compliance, location, session controls) apply to Power BI Service automatically via Entra.
SCIM-free: Entra groups are directly assignable to workspace roles, app audiences, and RLS roles — no provisioning pipeline required.
Service principals: Entra app registrations used for automation/embedded scenarios (no MFA prompt, no password expiry).
OIDC external MFA (2025): third-party MFA providers supported alongside Conditional Access.""",
  g="""Capable but requires explicit setup.
Tableau Cloud and Server both support SAML 2.0 and OpenID Connect (OIDC) with Entra ID as IdP, plus SCIM 2.0 for user/group provisioning. MFA is enforced via Entra Conditional Access (delegated — Tableau itself has no built-in MFA).
Tableau Cloud: use the 'Tableau Cloud' app from the Entra enterprise-apps gallery; exchange SAML metadata XML. Tableau Server: requires valid SHA-2 SSL certificate first; configured via TSM.
Multiple auth configurations per site (2024.3+): up to 18-20 auth methods on one site. SAML attribute assertions feed Tableau user-attribute functions (2025.3+).
Limitation: no native Entra group support in permissions — groups must be synced via SCIM first.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/enterprise/service-admin-azure-ad-b2b
https://learn.microsoft.com/en-us/entra/identity/conditional-access/concept-conditional-access-cloud-apps
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-service-principal""",
  i="""Official documentation:
https://help.tableau.com/current/online/en-us/saml_config_azure_ad.htm
https://help.tableau.com/current/online/en-us/scim_config_azure_ad.htm
https://help.tableau.com/current/server/en-us/saml_config_steps_tsm.htm""",
  j="""Power BI wins decisively for Azure AD / Microsoft-centric organisations.
Power BI's Entra ID integration is zero-configuration — SSO, MFA, and Conditional Access work automatically for any M365 tenant. Tableau requires explicit SAML/OIDC setup, SCIM provisioning, and certificate management (for Server).
GTF is a Microsoft shop (Azure EDW, ADF, Databricks on Azure) — Power BI's native Entra integration means SSO just works, with no ongoing IdP configuration maintenance.
Advantage: POWER BI — intrinsic SSO/MFA for Entra tenants; Tableau requires explicit SAML setup and ongoing maintenance.""",
  k="COMPLETED"
)

wb.save(FILEPATH)
print("Batch 1 saved: S.No 7, 9, 21, 22, 23")
