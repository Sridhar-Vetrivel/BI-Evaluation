import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 31: Query response time under concurrent load ──
u(31,
  f="""Performance varies significantly by connection mode.
Import mode (VertiPaq): sub-100ms typical for well-modeled datasets; engine is columnar in-memory — most interactive queries return near-instantaneously. Premium dedicated capacity eliminates shared-resource contention.
DirectQuery / Live to AtScale: performance is governed by AtScale's autonomous aggregate engine. AtScale's in-memory aggregates deliver near-import speed for common query patterns; cold queries against Databricks may take 2-10s.
Diagnostics: Performance Analyzer (Desktop) shows per-visual query time; DAX Studio for server-side traces; Fabric Capacity Metrics for throughput/throttling.
Concurrency: Premium capacity scales via CU allocation; shared capacity throttles above a query concurrency ceiling. Dedicated capacity is required for guaranteed SLAs at 100+ concurrent users.""",
  g="""Performance depends on extract vs live and underlying source.
Extract (Hyper engine): sub-second for most analytical queries on millions of rows; Hyper is a high-performance in-process columnar engine. Reliable SLA for dashboards built on extracts.
Live to AtScale: same AtScale aggregate benefits as Power BI; Tableau fires SQL against AtScale — performance is comparable once AtScale aggregates warm.
Diagnostics: Performance Recording (Help > Settings and Performance > Start Performance Recording) captures per-sheet query time and SQL; Tableau Server Admin Views show workbook load times across users.
Concurrency: Tableau Cloud autoscales; Tableau Server capacity is fixed to licensed cores — plan for peak concurrent sessions in sizing.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/guidance/power-bi-optimization
https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-performance-analyzer
https://learn.microsoft.com/en-us/fabric/enterprise/metrics-app""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/perf_collect_server_repo.htm
https://help.tableau.com/current/pro/desktop/en-us/perf_record_create_desktop.htm
https://help.tableau.com/current/server/en-us/adminview_stats_load_time.htm""",
  j="""Tie for the GTF/AtScale architecture.
Both tools connect live to AtScale and benefit equally from AtScale's autonomous aggregates — the semantic layer, not the BI tool, determines query speed for common operational reports. For Power BI's Import-mode workloads (non-AtScale sources), VertiPaq delivers consistently faster queries than Tableau's live equivalent.
Premium dedicated capacity gives Power BI more predictable SLA guarantees at high concurrency; Tableau Cloud autoscales but at unpredictable cost.
Advantage: TIE on AtScale path; Power BI edge for Import-mode non-AtScale sources.""",
  k="BLOCKER"
)

# ── S.No 32: Dataset size limits ──
u(32,
  f="""Size limits depend heavily on licensing tier.
Import mode: 1 GB model limit on Pro (shared capacity); up to 400 GB on Premium P1 / Fabric F64 (large dataset storage). Model is compressed in-memory — a 400 GB uncompressed dataset may compress to 50-100 GB in VertiPaq.
DirectQuery / Live to AtScale: no model size limit — the semantic model lives in AtScale; Power BI stores only the schema and rendering layer.
Fabric: large dataset storage mode allows models beyond the in-memory RAM ceiling using SSD spill; upper bound scales with Fabric capacity purchased.
Row limits: no enforced row cap in Import within the size ceiling; VertiPaq compression handles billions of rows if memory allows.""",
  g="""Extract (Hyper engine) has no enforced row limit.
Hyper engine handles 100M+ row extracts in production environments; practical ceiling is disk and RAM on the Tableau Server/Cloud hosting. Tableau Cloud storage quota: 100 GB per site (Basic), 500 GB (Enterprise/Tableau+).
Live to AtScale: no extract size limit — all data stays in AtScale; Tableau renders results from live queries.
Column limit: no strict column cap in Hyper; wide tables (500+ columns) can degrade performance.
For GTF's ~118M row synthetic dataset: Hyper extract is viable; Live to AtScale is the recommended production pattern.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-large-models
https://learn.microsoft.com/en-us/power-bi/fundamentals/service-features-license-type
https://learn.microsoft.com/en-us/power-bi/connect-data/service-dataset-modes-understand""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/extracting_data.htm
https://help.tableau.com/current/online/en-us/to_site_storage.htm
https://help.tableau.com/current/pro/desktop/en-us/hyper_overview.htm""",
  j="""Tableau edge for raw volume without license constraints.
Power BI Import mode size limits are license-tiered (1 GB on Pro — unusable for GTF's dataset; 400 GB on Premium). For GTF, Premium/Fabric is required to handle the full dataset in Import mode — an additional cost gate.
Tableau Hyper extract has no enforced row limit and can handle GTF's 118M row dataset on Basic Cloud storage.
However, for the production AtScale architecture, both tools use Live Connection — dataset size limits become irrelevant as data never leaves AtScale.
Advantage: TABLEAU for extract-based workloads without Premium licensing; TIE on AtScale Live path.""",
  k="COMPLETED"
)

# ── S.No 33: Aggregation and caching strategies ──
u(33,
  f="""Rich native aggregation framework — strongest in the market.
VertiPaq in-memory engine: automatic columnar compression and bitmap indexes — most aggregations execute in-engine at microsecond speed.
Aggregation tables: define a pre-aggregated summary table in the model; Power BI routes matching queries to it transparently (no report change needed). Dual storage mode: aggregation table in-memory, detail table in DirectQuery.
Service-level caching: Power BI Service caches query results per filter state — identical queries from different users hit cache, not the source.
With AtScale: AtScale provides its own autonomous aggregate engine (in-memory + Databricks push). Power BI's VertiPaq adds a second caching layer on top — effectively two-level caching for Import-connected models.""",
  g="""Caching through Hyper extract + AtScale aggregates.
Hyper extract is itself a pre-aggregated columnar cache — Tableau materializes the dataset locally, and Hyper executes queries against this in-memory/on-disk store at high speed.
Live connection to AtScale: no local Hyper cache; all aggregation happens in AtScale. Tableau passes SQL GROUP BY queries to AtScale; AtScale's aggregate store handles them.
Context filters: Tableau materializes a temp table for the first large-reduction filter — acts as a session-level cache for that query subset.
No equivalent to Power BI's Aggregation Tables or Service query-result cache. Tableau performance tuning relies more on data source optimization and AtScale than on BI-layer caching.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/transform-model/aggregations-advanced
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-incremental-refresh
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-storage-mode""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/extracting_data.htm
https://help.tableau.com/current/pro/desktop/en-us/data_extract_faq.htm
https://help.tableau.com/current/pro/desktop/en-us/filtering_context_overview.htm""",
  j="""Power BI edge for multi-layer caching architecture.
Power BI offers Aggregation Tables (BI-layer pre-agg), VertiPaq in-memory compression, and Service query-result caching — three independent caching mechanisms. Tableau relies on Hyper extracts (good for static snapshots) and passes aggregation responsibility to AtScale for live connections.
For GTF/AtScale architecture: AtScale's autonomous aggregates are the primary performance layer for both tools — the BI-layer caching difference narrows on the live path. Power BI's Aggregation Tables remain relevant for any non-AtScale Import-mode models.
Advantage: POWER BI for BI-layer caching sophistication; TIE on AtScale live architecture.""",
  k="BLOCKER"
)

# ── S.No 36: High-cardinality dimensions ──
u(36,
  f="""Handles high cardinality well in Import mode via VertiPaq dictionary encoding.
VertiPaq compresses high-cardinality string columns using hash encoding + bitmaps — millions of distinct values are manageable within the model size ceiling. Example: a franchise_id or customer_id dimension with 10M distinct values compresses significantly vs raw storage.
DirectQuery: high cardinality impacts join pushdown; CROSSJOIN or high-cardinality slicer selections can generate slow multi-value IN clauses sent to the source. Use query reduction settings (add Apply button) to defer query firing.
RLS scaling: dynamic RLS on very-high-cardinality security columns (e.g., 1M+ user→row mappings) can degrade; AtScale-side enforcement is more scalable for those patterns.""",
  g="""Handles high cardinality well in Hyper extract mode.
Hyper engine uses columnar storage with dictionary encoding — high-cardinality dimensions (millions of distinct values) are efficiently stored and queried. No enforced cardinality ceiling.
Live to AtScale: cardinality is managed by AtScale and the underlying Databricks store; Tableau passes the query and renders the result. No Tableau-side cardinality limit.
LOD expressions and high cardinality: FIXED LODs can cause cardinality explosion if the dimension key in the FIXED expression is very high-cardinality — produces a large intermediate result set before re-aggregating. Design LOD expressions carefully on high-cardinality keys.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/guidance/star-schema
https://learn.microsoft.com/en-us/power-bi/connect-data/directquery-troubleshoot
https://learn.microsoft.com/en-us/power-bi/guidance/power-bi-optimization#high-cardinality-columns""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/extracting_data.htm
https://help.tableau.com/current/pro/desktop/en-us/calculations_calculatedfields_lod.htm
https://www.tableau.com/about/blog/2019/3/hyper-performance""",
  j="""Tie — both handle high cardinality within their respective engines.
Power BI VertiPaq and Tableau Hyper both use columnar dictionary encoding and handle millions of distinct values efficiently in cached/extract mode. DirectQuery/Live behavior is governed primarily by AtScale and Databricks, not the BI tool.
Key watch-out for Tableau: poorly designed LOD expressions on high-cardinality keys can cause query fan-out; DAX CROSSJOIN in Power BI has the same risk. Both require deliberate data modeling.
Advantage: TIE — cardinality handling is equivalent for the GTF/AtScale architecture.""",
  k="COMPLETED"
)

# ── S.No 37: Scheduled refresh frequency ──
u(37,
  f="""Refresh frequency is license-tiered — a meaningful constraint on Pro.
Pro (shared capacity): maximum 8 scheduled refreshes per day per dataset. Premium Per User / Premium capacity / Fabric: maximum 48 refreshes per day.
XMLA endpoint (Premium): programmatic refresh via TMSL/TOM with no daily-count ceiling — enables hourly or event-driven refresh patterns.
Incremental refresh: reduces refresh time by only processing recent partitions; pairs with near-real-time hybrid table on Premium for last N minutes of live data.
Concurrency: up to 6 active refreshes simultaneously per Premium capacity node; shared capacity limits concurrent refreshes more aggressively.""",
  g="""No enforced daily refresh count ceiling.
Tableau Server: fully configurable refresh schedules (minutes, hours, days, weeks, monthly) with no enforced count cap. Administrator sets schedule frequency.
Tableau Cloud: minimum extract refresh interval is 15 minutes (no sub-15-min extract refresh). Live connections refresh on user interaction, not on a clock — no refresh scheduling needed for live.
Tableau Bridge: 10 concurrent extract refreshes per Bridge client (tunable); pool multiple clients for higher concurrency.
Incremental refresh: append-based (new rows since last key value); less sophisticated than Power BI's partition-based approach but covers the common case.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/connect-data/refresh-scheduled-refresh
https://learn.microsoft.com/en-us/power-bi/enterprise/service-premium-incremental-refresh
https://learn.microsoft.com/en-us/analysis-services/xmla/xml-elements-methods-execute""",
  i="""Official documentation:
https://help.tableau.com/current/online/en-us/schedule_add.htm
https://help.tableau.com/current/online/en-us/to_keep_data_fresh.htm
https://help.tableau.com/current/server/en-us/schedule_manage_extract.htm""",
  j="""Tableau wins on scheduling flexibility; moot for AtScale live architecture.
Power BI's 8-refresh/day cap on Pro is a hard operational constraint for high-frequency data — requires Premium to reach 48/day or XMLA for programmatic refresh. Tableau has no equivalent daily cap.
However, for GTF's AtScale architecture: both tools use Live Connection to AtScale, which eliminates extract refresh scheduling entirely. AtScale manages its own aggregate refresh against Databricks — the BI tool does not schedule refreshes.
Advantage: TABLEAU for extract-based workloads; TIE for AtScale live architecture where refresh scheduling is irrelevant.""",
  k="COMPLETED"
)

# ── S.No 38: Complex calculations on render time ──
u(38,
  f="""DAX in Import mode is extremely fast for complex business logic.
VertiPaq executes DAX filter context evaluations in-engine; well-written measures (CALCULATE, SWITCH, DIVIDE) return in milliseconds even with complex hierarchies. Calculation Groups (via Tabular Editor) reuse time-intelligence logic across all measures without measure proliferation.
Complex iterators (SUMX over many rows, CROSSJOIN of large tables) are the main performance risk — identify with DAX Studio's Server Timings trace.
DirectQuery path: each DAX measure translates to a SQL query against AtScale; complex DAX with many CALCULATE transitions can generate verbose SQL — risk of slow renders on deeply nested calculations.""",
  g="""Tableau calculated fields compile at query time.
Simple calculated fields (arithmetic, string, date) add minimal overhead. LOD expressions (FIXED, INCLUDE, EXCLUDE) generate nested sub-queries — FIXED LOD on a high-cardinality key against a live source can be slow.
Context filters + LOD combinations: multiple context filters compounded with LOD calcs can cause exponential query fan-out on live connections. Use Performance Recording to identify the offending calculation.
With AtScale: Tableau passes SQL + computed expressions to AtScale; AtScale's query optimizer handles most standard aggregations. Complex Tableau calcs that cannot push down to SQL execute in Tableau's query engine — potential bottleneck for intricate calculated fields.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/guidance/dax-variables
https://learn.microsoft.com/en-us/power-bi/guidance/directquery-model-guidance
https://daxstudio.org/ (free tool — use Server Timings to profile DAX)""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/calculations_calculatedfields_lod.htm
https://help.tableau.com/current/pro/desktop/en-us/perf_record_create_desktop.htm
https://help.tableau.com/current/pro/desktop/en-us/filtering_context_overview.htm""",
  j="""Power BI edge for complex financial business logic.
DAX in VertiPaq (Import mode) handles deeply nested financial calculations — time intelligence, parent-child hierarchies, rolling averages — fast and elegantly, especially with Calculation Groups. Tableau's LOD expressions are powerful for analytical patterns but require careful design to avoid slow sub-query fan-out on live sources.
For GTF P&L reporting with many derived measures (Variance %, MTD, YTD, Prior Year): DAX is better suited and better documented for this pattern.
Advantage: POWER BI for complex financial calculations; TABLEAU holds ground for analytical/exploratory LOD patterns.""",
  k="BLOCKER"
)

wb.save(FILEPATH)
print("Batch 2 saved: S.No 31, 32, 33, 36, 37, 38")
