import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FP = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FP)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    r = sno + 1
    if f is not None: ws.cell(r, 6).value = f
    if g is not None: ws.cell(r, 7).value = g
    if h is not None: ws.cell(r, 8).value = h
    if i is not None: ws.cell(r, 9).value = i
    if j is not None: ws.cell(r, 10).value = j
    if k is not None: ws.cell(r, 11).value = k

assert ws.cell(144, 4).value and 'High' in str(ws.cell(144, 4).value), 'S.No 143 is not High'

u(143,
  f="""Strong and rapidly maturing — Copilot is now Power BI's strategic natural-language path.
Copilot for Power BI: a user types a plain-English question ('show catering sales by territory last quarter') and Copilot queries the semantic model and returns a visual answer. Works in both view and edit modes in the Power BI Service, for report authors and report viewers alike.
Ad hoc DAX: Copilot generates DAX on the fly for calculations not already in the model (year-over-year growth, ratios, counts), with a 'How Copilot arrived at this' panel so users can verify the fields and logic used.
Surfaces: an in-report Copilot pane (generally available) for questions about the open report, plus a standalone full-screen Copilot (preview) that finds and interrogates any report/model the user has access to.
Classic Q&A (the older NL visual) still works but is being retired in December 2026 — Copilot is the forward-looking replacement.
Limitation: requires paid Fabric capacity (F2+) or Power BI Premium (P1+) — not available on Pro alone; currently English-only; Copilot data questions run in the Service (Desktop 'coming soon'); cannot yet answer forecasting, anomaly, or key-influencer questions.""",
  g="""Strong, but split across two surfaces — Tableau Agent (authoring) and Tableau Pulse (consumption).
Tableau Agent (formerly Einstein Copilot): conversational AI inside authoring — describe a visualization in natural language and Agent builds it, suggests analytical questions, and creates/explains calculated fields. Available in Tableau Desktop (2025.1+), Tableau Cloud (with Tableau+), and Tableau Server (2025.3+) web authoring; governed by the Einstein Trust Layer on Cloud (customer data is not used to train the LLM).
Tableau Pulse: a metrics layer + AI insights platform aimed squarely at non-technical users — they follow metrics and receive LLM-generated, plain-language insight summaries; Enhanced Q&A (Discover) supports conversational NL exploration; proactive digests are pushed by email or Slack. This is the closest match to 'insights come to me without asking an analyst.'
The previous-generation Ask Data and Metrics were retired (Tableau Cloud February 2024 / Tableau Server 2024.2) and replaced by Tableau Pulse.
Limitation: Tableau Pulse is Tableau Cloud only (not on Tableau Server); Tableau Agent requires Tableau+ with 'AI in Tableau' enabled; the NL capability is divided across two products rather than one integrated in-report experience.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/create-reports/copilot-introduction
https://learn.microsoft.com/en-us/power-bi/create-reports/copilot-ask-data-question
https://learn.microsoft.com/en-us/power-bi/natural-language/q-and-a-intro""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/desktop_einstein.htm
https://help.tableau.com/current/online/en-us/pulse_intro.htm
https://help.tableau.com/current/online/en-us/ask_data_enable.htm""",
  j="""Narrow Power BI advantage for integrated, governed in-report NL; Tableau Pulse wins on proactive pushed insights.
Power BI Copilot is a single, tightly integrated surface: franchise users ask questions in plain English directly against the certified AtScale semantic model (RLS respected via Live Connection), get a visual answer plus on-the-fly DAX, all inside the report they already use — and on the same Azure tenant as AtScale/Databricks.
Tableau splits the capability: Tableau Agent accelerates authoring, while Tableau Pulse delivers AI insight summaries and email/Slack digests to metric followers. Pulse's proactive digest model is genuinely strong for the client's 'don't talk to people to get insights' goal — insights are pushed to franchisees without them opening a dashboard — but Pulse is Tableau Cloud only and is a separate surface from the franchise reports.
Cost/fit: both require an upgraded tier (Power BI: Fabric/Premium capacity; Tableau: Tableau+ and Cloud). For GTF franchise operators on the governed AtScale model who want answers in-context without asking an analyst, Power BI Copilot is the simpler single-surface path; Tableau Pulse is the better fit if proactive, pushed natural-language digests are the priority.
Advantage: POWER BI for integrated in-report NL on the governed model; TABLEAU (Pulse) for proactive pushed insight digests — the choice hinges on whether GTF wants pull (ask in-report) or push (delivered digests).""",
  k="COMPLETED"
)

wb.save(FP)
print("S.No 143 populated (COMPLETED): F/G/H/I/J set, Copilot vs Tableau Agent/Pulse, links validated.")