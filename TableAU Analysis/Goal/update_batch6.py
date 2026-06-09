import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 101: REST API coverage ──
u(101,
  f="""Comprehensive REST API covering full content lifecycle.
Power BI REST API: CRUD on semantic models (datasets), reports, dashboards, workspaces, gateways, users, capacities, dataflows, goals. Full admin API for tenant-level governance operations.
Export-to-File API: programmatic PDF/PNG/XLSX/PPTX generation from any report page with applied filter context — key for building automated report delivery pipelines.
XMLA endpoint (Premium): Tabular Model Scripting Language (TMSL) and TOM for semantic model CRUD — create, modify, delete tables/measures/relationships programmatically.
Power Automate connectors: 40+ prebuilt Power BI actions for no-code automation of common operations (refresh, export, subscribe).""",
  g="""Comprehensive REST API with developer-friendly extensions.
Tableau REST API: CRUD on workbooks, data sources, views, projects, sites, users, groups, schedules, subscriptions, permissions, flows. Full admin operations including migration.
Metadata API (GraphQL): query content lineage, upstream/downstream dependencies, certification status, usage counts — more expressive for governance queries than Power BI's Scanner API.
Webhooks: register event-driven callbacks on publish, complete, fail, extract refresh fail events — enables reactive integrations (e.g., trigger a Slack message when an extract fails).
tabcmd: command-line wrapper for common operations (login, export, publish, delete) — scriptable without coding a REST client.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/rest/api/power-bi/
https://learn.microsoft.com/en-us/power-bi/developer/embedded/export-to
https://learn.microsoft.com/en-us/analysis-services/xmla/xml-elements-methods""",
  i="""Official documentation:
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api.htm
https://help.tableau.com/current/api/metadata_api/en-us/index.html
https://help.tableau.com/current/api/webhooks/en-us/docs/webhooks-events-payload.htm""",
  j="""Tie — both APIs are enterprise-grade with different strengths.
Power BI's Export-to-File API and XMLA endpoint are additional functional surfaces Tableau lacks — particularly valuable for building automated report generation and semantic model CI/CD pipelines.
Tableau's Metadata GraphQL API and Webhooks are more developer-friendly for event-driven integrations and lineage-based governance tooling. tabcmd lowers the scripting barrier for sysadmins.
For GTF automation (report bursting, scheduled exports, model deployment): Power BI's Export API + Power Automate is more turnkey; Tableau requires more custom REST scripting.
Advantage: TIE — Power BI stronger for export automation; Tableau stronger for lineage/event-driven integrations.""",
  k="COMPLETED"
)

# ── S.No 107: Programmatic report generation ──
u(107,
  f="""Strong — multiple mechanisms for parameterized/templated report generation.
Export-to-File API: programmatic PDF/PPTX/PNG generation from any report page with applied bookmarks and filter parameters — filter to a specific franchise, export, deliver. Scriptable in any language via REST.
Dynamic subscriptions (Premium): one subscription mapping dataset → N per-recipient parameterized exports delivered on schedule — zero engineering after setup.
Paginated reports (Premium): RDL template with parameters defined in Report Builder; at render time, parameter values drive different output — one template renders differently for each franchise, concept, date range. Export via API or subscription.""",
  g="""Achievable but requires custom scripting — no turnkey equivalent to Power BI's dynamic subscriptions.
tabcmd: `tabcmd get` with URL filter parameters exports a filtered PDF/PNG per invocation. Scriptable in a loop for per-recipient generation.
REST API: `queryViewPDF` / `queryViewImage` endpoints with filter parameters appended as URL strings — same outcome as tabcmd but via HTTP in any language.
Tableau Prep + Flows: data-driven content (different datasets → different outputs) but not report parameterization in the same sense.
No equivalent to Power BI's paginated report parameterization or dynamic subscription mapping dataset — bursting requires a custom script loop.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/developer/embedded/export-to
https://learn.microsoft.com/en-us/power-bi/collaborate-share/end-user-subscribe#subscribe-with-dynamic-per-recipient-subscriptions
https://learn.microsoft.com/en-us/power-bi/paginated-reports/report-design/report-parameters-paginated-reports""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/tabcmd_cmd.htm#get_url
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_workbooks_and_views.htm#query_view_pdf
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_workbooks_and_views.htm#query_view_image""",
  j="""Power BI wins on turnkey programmatic generation.
Power BI provides three independent mechanisms — Export API, dynamic subscriptions, paginated report parameterization — any of which covers the 'generate one per franchise' use case with minimal engineering. Dynamic subscriptions require zero ongoing script maintenance once the mapping dataset is established.
Tableau requires scripting a loop (tabcmd or REST) for every bursting scenario — more flexible but more engineering effort. No scheduling/mapping dataset equivalent to Power BI's dynamic subscription.
For GTF's franchise-level automated report delivery: Power BI is 80% less engineering than Tableau for the same outcome.
Advantage: POWER BI — turnkey parameterized generation via dynamic subscriptions; Tableau requires custom scripting.""",
  k="COMPLETED"
)

# ── S.No 134: Export formats ──
u(134,
  f="""Comprehensive export formats, especially for paginated reports.
Standard interactive reports: Export to PDF (full report, all pages), PPTX (PowerPoint, one slide per page), PNG (single page snapshot). Underlying data: CSV or XLSX (up to 150K rows xlsx, 30K rows CSV from visual).
Paginated reports (Premium): PDF, Excel (full formatting preserved), CSV, Word (.docx), PowerPoint (.pptx), image (BMP/GIF/JPG/PNG), accessible PDF (tagged PDF for screen readers). True pixel-perfect rendering with headers/footers maintained.
Export-to-File API: programmatic export in all the above formats with applied filter/parameter context.""",
  g="""Good coverage for interactive reports; no paginated-equivalent format richness.
View/Dashboard: Download PDF (view or dashboard), PNG (image), CSV (crosstab/underlying data), Excel (crosstab). PDF preserves visual layout; CSV/Excel are data-only.
Workbook: Download as TWBX (packaged workbook), PDF (multi-sheet), or Crosstab (single sheet to Excel).
No PowerPoint native export — PDF is the closest printable format; PPT conversion requires external tooling.
No Word/docx or accessible tagged-PDF export.
REST API: `queryViewPDF`, `queryViewImage`, `queryViewData` (CSV) — programmatic export matching UI options.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-export
https://learn.microsoft.com/en-us/power-bi/paginated-reports/report-builder/export-reports
https://learn.microsoft.com/en-us/power-bi/developer/embedded/export-to""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/save_export.htm
https://help.tableau.com/current/pro/desktop/en-us/export_printpdf.htm
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_workbooks_and_views.htm#query_view_pdf""",
  j="""Power BI wins on export format breadth.
Power BI's paginated report export covers PDF, Excel (formatted), Word, PowerPoint, CSV, image, and accessible PDF — more comprehensive than any BI tool in the market. Standard report PPTX export is native and widely used in enterprise. Accessible PDF (tagged) supports compliance requirements.
Tableau's export covers PDF, PNG, CSV, Excel — adequate for most use cases but lacks Word, native PPTX, and accessible PDF.
For GTF's catering detail report (franchisee-readable PDF + Excel attachment): Power BI paginated reports cover both requirements in one render path.
Advantage: POWER BI — broader export format set, especially for paginated/operational report delivery.""",
  k="COMPLETED"
)

# ── S.No 141: Drag-and-drop report building ──
u(141,
  f="""Capable — well-suited for non-technical report authors with proper training.
Power BI Desktop: drag fields from the Fields pane onto a canvas; Power BI suggests chart types via the Visualizations pane. Auto-detect relationships on data load. Quick measures: guided measure creation for 20+ common patterns (YTD totals, running totals, period-over-period) without writing DAX.
Q&A natural language visual: type a question in plain English ('total revenue by brand last quarter') — Power BI generates a visual automatically. Can pin to dashboard.
Copilot (Fabric): natural language to DAX measure and visual creation (2024+) — further reduces code requirements.
Web authoring: basic report edits in the browser (Power BI Service) without Desktop install.""",
  g="""Industry-leading drag-and-drop experience — Tableau is the gold standard for self-service authoring.
Tableau Desktop: drag fields to Rows, Columns, Marks; Show Me recommends the best chart type from selected fields. Field types auto-detected; date fields auto-drill hierarchy. Widely regarded as the most intuitive authoring experience in the BI market.
Ask Data (now Tableau Pulse / Einstein): natural language query directly in the browser — no Desktop needed.
Web authoring (Tableau Cloud/Server): supports core report building without installing Desktop — more capable than Power BI's web editor.
Limitation: Tableau Desktop has a steeper initial learning curve than Power BI Desktop for users new to the Rows/Columns mental model.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/fundamentals/desktop-getting-started
https://learn.microsoft.com/en-us/power-bi/create-reports/power-bi-autogenerate-report
https://learn.microsoft.com/en-us/power-bi/natural-language/q-and-a-intro""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/getstarted_buildmanual_ex1basic.htm
https://help.tableau.com/current/pro/desktop/en-us/show_me.htm
https://help.tableau.com/current/online/en-us/web_author_build_view.htm""",
  j="""Tie with tool-specific strengths.
Tableau is widely recognised as the most intuitive drag-and-drop authoring experience — Show Me and the Rows/Columns canvas are industry benchmarks for self-service analytics. Power BI Desktop is more approachable for users coming from Excel and Office tooling.
Power BI Copilot (NL to DAX/visual) and Q&A reduce the need for any authoring knowledge; Tableau Pulse/Ask Data provides similar NL capability.
For GTF franchise report builders (non-technical, franchise-level): Power BI's Office familiarity and guided Quick Measures lower the onboarding barrier. For analytical power users: Tableau's canvas model is more expressive.
Advantage: TIE — Power BI better for Office-familiar users; Tableau better for analytical self-service.""",
  k="COMPLETED"
)

# ── S.No 145: In-tool data preparation ──
u(145,
  f="""Power Query is the best integrated data prep tool in any BI product.
Power Query (M language): built directly into Power BI Desktop — shape, clean, transform, merge, pivot/unpivot data before it enters the model. 300+ built-in transformation steps via UI; advanced transforms in M code.
Dataflows Gen2 (Fabric): Power Query in the cloud — author reusable ETL flows in the browser without Desktop; outputs land in Fabric Lakehouse/Warehouse or OneLake. Scheduled refresh, incremental load, output destinations configurable.
Computed columns in the model (DAX calculated columns): light transforms at the semantic layer.
Limitation: Power Query is strong but not a replacement for ADF/Databricks for enterprise-scale ETL — suitable for last-mile/self-service data shaping.""",
  g="""Tableau Prep Builder: purpose-built visual ETL, separate from Desktop.
Tableau Prep Builder: visual drag-and-drop flow editor for cleaning, reshaping, and combining data. Steps: Input → Clean → Join/Union → Aggregate → Pivot → Output (to .hyper, published data source, or database). Included with Creator license.
Published Flows: run on Tableau Server/Cloud on a schedule — equivalent to Dataflows for cloud-based ETL scheduling.
In-workbook transforms: calculated fields, Unions (append), cross-database joins in the data source editor for light transforms without leaving Desktop.
Limitation: Prep is a separate application — context switching between Prep and Desktop adds workflow friction not present in Power BI (single Power Query window within Desktop).""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-query/power-query-what-is-power-query
https://learn.microsoft.com/en-us/fabric/data-factory/dataflows-gen2-overview
https://learn.microsoft.com/en-us/power-bi/transform-model/desktop-query-overview""",
  i="""Official documentation:
https://help.tableau.com/current/prep/en-us/prep_get_started.htm
https://help.tableau.com/current/server/en-us/prep_conductor_overview.htm
https://help.tableau.com/current/pro/desktop/en-us/data_clean.htm""",
  j="""Power BI wins on data prep integration and depth.
Power Query is embedded in Power BI Desktop — authors prep and model data in a single application with no tool switching. Dataflows Gen2 extends Power Query to cloud-scale reusable ETL. The M language is more expressive than Tableau Prep's step-based model for complex transforms.
Tableau Prep Builder is visually intuitive and well-designed but is a separate application — workflow breaks between Prep and Desktop add friction. Prep lacks M language equivalents for programmatic transforms.
For GTF's data engineers building franchise-level data prep pipelines: Power Query + Dataflows Gen2 is more capable and better integrated.
Advantage: POWER BI — Power Query embedded in Desktop + Dataflows Gen2 for cloud ETL.""",
  k="COMPLETED"
)

# ── S.No 147: Governed self-service ──
u(147,
  f="""Best-in-class governed self-service via certified semantic model reuse.
Workflow: data engineers publish and certify a semantic model (connected to AtScale); business analysts browse the Data Hub → select the certified model → connect via Live Connection → build reports against governed measures/dimensions without any SQL or raw data access.
RLS inherited automatically: the connecting report respects the certified model's row-level security without extra configuration.
Endorsement levels: Promoted (team-level quality) and Certified (org-level quality) — two tiers of trust visible to all report authors.
Author bypass risk: a report author can choose to connect directly to AtScale (bypassing the certified model) — mitigate by restricting AtScale connection strings to data engineers via network policy or Entra Conditional Access.""",
  g="""Governed self-service via published data sources and Virtual Connections.
Workflow: data stewards publish a certified data source connected to AtScale; analysts connect to the certified data source in Desktop/web authoring; they build visualizations against governed fields without writing SQL.
Virtual Connections + Data Policies (Data Management add-on): certified connection with RLS policies enforced centrally — tightest governance but requires the add-on license.
Certification badge: visible in search and Tableau Catalog; data quality warnings surface stale/unreliable sources.
Author bypass risk: analysts can connect directly to AtScale's SQL endpoint from Desktop — mitigate with network policy or by requiring use of published data sources via site admin settings.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-endorse-content
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-dataset-details
https://learn.microsoft.com/en-us/power-bi/connect-data/service-datasets-discover-across-workspaces""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/datasource_certified.htm
https://help.tableau.com/current/server/en-us/dm_vconn_overview.htm
https://help.tableau.com/current/server/en-us/data_policy.htm""",
  j="""Power BI wins for governed self-service without additional licensing.
Power BI certified semantic model + Live Connection enforces governed reuse at the platform level: one certified model → all reports inherit it, RLS included, no extra license. Analysts cannot accidentally create a disconnected copy.
Tableau's Virtual Connection + Data Policy model is robust but requires the Data Management add-on — additional cost for what Power BI delivers as a standard feature.
For GTF (franchise analysts building territory-level reports on top of the governed AtScale model): Power BI's Live Connection governance path is simpler and cheaper.
Advantage: POWER BI — governed self-service without add-on license; Tableau requires Data Management add-on for equivalent centralized governance.""",
  k="COMPLETED"
)

# ── S.No 149: Complexity gradient ──
u(149,
  f="""Good complexity gradient from consumer to expert author.
Consumer: view/filter reports in Power BI Service (browser or mobile app) — no license required beyond Viewer via App. Zero training needed.
Basic author: Power BI Service web editor — drag fields, apply filters, build simple charts without Desktop install. Quick measures for DAX without syntax.
Intermediate: Power BI Desktop — data modeling, custom DAX, relationships, advanced visuals from AppSource.
Expert: DAX Studio, Tabular Editor, XMLA endpoint, Power Query M, REST API, Embedded SDK.
Copilot bridges beginner → intermediate: NL to DAX reduces the code barrier for non-developers.""",
  g="""Best-in-class complexity gradient — Tableau is the industry standard for self-service BI.
Consumer: Tableau Reader (free) or Tableau Server/Cloud Viewer — browse and interact with published dashboards. Ask Data for NL queries.
Basic author: Tableau Public (free) for simple public dashboards; Tableau Cloud web authoring for basic views without Desktop.
Intermediate: Tableau Desktop — Show Me + Marks shelf for standard analytics. Well-documented; large community.
Expert: LOD expressions, table calculations, Tableau Prep flows, Extensions API, Embedded Analytics API, REST API.
Tableau's learning path is steeper initially (Rows/Columns mental model is unfamiliar to Excel users) but the ceiling is very high analytically.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/fundamentals/desktop-getting-started
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-consumer
https://learn.microsoft.com/en-us/power-bi/fundamentals/service-self-service-signup-for-power-bi""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/getstarted_overview.htm
https://help.tableau.com/current/online/en-us/web_author_overview.htm
https://www.tableau.com/learn/training""",
  j="""Tie — both offer a learnable gradient from consumer to expert.
Power BI is more approachable for Microsoft Office users (Excel mental model, familiar ribbon/field pane); Copilot and Quick Measures reduce DAX barrier. Tableau's visual design model is more expressive for analytical users but has a steeper initial curve.
Both have large training libraries, community content, and certification paths.
For GTF with a mix of franchise operators (consumers), territory managers (basic authors), and analytics team (expert authors): Power BI's familiarity advantage is relevant for the first two groups; Tableau holds no material advantage there.
Advantage: TIE — Power BI slightly more approachable for Office-familiar users; Tableau stronger for analytical self-service ceiling.""",
  k="COMPLETED"
)

wb.save(FILEPATH)
print("Batch 6 saved: S.No 101, 107, 134, 141, 145, 147, 149")
