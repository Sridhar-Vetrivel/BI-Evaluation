import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 81: Star/snowflake schema support (AtScale consumption lens) ──
u(81,
  f="""Power BI receives AtScale's pre-built dimensional model via XMLA — schema arrives fully formed.
When connecting Live (XMLA), Power BI's model pane shows AtScale's fact and dimension tables with all relationships already defined and correct cardinality set. Report authors see a ready star schema with no relationship wiring needed.
Additional local tables (e.g., a budget Excel sheet) can be added via Composite model on top of the AtScale Live Connection — Power BI supports this; Tableau does not equivalently.
VertiPaq optimizes star schema queries extremely well (single-table filtering via bitmap joins).""",
  g="""Tableau receives AtScale as SQL tables — relationships must be wired manually in Tableau's logical layer.
When connecting to AtScale as a SQL source, Tableau sees individual tables. Authors must define Relationships between facts and dimensions in the logical layer (drag dimension primary key to fact foreign key, set cardinality and referential integrity).
Tableau's logical layer (noodle model) supports star schemas well once wired — each relationship is a JOIN that Tableau generates on demand. Snowflake schemas supported by chaining relationships.
Extra authoring step vs Power BI's pre-formed XMLA model — adds setup time per workbook/published data source.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/transform-model/desktop-relationships-understand
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-composite-models
https://atscale.com/resource/power-bi-live-connection/""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/relate_tables.htm
https://help.tableau.com/current/pro/desktop/en-us/datasource_multitable_normalized.htm
https://atscale.com/resource/tableau-connector/""",
  j="""Power BI edge for AtScale schema consumption.
Power BI's XMLA Live Connection to AtScale delivers the star schema pre-formed — report authors start from a ready model with no relationship configuration. Tableau authors must manually wire every fact-dimension relationship from AtScale's SQL tables, adding setup time per workbook or published data source.
Composite models in Power BI allow blending the AtScale XMLA model with additional local data sources — a unique hybrid capability Tableau lacks.
Advantage: POWER BI — schema arrives pre-defined via XMLA; less authoring overhead per report.""",
  k="BLOCKER"
)

# ── S.No 82: Calculated columns, measures, KPI definitions ──
u(82,
  f="""AtScale certified measures surface in Power BI as native model measures.
Via XMLA Live Connection, all AtScale-defined measures (revenue, margin, YTD, etc.) appear in the Fields pane exactly as certified — no recreation needed.
Report-level measures (DAX): Power BI uniquely allows adding DAX measures on top of a Live Connection model without modifying AtScale — last-mile report-specific calculations that do not pollute the governed semantic model.
Calculation Groups (via Tabular Editor): reusable time-intelligence and formatting patterns applied across all measures — 1 Calculation Group replaces 20+ individual time-intelligence measures.
KPI visual: built-in KPI visual type for target vs actual vs trend in a single widget.""",
  g="""AtScale measures arrive as fields in the Tableau data pane.
When connecting to AtScale via SQL, Tableau exposes AtScale's pre-calculated measures as available fields — report authors drag them onto the canvas without recreating SQL.
Report-level calculated fields: Tableau calculated fields can be defined on top of AtScale's fields — using standard Tableau expression language (not DAX). LOD expressions, table calculations (WINDOW_SUM, RANK, INDEX), and string/date functions available.
No Calculation Group equivalent — time intelligence variants (YTD, MTD, Rolling) must be authored as separate calculated fields.
KPI: no dedicated KPI visual type — achieved via BANs (Big Ass Numbers), text tables with conditional formatting, or BAN+sparkline combos.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/transform-model/desktop-measures
https://learn.microsoft.com/en-us/power-bi/transform-model/calculation-groups
https://learn.microsoft.com/en-us/power-bi/visuals/power-bi-visualization-kpi""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/calculations_calculatedfields_create.htm
https://help.tableau.com/current/pro/desktop/en-us/calculations_calculatedfields_lod.htm
https://help.tableau.com/current/pro/desktop/en-us/calculations_tablecalculations.htm""",
  j="""Power BI edge for business calculation layer.
AtScale-defined measures surface cleanly in both tools. Power BI's unique advantage: report-level DAX measures extend the model without modifying AtScale — powerful for last-mile business logic (custom P&L lines, blended metrics). Calculation Groups avoid measure proliferation elegantly.
Tableau calculated fields are more accessible to analysts (familiar syntax) but lack DAX's expressiveness for complex financial logic. No Calculation Group equivalent means more manual measure maintenance.
For GTF's P&L and franchise performance KPIs: DAX Calculation Groups + report-level measures are a meaningful authoring productivity advantage.
Advantage: POWER BI for complex financial calculations and calculation reuse; TIE for standard aggregation and analytical calcs.""",
  k="BLOCKER"
)

# ── S.No 83: Reusable certified dataset / semantic model ──
u(83,
  f="""First-class certified semantic model ecosystem.
Power BI semantic model endorsement: workspace admins can 'Promote' or 'Certify' a semantic model — certified badge appears in the Data Hub and connector dialogs. Certification signals the model is authoritative and governed.
Live Connection reuse: any report author in the tenant can connect to a certified semantic model via Live Connection (Get Data → Power BI semantic models) — no data copy, no duplication, inherits all RLS/OLS definitions automatically.
In the GTF AtScale pattern: AtScale's model becomes a certified Power BI semantic model; hundreds of reports all connect to the same live source — one update propagates everywhere.
Lineage view: tracks which reports consume which semantic model.""",
  g="""Published data sources with certification support.
Tableau published data sources: authors connect to a published data source (workbook → data pane → right-click → Publish to Server) — makes it available to all site users as a reusable, live connection or extract.
Certification badge: admins can certify a published data source — appears with badge in Tableau's search and data catalog. Data quality warnings flag stale or untrustworthy sources.
Governed catalog: Tableau Data Management add-on provides a full data catalog (Tableau Catalog) with lineage, impact analysis, certified source filtering, and sensitivity labels.
Without Data Management: certification badge exists but catalog and lineage views are not available — reduces governance visibility.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-endorse-content
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-dataset-details
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-data-lineage""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/datasource_certified.htm
https://help.tableau.com/current/server/en-us/dm_catalog_overview.htm
https://help.tableau.com/current/server/en-us/dm_lineage.htm""",
  j="""Power BI wins for governed reuse without additional licensing.
Power BI certified semantic model + Live Connection is a tightly integrated, no-add-on pattern: certify once in the Data Hub, connect from anywhere, inherit all security — zero data copy.
Tableau's published data source certification works but the lineage and catalog features that make governance visible require the Data Management add-on (extra license cost).
For GTF (AtScale as the single certified semantic layer): Power BI's model promotion/certification + Live Connection is the natural fit — certify the AtScale-connected model once, propagate to all franchise reports.
Advantage: POWER BI — certified semantic model reuse + lineage included without add-on.""",
  k="BLOCKER"
)

# ── S.No 85: Hierarchy definition and drill path management ──
u(85,
  f="""Hierarchies from AtScale arrive pre-built via XMLA — no author reconfiguration.
AtScale's dimensional hierarchies (Brand → Concept → Unit; Year → Quarter → Month) are transmitted through the XMLA connection and appear as expandable hierarchies in Power BI's field list. Authors simply drag a hierarchy onto a visual's axis — drill up/down controls appear automatically.
Date hierarchy: Power BI auto-generates a date hierarchy from any date column if no hierarchy is explicitly defined — Year, Quarter, Month, Day without any authoring.
Additional hierarchies: Power BI authors can define supplemental hierarchies within the local model layer (for composite model scenarios).
Drill-through pages: full context-passing drill-through to a detail page is a separate but complementary capability.""",
  g="""Hierarchies must be manually defined from AtScale's SQL columns.
AtScale exposes flat SQL columns (brand_name, unit_name, year, quarter, month, day) — Tableau authors must define drill paths by dragging dimension fields into a parent-child structure in the Data pane (e.g., drag Unit onto Brand).
Date dimension: Tableau auto-creates date drill hierarchy (Year > Quarter > Month > Day > Hour > Minute) from any date/datetime column — no manual setup needed.
Custom hierarchies: drag one field onto another in the Data pane; Tableau creates a hierarchy folder. Limited to the order defined — cannot reorder drill levels at runtime.
Drill-through equivalent: Dashboard Action (Go to Sheet) with filter context — achieves the same outcome with more manual configuration.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-drill-through-buttons
https://learn.microsoft.com/en-us/power-bi/visuals/desktop-matrix-visual#expand-and-collapse-row-headers
https://learn.microsoft.com/en-us/power-bi/transform-model/desktop-tutorial-create-measures""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/qs_hierarchies.htm
https://help.tableau.com/current/pro/desktop/en-us/actions_dashboards.htm
https://help.tableau.com/current/pro/desktop/en-us/drilldown.htm""",
  j="""Power BI advantage for AtScale hierarchy consumption.
AtScale-defined hierarchies (Brand → Concept → Unit; fiscal calendar hierarchy) are transmitted pre-built through the XMLA connection — Power BI authors never manually wire drill paths. Tableau authors must manually build each hierarchy from AtScale's flat SQL columns, adding setup time per published data source.
Date hierarchies are auto-handled equally in both tools.
For GTF with a multi-level franchise hierarchy (Brand → Territory → Unit): Power BI authors save meaningful setup time per workbook.
Advantage: POWER BI — AtScale hierarchies arrive pre-defined via XMLA; Tableau requires manual hierarchy configuration from SQL columns.""",
  k="BLOCKER"
)

# ── S.No 86: Time intelligence functions ──
u(86,
  f="""Native DAX time intelligence is the strongest in the BI market.
AtScale exposes time-grain columns and pre-calculated period measures; Power BI authors add last-mile DAX time intelligence on top — TOTALYTD, TOTALMTD, DATESINPERIOD, SAMEPERIODLASTYEAR, PARALLELPERIOD, DATEADD — all execute natively in the Live Connection model.
Calculation Groups (Tabular Editor): define YTD, MTD, QTD, Rolling 12M, Prior Year, Variance vs PY as one Calculation Group — applies to every measure automatically. Eliminates measure sprawl for time intelligence.
Report-level time intelligence measures can be added without modifying AtScale — authors work on top of the governed model independently.""",
  g="""Tableau date functions are capable but more manual to set up for enterprise time intelligence.
AtScale provides time-grain columns; Tableau adds DATETRUNC, DATEADD, DATEDIFF, DATENAME calculated fields for period comparisons. YTD: FIXED LOD on year + conditional date filter. MTD: similar pattern with month filter. Rolling N periods: WINDOW_SUM over date-ordered partition.
No Calculation Group equivalent — every time intelligence variant (YTD Revenue, YTD Margin, MTD Revenue, Prior Year Revenue…) is a separate calculated field. For 10 measures × 5 period variants = 50 calculated fields vs 1 Calculation Group in Power BI.
Dynamic periods via Parameter: parameter + CASE calculated field to switch between period types — works but adds authoring complexity.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/transform-model/calculation-groups
https://dax.guide/st/time-intelligence/
https://learn.microsoft.com/en-us/dax/time-intelligence-functions-dax""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/functions_functions_date.htm
https://help.tableau.com/current/pro/desktop/en-us/calculations_calculatedfields_lod.htm
https://help.tableau.com/current/pro/desktop/en-us/calculations_tablecalculations.htm""",
  j="""Power BI wins on time intelligence, particularly at scale.
DAX time intelligence functions are purpose-built for financial period calculations — no workarounds needed for standard enterprise patterns. Calculation Groups reduce measure count dramatically (1 group vs dozens of individual measures).
Tableau achieves the same results but requires more calculated fields and more authoring effort per time period variant. For GTF with P&L reports requiring YTD, MTD, QTD, Rolling 12M, Prior Year across 10+ measures, Calculation Groups are a significant productivity and maintenance advantage.
Advantage: POWER BI — native time intelligence functions + Calculation Groups vs Tableau's manual per-variant calculated fields.""",
  k="BLOCKER"
)

# ── S.No 87: Centralized metric layer support ──
u(87,
  f="""First-class AtScale integration via XMLA — and growing Fabric metric capabilities.
AtScale native Power BI connector (XMLA endpoint, compatibility level 1600): Power BI connects in Live Connection mode with full DAX support. AtScale's governed metrics (revenue, margin, volume) are the single source of truth — Power BI consumes, not re-defines.
Fabric Metrics Hub (2025+): Microsoft's own metric layer within Fabric, defining KPIs with owners, units, targets, and status — complements AtScale for Fabric-managed metrics.
dbt Semantic Layer: Power BI does not have a native dbt Semantic Layer (MetricFlow) connector — dbt metrics are accessed via the dbt-connected data source (Databricks/Synapse), not through the MetricFlow API directly.""",
  g="""AtScale SQL connector + dbt Semantic Layer support — dual metric layer compatibility.
AtScale Tableau Cloud Connector (OAuth SSO, no Bridge required): Tableau connects live to AtScale's SQL endpoint — AtScale's certified measures and dimensions are available as standard SQL fields. Cleanest setup for Tableau Cloud.
dbt Semantic Layer (MetricFlow): Tableau has an official dbt Semantic Layer connector (2024) — connects to MetricFlow-defined metrics via the dbt Semantic Layer API. If GTF uses dbt for metric definitions, Tableau can query them natively.
Semantic layer co-existence: Tableau can connect to multiple semantic layers (AtScale for operational metrics, dbt for transformational metrics) in separate published data sources.""",
  h="""Official documentation:
https://atscale.com/resource/power-bi-live-connection/
https://learn.microsoft.com/en-us/fabric/fundamentals/metrics-hub
https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-directquery-datasets-azure-analysis-services""",
  i="""Official documentation:
https://atscale.com/resource/tableau-connector/
https://help.tableau.com/current/pro/desktop/en-us/examples_dbt.htm
https://docs.getdbt.com/docs/use-dbt-semantic-layer/tableau""",
  j="""Tie — both integrate well with AtScale; Tableau has dbt Semantic Layer advantage.
Power BI's XMLA Live Connection to AtScale delivers the richest model fidelity (pre-formed schema, hierarchies, certified measures via DAX). Tableau's AtScale SQL connector is equally capable and is AtScale's most mature 'out-of-the-box' integration per AtScale's documentation.
If GTF uses dbt for metric management alongside AtScale, Tableau's native dbt Semantic Layer connector is a differentiator Power BI currently lacks.
For GTF (AtScale is the primary semantic layer, no dbt Semantic Layer): TIE.
Advantage: TIE for AtScale-only; TABLEAU if dbt Semantic Layer is also used.""",
  k="BLOCKER"
)

# ── S.No 88: Govern and version the semantic model independently ──
u(88,
  f="""Strong native Git integration for Power BI semantic models (GA 2024).
Fabric/Power BI Git integration: semantic model (.bim / TMDL) and report (report.json) stored as files in Azure DevOps or GitHub. Every publish from the Service commits to the branch; branch pull updates the workspace. Standard Git workflows (branch, PR, merge, revert) apply to BI content.
Deployment pipelines: structured Dev → Test → Prod promotion with per-stage data source override rules — separates environment concerns from model versioning.
AtScale semantic model versioning: AtScale governs its own model version independently. Power BI Live Connection picks up AtScale model changes automatically — no Power BI model republish needed when AtScale measures update.""",
  g="""No native Git integration — relies on revision history and third-party tools.
Tableau revision history: Server/Cloud stores the last N published versions of each workbook and data source (configurable N) — lightweight rollback per item. Not true version control: no branching, no diff, no PR workflow.
Content Migration Tool: moves workbooks/data sources between projects/sites/servers — useful for environment promotion (dev → prod) but not version control.
Third-party Git integration: Broadcom Catalyst Tableau Extension, ZAP BI's Tableau Git — enable committing .twb/.twbx to Git repos. Not native; requires external tooling and process discipline.
AtScale model changes: Tableau's SQL connection to AtScale picks up schema changes automatically (new columns/measures appear); no workbook republish needed for pure AtScale metadata changes.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/developer/projects/projects-git
https://learn.microsoft.com/en-us/power-bi/create-reports/deployment-pipelines-overview
https://learn.microsoft.com/en-us/power-bi/developer/projects/projects-overview""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/revision_history_admin.htm
https://help.tableau.com/current/server/en-us/cmt-intro.htm
https://help.tableau.com/current/server/en-us/tp_about.htm""",
  j="""Power BI wins on semantic model version control.
Power BI's native Git integration (GA 2024) brings full branch/PR/merge workflows to semantic model and report files — standard software engineering practices applied to BI content. Deployment pipelines enforce Dev → Test → Prod promotion with governance gates.
Tableau's revision history is a lightweight rollback tool, not version control. Native Git for Tableau content requires third-party tooling and process enforcement — not built into the platform.
For GTF with multiple developers working on the BI layer simultaneously (franchise reports, brand analytics, catering): Git-based branching prevents content conflicts and enables code review of DAX measures.
Advantage: POWER BI — native Git integration + deployment pipelines vs Tableau's revision history only.""",
  k="BLOCKER"
)

wb.save(FILEPATH)
print("Batch 5 saved: S.No 81, 82, 83, 85, 86, 87, 88")
