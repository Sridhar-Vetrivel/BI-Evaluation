import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

FILEPATH = r'c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Goal\Discovery.xlsx'
wb = openpyxl.load_workbook(FILEPATH)
ws = wb['Evaluation Matrix - 1']

def u(sno, f=None, g=None, h=None, i=None, j=None, k=None):
    row = sno + 1
    if f is not None: ws.cell(row, 6).value = f
    if g is not None: ws.cell(row, 7).value = g
    if h is not None: ws.cell(row, 8).value = h
    if i is not None: ws.cell(row, 9).value = i
    if j is not None: ws.cell(row, 10).value = j
    if k is not None: ws.cell(row, 11).value = k

# ── S.No 41: Workspace / project-based collaboration ──
u(41,
  f="""Capable. Workspaces are collaboration containers: Admin, Member, Contributor, Viewer roles. Multiple users/groups share content in a workspace during development.
Git integration (GA 2024+): connect workspace to Azure DevOps/GitHub branch — commit/pull from within the Power BI Service UI. Supports parallel development branches.
Deployment pipelines (Premium/Fabric): Dev → Test → Prod promotion with automated content moving, data-source rule overrides per stage, and deployment comparisons.
Org Apps: packaged consumer experience published from workspace; audiences control which groups see which tabs.
Limitation: flat hierarchy — workspaces are peers with no nesting; complex orgs (e.g., multi-brand, multi-territory) must manage many peer workspaces.""",
  g="""Capable. Projects are nested containers with role-based access; support sub-projects for folder-like hierarchy. Project Leader role enables delegated governance without Server Admin.
Content can be developed and tested in separate dev/test projects or sites; the Content Migration Tool moves content between projects/sites/servers.
No native Git integration — third-party tools (Broadcom Catalyst Tableau Extension) provide version control; revision history in Server/Cloud is the built-in fallback (last N versions per workbook).
No deployment pipeline equivalent — teams use separate projects (dev/prod) or sites (strict isolation) as the promotion mechanism.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-new-workspaces
https://learn.microsoft.com/en-us/power-bi/developer/projects/projects-git
https://learn.microsoft.com/en-us/power-bi/create-reports/deployment-pipelines-overview""",
  i="""Official documentation:
https://help.tableau.com/current/server/en-us/projects.htm
https://help.tableau.com/current/server/en-us/projects_admin.htm
https://help.tableau.com/current/server/en-us/cmt-intro.htm""",
  j="""Tie with tool-specific strengths.
Power BI: cleaner Git integration (native, GA), deployment pipelines for structured promotion, and App-based consumer packaging — better for teams with DevOps discipline. Limitation: flat workspace hierarchy.
Tableau: better for hierarchical governance (nested projects, Project Leader delegation), more granular content permissions. Limitation: no native Git, no deployment pipeline.
For GTF: Power BI's deployment pipelines (Dev → Test → Prod) directly support the phased delivery approach; Git integration enables parallel brand-team development.
Advantage: POWER BI for DevOps/pipeline workflows; TABLEAU for hierarchical governance.""",
  k="COMPLETED"
)

# ── S.No 43: Sharing via link, embed, or direct user assignment ──
u(43,
  f="""Multiple sharing mechanisms available.
Share link: generates a read-only report link for internal org users (Entra-authenticated); optional per-user permission grant included. Links respect report-level RLS.
Per-item sharing: share specific reports or semantic models to individual users/groups without workspace access.
Embed: Power BI Embedded (for ISV/custom portals), SharePoint Online web part (native), Teams tab/meeting app (native), website embed with iframe + Entra token.
App sharing: publish App to specific Entra groups — cleanest enterprise consumer model.
Power BI Embedded (A-SKU or Fabric): full API control for white-label portals; 'embed for your customers' (service principal auth) for external users without Power BI licenses.""",
  g="""Multiple sharing mechanisms available.
Share link: Tableau Cloud/Server generates authenticated links; recipient must have a Tableau license.
Per-content sharing: assign permission to individual users/groups at workbook, view, or data-source level.
Embed: Tableau Embedded Analytics (JavaScript API v2/v3) for custom portals; iframe embed for simple use cases. OAuth SSO for embedded user identity pass-through.
Guest/External: Tableau Cloud supports guest access for Viewer-licensed users from external Entra tenants via Connected Apps.
Teams: embed via iframe or Tableau for Microsoft Teams connector (not as native as Power BI Teams integration). SharePoint: no native web part — use iframe or Tableau for SharePoint extension.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-share-reports
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-embed-secure
https://learn.microsoft.com/en-us/power-bi/developer/embedded/embedded-analytics-power-bi""",
  i="""Official documentation:
https://help.tableau.com/current/api/embedding_api/en-us/index.html
https://help.tableau.com/current/server/en-us/permissions.htm
https://help.tableau.com/current/online/en-us/connected_apps.htm""",
  j="""Power BI wins for Microsoft ecosystem sharing.
SharePoint native web part and Teams native tab/bot are zero-config for M365 orgs — no iframe wrangling. GTF already uses Microsoft 365, making Power BI the path of least resistance for sharing into Teams channels and SharePoint intranet portals.
Tableau's embedded analytics API is more powerful for custom-portal scenarios but requires developer build time and a Connected App configuration.
Both cover standard link-share and user-assignment patterns adequately.
Advantage: POWER BI for M365-ecosystem sharing (Teams, SharePoint); TIE for custom embedded analytics portals.""",
  k="COMPLETED"
)

# ── S.No 45: Subscription and snapshot sharing ──
u(45,
  f="""Strong — most capable subscription system in the comparison.
Standard subscriptions: schedule email delivery of report page as PDF/PPTX/PNG snapshot; configure per page, per recipient group, with conditional send ('Only send if data changed').
Dynamic per-recipient subscriptions (Premium/Fabric): one subscription definition, N rows in a mapping dataset — each row drives a differently filtered snapshot to a different recipient email. True data-driven bursting for franchise-level delivery.
Paginated report subscriptions (Premium): full pixel-perfect PDF/Excel attachment per schedule, with parameter-per-recipient rows. Ideal for catering detail report per franchise.
Power Automate integration: conditional/branded email workflows, Teams delivery, SharePoint archiving.""",
  g="""Basic subscriptions — no per-recipient dynamic filtering natively.
Subscriptions: email delivery of a view or workbook as PDF or PNG image on a configured schedule. Recipients receive the same content (filtered only to their RLS if configured).
'Only when data changes' option available on Tableau Cloud subscriptions.
Per-recipient bursting: not natively supported. Must script via REST API — create one subscription per recipient with applied URL filters, or use tabcmd to export filtered PDFs and deliver via email script. Third-party tools (ChristianSteven BI360, Rollstack) provide turnkey bursting.
Data-Driven Alerts: threshold-based email notification when a measure crosses a value — not a report snapshot, but useful for exception-based alerting.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-subscribe
https://learn.microsoft.com/en-us/power-bi/collaborate-share/end-user-subscribe#subscribe-with-dynamic-per-recipient-subscriptions
https://learn.microsoft.com/en-us/power-bi/paginated-reports/subscriptions/paginated-reports-subscriptions""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/subscribe_user.htm
https://help.tableau.com/current/server/en-us/subscribe.htm
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_subscriptions.htm""",
  j="""Power BI wins clearly — dynamic subscriptions are a core GTF requirement.
The 'Robots' use case (per-franchise filtered email delivery) is a native Power BI Premium feature — one subscription definition, one mapping dataset, automatic per-recipient bursting. Tableau requires REST API scripting or a third-party tool to achieve the same outcome.
For GTF with potentially hundreds of franchise recipients needing unique territory-filtered reports on a schedule, building a custom bursting script in Tableau is engineering overhead Power BI avoids.
Advantage: POWER BI — dynamic per-recipient subscriptions are native (Premium); Tableau requires custom build.""",
  k="COMPLETED"
)

# ── S.No 47: Notification and change alert mechanisms ──
u(47,
  f="""Rich alert ecosystem across multiple mechanisms.
Data alerts: set on dashboard tile visuals (card, gauge, KPI) — fires an email + Teams notification when a threshold is crossed. Reusable across org.
Power BI anomaly detection (Premium): AI-driven anomaly flagging on line chart visuals with one-click explanation. No threshold configuration needed.
Power Automate: trigger complex conditional workflows from Power BI data alerts — multi-step notifications, Teams messages, conditional branching, escalation chains.
Smart narratives and on-canvas AI insights (Fabric): surface automatic trend callouts in reports — not a push alert, but a reader-facing callout.""",
  g="""Threshold-based email alerts — simpler scope.
Data-Driven Alerts: set on a continuous axis mark in a published view on Server/Cloud. Fires email when measure crosses threshold; configurable check frequency.
Tableau Pulse (Cloud, 2024+): AI-driven metric monitoring with natural language insights delivered via email digest or embedded in Slack/Teams. Designed for executive-level metric health monitoring without requiring users to open Tableau.
Webhook API: register webhook callbacks on content events (workbook/data source publish, job complete/fail, extract refresh fail) — enables event-driven integrations.
Limitation: no equivalent to Power BI anomaly detection on arbitrary line charts; no native Teams notification without custom webhook/connector.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/consumer/end-user-alerts
https://learn.microsoft.com/en-us/power-bi/transform-model/desktop-anomaly-detection
https://learn.microsoft.com/en-us/power-bi/connect-data/power-bi-data-sources#power-automate""",
  i="""Official documentation:
https://help.tableau.com/current/pro/desktop/en-us/data_alerts.htm
https://help.tableau.com/current/online/en-us/pulse_set_up.htm
https://help.tableau.com/current/api/rest_api/en-us/REST/rest_api_ref_notifications.htm""",
  j="""Power BI wins on alert richness; Tableau Pulse is a differentiator for AI-driven monitoring.
Power BI's anomaly detection, Power Automate integration, and multi-channel delivery (email, Teams, mobile push) give a more comprehensive alerting ecosystem. Threshold-based data alerts are comparable.
Tableau Pulse is a notable differentiator for executive metric health monitoring with NL summaries — no direct Power BI equivalent in the same packaging.
For GTF operational alerting (franchise missing sales target, catering order spike): Power BI Power Automate flows give more customization. For board-level metric digests: Tableau Pulse is compelling.
Advantage: POWER BI for operational alerting; TABLEAU PULSE for executive metric digest.""",
  k="COMPLETED"
)

# ── S.No 48: Integration with collaboration platforms (Teams, Slack) ──
u(48,
  f="""Native Microsoft 365 ecosystem integration — strongest in market for Teams/SharePoint.
Microsoft Teams: native Power BI tab (add report as channel tab), Power BI meeting app (share report during meeting), Power BI Teams bot (query data in chat), subscribe reports directly from Teams UI.
SharePoint Online: native Power BI web part — embed live, filterable report in any SharePoint page without iFrame or custom code.
Outlook: email subscriptions with PDF/PPTX snapshots delivered to shared mailboxes; Power Automate for conditional branded HTML emails.
Viva Insights / M365 analytics integration: Power BI datasets feed into M365 productivity analytics.
Slack: not native — requires Power Automate connector or incoming webhook script.""",
  g="""Salesforce ecosystem strength; Microsoft ecosystem via configuration.
Slack: official Tableau for Slack connector — share views, receive alerts, and query Tableau Pulse metrics from Slack channels. Best-in-class Slack integration among BI tools.
Microsoft Teams: Tableau for Microsoft Teams connector — subscribe to views, get alerts, run Ask Data queries from Teams chat. Less native than Power BI's built-in Teams tab experience; no meeting app.
SharePoint: no native web part — embed via iframe or use Tableau for SharePoint (3rd-party extension by Tableau).
Salesforce: native integration — Tableau dashboards embed natively in Salesforce CRM Lightning pages, Slack (Salesforce owns both).
Email: standard subscriptions; no dynamic per-recipient without scripting.""",
  h="""Official documentation:
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-collaborate-microsoft-teams
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-embed-report-spo
https://learn.microsoft.com/en-us/power-bi/collaborate-share/service-microsoft-teams-app""",
  i="""Official documentation:
https://help.tableau.com/current/online/en-us/slack_admin.htm
https://help.tableau.com/current/online/en-us/teams_admin.htm
https://www.tableau.com/products/add-ons/tableau-for-microsoft-teams""",
  j="""Power BI wins decisively for the GTF Microsoft-centric stack.
GTF operates on Azure / Microsoft 365 — Power BI's native Teams tab, SharePoint web part, and Teams bot are zero-configuration integrations that Teams users encounter daily. No iFrame, no connector setup, no developer involvement.
Tableau's Slack integration is best-in-class (if GTF used Slack) and its Salesforce embedding is superior — but GTF is Teams + SharePoint, not Slack + Salesforce.
Advantage: POWER BI for Microsoft 365 collaboration stack (Teams, SharePoint, Outlook). TABLEAU if Slack or Salesforce CRM are primary collaboration tools.""",
  k="COMPLETED"
)

wb.save(FILEPATH)
print("Batch 3 saved: S.No 41, 43, 45, 47, 48")
