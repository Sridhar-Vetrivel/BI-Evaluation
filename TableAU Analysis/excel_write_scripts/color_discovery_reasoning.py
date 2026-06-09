"""Color column E (Discovery Reasoning) in Discovery.xlsx based on the four
buckets defined in the Legend sheet:

  Client (#1)  — theme=9, tint=0.40  (orange/peach)
  Docs   (#2)  — theme=8, tint=0.80  (light blue)
  GTF    (#3)  — theme=9, tint=0.60  (orange, lighter)
  POC    (#4)  — theme=6, tint=0.60  (light green)
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

FILLS = {
    "client": PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.40)),
    "docs":   PatternFill(patternType="solid", fgColor=Color(theme=8, tint=0.80)),
    "gtf":    PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.60)),
    "poc":    PatternFill(patternType="solid", fgColor=Color(theme=6, tint=0.60)),
}

# Per-breakdown-item classification.
# Keyed by exact text of column C (Breakdown Item).
CLASSIFY = {
    # 1. Data Connectivity & Integration
    "Native connectors to structured sources (SQL Server, Oracle, Snowflake, BigQuery, Redshift, SAP HANA)": "docs",
    "Support for semi-structured and unstructured sources (JSON, XML, REST APIs, web scraping)": "docs",
    "File-based ingestion (Excel, CSV, Parquet, Avro)": "docs",
    "Cloud storage connectors (S3, Azure Blob, GCS)": "docs",
    "ERP/CRM connectors (SAP, Salesforce, Dynamics 365, ServiceNow)": "docs",
    "Custom connector development capability (SDK support)": "docs",
    "Live connection vs. import mode availability per source": "poc",
    "OAuth and token-based authentication support for API sources": "docs",
    "Support for on-premises data sources via gateway": "poc",
    "Connector certification and update frequency by vendor": "docs",

    # 2. Visualization & UX
    "Library of standard chart types (bar, line, scatter, pie, waterfall, funnel, bullet)": "docs",
    "Advanced/specialized chart types (heatmaps, treemaps, Sankey, chord, small multiples)": "docs",
    "Custom visualization support (D3.js, Vega-Lite, marketplace extensions)": "docs",
    "Conditional formatting and dynamic visual properties": "poc",
    "Cross-filtering and drill-down/drill-through behavior": "poc",
    "Tooltip customization and rich interactivity": "docs",
    "Dashboard layout flexibility (pixel-perfect vs. responsive grid)": "poc",
    "Theming and branding consistency across reports": "docs",
    "Animation and transitions for storytelling": "docs",
    "Accessibility of visuals (screen reader compatibility, color-blind palettes)": "docs",

    # 3. Governance, Security & Compliance
    "Role-based access control (RBAC) at report, dataset, and workspace level": "poc",
    "Row-level security (RLS) and column-level security (CLS)": "poc",
    "Single Sign-On (SSO) integration (SAML, OIDC, Azure AD, Okta)": "poc",
    "Data classification and sensitivity label enforcement": "docs",
    "Compliance certifications held by vendor (SOC 2, ISO 27001, HIPAA, FedRAMP, GDPR)": "docs",
    "Data residency and sovereignty controls": "docs",
    "Tenant-level isolation in multi-tenant deployments": "docs",
    "Policy enforcement for sharing (prevent external sharing, download restrictions)": "docs",
    "Encryption at rest and in transit": "docs",
    "Periodic access reviews and certification workflows": "docs",

    # 4. Performance & Scalability
    "Query response time under concurrent user load (SLA benchmarks)": "poc",
    "Dataset size limits (row count, in-memory vs. DirectQuery thresholds)": "poc",
    "Aggregation and caching strategies (pre-aggregation, query caching, dual mode)": "poc",
    "Horizontal scaling capability for large deployments": "docs",
    "Performance profiling and query diagnostics tooling": "poc",
    "Handling of high-cardinality dimensions": "poc",
    "Scheduled refresh frequency limits and concurrency": "docs",
    "Impact of complex calculations on render time": "poc",
    "CDN support for global report distribution": "docs",
    "Load testing support and documented capacity planning models": "docs",

    # 5. Collaboration & Sharing
    "Workspace or project-based collaboration with role assignments": "gtf",
    "Report commenting and annotation features": "docs",
    "Sharing via link, embed, or direct user assignment": "docs",
    "External user sharing (guest/B2B access)": "client",
    "Subscription and snapshot sharing (static vs. live)": "poc",
    "Co-authoring or concurrent editing support": "docs",
    "Notification and change alert mechanisms": "docs",
    "Integration with collaboration platforms (Teams, Slack, email)": "poc",
    "Version history and rollback on shared content": "docs",
    "Approval workflows for publishing reports": "docs",

    # 6. Licensing & TCO
    "Licensing model (per user, per capacity/core, consumption-based)": "docs",
    "Cost at different scale points (50, 500, 5000 users)": "gtf",
    "Distinction between author/creator vs. viewer licensing costs": "gtf",
    "Premium or capacity-based tiers and what they unlock": "docs",
    "Cost of connectors, add-ons, and marketplace extensions": "docs",
    "On-premises vs. cloud licensing cost differential": "client",
    "Contractual flexibility (annual vs. monthly, enterprise agreements)": "gtf",
    "Hidden costs — gateway infrastructure, training, professional services": "gtf",
    "Vendor discount structures for large enterprise commitments": "gtf",
    "Total 3-year TCO model including infrastructure and admin overhead": "gtf",

    # 7. AI/ML & Advanced Analytics
    "Native AI visuals (anomaly detection, forecasting, key influencers)": "docs",
    "Natural language query interface (Q&A, Ask Data)": "docs",
    "Integration with ML platforms (Azure ML, SageMaker, Databricks, Vertex AI)": "docs",
    "Python and R script execution within the tool": "docs",
    "AutoML or no-code ML model building": "docs",
    "Explainability features for AI-generated insights": "docs",
    "Smart narrative and automated insight generation": "docs",
    "Predictive analytics and what-if scenario modeling": "poc",
    "Vector/embedding search or GenAI integration roadmap": "docs",
    "Ability to surface model outputs as first-class visuals": "docs",

    # 8. Deployment & IT Operations
    "Deployment modes — SaaS, self-hosted, hybrid": "client",
    "Cloud platform support (Azure, AWS, GCP, on-premises)": "docs",
    "High availability and disaster recovery architecture": "docs",
    "Upgrade and patch management process (vendor-managed vs. self-managed)": "docs",
    "Infrastructure-as-code support for provisioning": "docs",
    "Admin portal capabilities (usage monitoring, capacity management, user management)": "docs",
    "Multi-environment support (dev, staging, production)": "docs",
    "Backup and restore capabilities for BI content": "docs",
    "SLA commitments and incident response times": "docs",
    "Container and Kubernetes deployment support (for self-hosted options)": "docs",

    # 9. Data Modeling & Semantic Layer
    "Support for star and snowflake schema modeling": "poc",
    "Calculated columns, measures, and KPI definitions (DAX, Tableau calculations)": "poc",
    "Reusable certified dataset / semantic model concept": "poc",
    "Relationships and cardinality handling (many-to-many, bi-directional)": "poc",
    "Hierarchy definition and drill path management": "poc",
    "Time intelligence functions (YTD, MTD, rolling periods)": "poc",
    "Centralized metric layer support (compatibility with dbt Semantic Layer, AtScale, Cube)": "poc",
    "Ability to govern and version the semantic model independently of reports": "poc",
    "Support for composite models (mixing live and imported data)": "poc",
    "Data type handling and implicit vs. explicit conversion behavior": "poc",

    # 10. Embedded Analytics — all out of scope
    "Embedding via iFrame, JavaScript SDK, or REST API": "docs",
    "Row-level security propagation in embedded context": "docs",
    "White-labeling and full UI customization for embedded scenarios": "docs",
    "Token-based authentication for embedded sessions (no vendor login required)": "docs",
    "Licensing model for embedded usage (app-owns-data vs. user-owns-data)": "docs",
    "Interaction API for programmatic control of filters and visuals": "docs",
    "Performance of embedded reports at scale": "docs",
    "Cross-origin (CORS) and Content Security Policy (CSP) compliance": "docs",
    "Support for embedding in mobile apps (iOS/Android)": "docs",
    "Analytics portal pattern support (multi-tenant embedded deployments)": "docs",

    # 11. Extensibility & Developer Ecosystem
    "REST API coverage (CRUD on reports, datasets, workspaces, users)": "poc",
    "Webhook and event-driven integration support": "docs",
    "Custom visual development framework and marketplace": "docs",
    "CLI tooling for automation and scripting": "docs",
    "SDK availability and language support (Python, .NET, JavaScript, REST)": "docs",
    "Third-party extension ecosystem maturity": "docs",
    "Programmatic report generation (templating, parameterized reports)": "poc",
    "Integration with orchestration tools (Azure Data Factory, Airflow, dbt)": "poc",
    "Community-contributed connectors and visuals quality and maintenance": "docs",
    "Developer documentation quality and completeness": "docs",

    # 12. Vendor Maturity & Ecosystem
    "Gartner Magic Quadrant / Forrester Wave positioning and trajectory": "docs",
    "Product release cadence and public roadmap transparency": "docs",
    "Vendor financial stability and long-term viability": "docs",
    "Size and activity of community (forums, Stack Overflow, GitHub)": "docs",
    "Training and certification program availability and industry recognition": "gtf",
    "Partner ecosystem for implementation and support": "docs",
    "User group and conference ecosystem (Tableau Conference, Microsoft Fabric Community)": "docs",
    "Support tier options (standard, premier, dedicated TAM)": "docs",
    "Responsiveness to enterprise support tickets": "docs",
    "History of backward compatibility and migration support during major version changes": "docs",

    # 13. Mobile & Accessibility
    "Native mobile app availability (iOS and Android)": "client",
    "Responsive layout design vs. dedicated mobile layout authoring": "client",
    "Offline report access on mobile": "docs",
    "Touch-optimized interaction design": "docs",
    "Push notification support for alerts on mobile": "docs",
    "WCAG 2.1 AA compliance for accessibility": "docs",
    "Keyboard navigation support in desktop and web interfaces": "docs",
    "Screen reader compatibility (NVDA, JAWS, VoiceOver)": "docs",
    "Color contrast and color-blind safe defaults": "docs",
    "Accessibility audit tooling or built-in checker": "docs",

    # 14. Interoperability & Standards
    "ODBC/JDBC connectivity for third-party tool access": "docs",
    "OData endpoint exposure for report data consumption": "docs",
    "Compatibility with enterprise metadata catalogs (Microsoft Purview, Alation, Collibra, Atlan)": "docs",
    "Export formats supported (PDF, Excel, CSV, PNG, PowerPoint)": "poc",
    "Open standard support (Apache Arrow, XMLA endpoint, ANSI SQL)": "poc",
    "Integration with data catalog and data dictionary platforms": "docs",
    "Compatibility with ETL/ELT tools (Informatica, Talend, Fivetran, dbt)": "docs",
    "Cross-tool report migration tooling availability": "gtf",
    "Support for industry-specific data standards (FHIR for healthcare, XBRL for finance)": "docs",
    "API versioning and deprecation policy": "docs",

    # 15. Self-Service & Authoring Experience
    "Drag-and-drop report building without SQL or coding knowledge": "docs",
    "Guided analytics and template-based starting points": "docs",
    "Natural language query for ad-hoc exploration": "docs",
    "Field suggestion and smart auto-complete during authoring": "docs",
    "In-tool data preparation and transformation (Power Query, Tableau Prep)": "docs",
    "Undo/redo depth and autosave behavior": "docs",
    "Governed self-service model — ability to expose certified datasets for user-built reports": "poc",
    "Sandbox or personal workspace for experimentation without impacting production": "docs",
    "Complexity gradient — simple reports accessible to novices, advanced features for power users": "docs",
    "In-product help, tutorials, and contextual guidance": "docs",

    # 16. Report Consumption & Alerting
    "Data-driven alerts on threshold breaches (KPI alerts)": "docs",
    "Scheduled email subscriptions with report snapshots": "poc",
    "Personalized filter state persistence per user": "poc",
    "In-report commenting and collaborative annotation": "docs",
    "Print and export options for consumers (not just authors)": "poc",
    "Bookmark and personal view saving": "poc",
    "Paginated report support for operational/pixel-perfect outputs": "poc",
    "Report embedding in email or intranet portals": "docs",
    "In-app notification center for alerts and updates": "docs",
    "Performance of report load time for consumers on low-bandwidth connections": "gtf",

    # 17. Data Lineage & Audit Trail
    "End-to-end lineage from data source through transformation to visual": "poc",
    "Impact analysis — which reports are affected if a dataset changes": "poc",
    "Column-level lineage visibility": "docs",
    "Report and dataset usage analytics (who viewed, how often, last accessed)": "docs",
    "Integration with enterprise data catalog for lineage federation": "docs",
    "Audit logs for user activity (logins, report access, data exports, permission changes)": "docs",
    "Data freshness visibility — last refresh timestamp surfaced to consumers": "poc",
    "Stale report detection and notification": "docs",
    "Lineage graph visualization within the tool": "docs",
    "Retention policy for audit logs and compliance archival": "docs",

    # 18. Version Control & CI/CD for BI
    "Native Git integration for report and dataset versioning": "docs",
    "Branching and merging support for parallel development": "docs",
    "Deployment pipeline support (dev → test → prod promotion)": "docs",
    "Diff and comparison tooling for BI artifact changes": "docs",
    "Rollback capability to previous versions": "docs",
    "Automated testing framework for reports and measures (data validation, regression)": "docs",
    "CLI or API-driven deployment for pipeline automation": "docs",
    "Environment variable and parameter management across environments": "docs",
    "Integration with CI/CD platforms (Azure DevOps, GitHub Actions, Jenkins)": "docs",
    "Change approval and release gating workflows": "docs",

    # 19. Real-time & Streaming Analytics
    "Support for streaming data ingestion (Kafka, Event Hubs, Kinesis, MQTT)": "docs",
    "Push dataset API for real-time dashboard updates": "docs",
    "DirectQuery mode availability and performance across source types": "poc",
    "Configurable auto-refresh intervals for near-real-time use cases": "docs",
    "Incremental refresh support to reduce full dataset reload overhead": "poc",
    "Latency SLA between data event and dashboard update": "docs",
    "Hybrid mode support (mix of real-time and historical data in one dashboard)": "docs",
    "Alert triggering on streaming threshold breaches": "docs",
    "Scalability of real-time connections under concurrent viewer load": "docs",
    "Compatibility with stream processing platforms (Spark Streaming, Flink, Azure Stream Analytics)": "docs",

    # 20. Localization & Internationalization
    "UI language support and number of available locale translations": "client",
    "Locale-aware number, date, and currency formatting": "poc",
    "Right-to-left (RTL) language support (Arabic, Hebrew)": "docs",
    "Multi-currency display and conversion handling": "client",
    "Time zone management for scheduled refreshes and report timestamps": "client",
    "Translation workflow for report labels and field names": "docs",
    "Regional compliance support (GDPR for EU, PDPA for Thailand, LGPD for Brazil)": "client",
    "Support for non-Latin character sets (CJK, Devanagari, Cyrillic)": "docs",
    "Locale-specific sorting and collation rules": "docs",
    "Regional data residency options aligned with localization needs": "client",
}

wb = load_workbook(PATH)

# Locate the matrix sheet by header signature
matrix_sheet_name = None
for sn in wb.sheetnames:
    s = wb[sn]
    if s.cell(row=1, column=1).value == "S.No" and s.cell(row=1, column=3).value == "Breakdown Item":
        matrix_sheet_name = sn
        break
if matrix_sheet_name is None:
    raise RuntimeError(f"Could not locate Evaluation Matrix sheet. Sheets: {wb.sheetnames}")
print(f"Using sheet: {matrix_sheet_name!r}")
ws = wb[matrix_sheet_name]

counts = {"client": 0, "docs": 0, "gtf": 0, "poc": 0}
unmatched = []

for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=3).value
    if not item:
        continue
    bucket = CLASSIFY.get(item)
    if bucket is None:
        unmatched.append((r, item))
        continue
    cell = ws.cell(row=r, column=5)
    cell.fill = FILLS[bucket]
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    counts[bucket] += 1

wb.save(PATH)

print(f"Counts: {counts}")
print(f"Total colored: {sum(counts.values())}")
if unmatched:
    print(f"Unmatched ({len(unmatched)}):")
    for r, item in unmatched:
        print(f"  Row {r}: {item}")
