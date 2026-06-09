"""Build Discovery.xlsx with prioritized evaluation breakdown items."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (Category, [(Breakdown item, Priority, Reasoning)])
DATA = [
    ("1. Data Connectivity & Integration", [
        ("Native connectors to structured sources (SQL Server, Oracle, Snowflake, BigQuery, Redshift, SAP HANA)",
         "Low", "AtScale is the single source the BI tool talks to; direct DB connectors are not the primary path."),
        ("Support for semi-structured and unstructured sources (JSON, XML, REST APIs, web scraping)",
         "Low", "Not in scope — all curated data flows through Databricks/ADF/AtScale."),
        ("File-based ingestion (Excel, CSV, Parquet, Avro)",
         "Medium", "Ad-hoc analyst workflows and one-off uploads may still need this."),
        ("Cloud storage connectors (S3, Azure Blob, GCS)",
         "Low", "Databricks/ADF handles cloud-storage ingestion; BI tool does not need it directly."),
        ("ERP/CRM connectors (SAP, Salesforce, Dynamics 365, ServiceNow)",
         "Not needed", "Source systems land in EDW upstream; BI tool never connects to them."),
        ("Custom connector development capability (SDK support)",
         "Low", "No bespoke source planned beyond AtScale."),
        ("Live connection vs. import mode availability per source",
         "High", "Live mode to AtScale is the core consumption pattern — must be validated for both tools."),
        ("OAuth and token-based authentication support for API sources",
         "Low", "Not relevant for AtScale/SSAS consumption."),
        ("Support for on-premises data sources via gateway",
         "High", "Maps directly to client requirement #2 (Gateways for refreshing reports)."),
        ("Connector certification and update frequency by vendor",
         "Low", "Only the AtScale/SSAS connector matters; vendor breadth is secondary."),
    ]),
    ("2. Visualization & UX", [
        ("Library of standard chart types (bar, line, scatter, pie, waterfall, funnel, bullet)",
         "High", "Baseline for client requirement #6 (Visualizations)."),
        ("Advanced/specialized chart types (heatmaps, treemaps, Sankey, chord, small multiples)",
         "Medium", "Nice-to-have for executive/pretty dashboards (#6c) but not blocking."),
        ("Custom visualization support (D3.js, Vega-Lite, marketplace extensions)",
         "Medium", "Useful for branded/executive storytelling (#6c)."),
        ("Conditional formatting and dynamic visual properties",
         "High", "Required for P&L reporting (#6d) and KPI signals."),
        ("Cross-filtering and drill-down/drill-through behavior",
         "High", "Required for #6a interactive dashboards and #6d P&L drilldowns."),
        ("Tooltip customization and rich interactivity",
         "High", "Directly supports interactive dashboards (#6a)."),
        ("Dashboard layout flexibility (pixel-perfect vs. responsive grid)",
         "High", "Executive 'pretty' dashboards (#6c) and P&L matrix layouts (#6d) need this."),
        ("Theming and branding consistency across reports",
         "Medium", "Important for executive reports but not a hard blocker."),
        ("Animation and transitions for storytelling",
         "Low", "Cosmetic; not core to enterprise reporting needs."),
        ("Accessibility of visuals (screen reader compatibility, color-blind palettes)",
         "Low", "No stated accessibility/compliance mandate for this audience."),
    ]),
    ("3. Governance, Security & Compliance", [
        ("Role-based access control (RBAC) at report, dataset, and workspace level",
         "High", "Maps to client requirements #3 (Apps equivalent) and #4 (RLS)."),
        ("Row-level security (RLS) and column-level security (CLS)",
         "High", "Direct match to client requirement #4 — franchise-level filtering."),
        ("Single Sign-On (SSO) integration (SAML, OIDC, Azure AD, Okta)",
         "High", "Direct match to client requirement #1 (Azure AD/Entra ID)."),
        ("Data classification and sensitivity label enforcement",
         "Low", "No stated classification requirement from the client."),
        ("Compliance certifications held by vendor (SOC 2, ISO 27001, HIPAA, FedRAMP, GDPR)",
         "Medium", "Standard enterprise hygiene — verify but unlikely to be a differentiator."),
        ("Data residency and sovereignty controls",
         "Low", "No multi-region/sovereignty requirement surfaced yet."),
        ("Tenant-level isolation in multi-tenant deployments",
         "Low", "Single-tenant internal deployment expected."),
        ("Policy enforcement for sharing (prevent external sharing, download restrictions)",
         "Medium", "Useful guardrail for franchise data leakage."),
        ("Encryption at rest and in transit",
         "Medium", "Table-stakes for enterprise — confirm but both vendors handle this."),
        ("Periodic access reviews and certification workflows",
         "Low", "Likely managed at IAM layer (Azure AD), not the BI tool."),
    ]),
    ("4. Performance & Scalability", [
        ("Query response time under concurrent user load (SLA benchmarks)",
         "High", "Initial AtScale perf test was a strong signal; needs rigorous validation (req #7)."),
        ("Dataset size limits (row count, in-memory vs. DirectQuery thresholds)",
         "High", "Catering detail report (#7) drives this — must handle large row counts."),
        ("Aggregation and caching strategies (pre-aggregation, query caching, dual mode)",
         "High", "How each tool interacts with AtScale's aggregations is critical."),
        ("Horizontal scaling capability for large deployments",
         "Medium", "Important for capacity planning but not day-one blocker."),
        ("Performance profiling and query diagnostics tooling",
         "Medium", "Needed for ongoing tuning; not a selection blocker."),
        ("Handling of high-cardinality dimensions",
         "High", "Franchise + product + date combinations create high cardinality in detail reports."),
        ("Scheduled refresh frequency limits and concurrency",
         "High", "Maps to client requirement #2 (refresh scheduling)."),
        ("Impact of complex calculations on render time",
         "High", "P&L calcs (#6d) and time-intelligence measures must render fast."),
        ("CDN support for global report distribution",
         "Low", "Internal franchise audience likely concentrated regionally."),
        ("Load testing support and documented capacity planning models",
         "Medium", "Useful during rollout; vendor docs usually suffice."),
    ]),
    ("5. Collaboration & Sharing", [
        ("Workspace or project-based collaboration with role assignments",
         "High", "Maps to client requirement #3 (Apps / Projects / Sites / Collections)."),
        ("Report commenting and annotation features",
         "Low", "Not surfaced as a requirement; consumers mostly view."),
        ("Sharing via link, embed, or direct user assignment",
         "High", "Core distribution mechanism for franchise users (#3)."),
        ("External user sharing (guest/B2B access)",
         "Medium", "Franchisees may be external identities — depends on Azure AD model."),
        ("Subscription and snapshot sharing (static vs. live)",
         "High", "Directly supports client requirement #8 (subscriptions/email delivery)."),
        ("Co-authoring or concurrent editing support",
         "Low", "Small author base; not a critical workflow."),
        ("Notification and change alert mechanisms",
         "Medium", "Useful for operational reports; not blocking."),
        ("Integration with collaboration platforms (Teams, Slack, email)",
         "High", "Email delivery is part of #8 (Power BI Robots equivalent)."),
        ("Version history and rollback on shared content",
         "Medium", "Standard governance hygiene."),
        ("Approval workflows for publishing reports",
         "Low", "Not surfaced as a requirement."),
    ]),
    ("6. Licensing & TCO", [
        ("Licensing model (per user, per capacity/core, consumption-based)",
         "High", "Power BI vs Tableau pricing models differ structurally — central to the decision."),
        ("Cost at different scale points (50, 500, 5000 users)",
         "High", "Franchise audience may scale broadly; must model both vendors."),
        ("Distinction between author/creator vs. viewer licensing costs",
         "High", "Most franchise users are viewers — viewer cost dominates TCO."),
        ("Premium or capacity-based tiers and what they unlock",
         "High", "Power BI Premium / Tableau capacity tiers gate key features."),
        ("Cost of connectors, add-ons, and marketplace extensions",
         "Medium", "Mostly relevant if custom visuals/extensions are adopted."),
        ("On-premises vs. cloud licensing cost differential",
         "Medium", "Tableau Server vs Cloud is a real fork in the decision."),
        ("Contractual flexibility (annual vs. monthly, enterprise agreements)",
         "Medium", "Procurement input; not a technical blocker."),
        ("Hidden costs — gateway infrastructure, training, professional services",
         "High", "Gateway sizing (#2) and admin overhead materially affect TCO."),
        ("Vendor discount structures for large enterprise commitments",
         "Medium", "Negotiation lever; relevant if EA exists with Microsoft."),
        ("Total 3-year TCO model including infrastructure and admin overhead",
         "High", "Final decision artifact — must be produced for the client."),
    ]),
    ("7. AI/ML & Advanced Analytics", [
        ("Native AI visuals (anomaly detection, forecasting, key influencers)",
         "Low", "Not part of stated requirements; AtScale + downstream tools likely handle this."),
        ("Natural language query interface (Q&A, Ask Data)",
         "Low", "Not a stated user need; franchise users follow fixed reports."),
        ("Integration with ML platforms (Azure ML, SageMaker, Databricks, Vertex AI)",
         "Low", "ML lives in Databricks layer, not in BI tool."),
        ("Python and R script execution within the tool",
         "Low", "Not a stated need."),
        ("AutoML or no-code ML model building",
         "Not needed", "Out of scope for this evaluation."),
        ("Explainability features for AI-generated insights",
         "Not needed", "Out of scope."),
        ("Smart narrative and automated insight generation",
         "Low", "Nice-to-have; not a differentiator for franchise reporting."),
        ("Predictive analytics and what-if scenario modeling",
         "Medium", "What-if analysis is called out in #6a — needs validation."),
        ("Vector/embedding search or GenAI integration roadmap",
         "Low", "Forward-looking; not part of current decision."),
        ("Ability to surface model outputs as first-class visuals",
         "Low", "Tie to AtScale measures is the more likely path."),
    ]),
    ("8. Deployment & IT Operations", [
        ("Deployment modes — SaaS, self-hosted, hybrid",
         "High", "Tableau Cloud vs Server is a fork; Power BI is SaaS-only on web."),
        ("Cloud platform support (Azure, AWS, GCP, on-premises)",
         "High", "Existing stack is Azure (ADF, Databricks, AAS) — alignment matters."),
        ("High availability and disaster recovery architecture",
         "Medium", "Standard for enterprise rollout."),
        ("Upgrade and patch management process (vendor-managed vs. self-managed)",
         "Medium", "Differs sharply between Tableau Server and Power BI Service."),
        ("Infrastructure-as-code support for provisioning",
         "Low", "Not surfaced as a requirement."),
        ("Admin portal capabilities (usage monitoring, capacity management, user management)",
         "High", "Required to operate at franchise scale."),
        ("Multi-environment support (dev, staging, production)",
         "Medium", "Standard SDLC need."),
        ("Backup and restore capabilities for BI content",
         "Medium", "Operational hygiene."),
        ("SLA commitments and incident response times",
         "Medium", "Procurement input."),
        ("Container and Kubernetes deployment support (for self-hosted options)",
         "Low", "Only matters if Tableau Server self-hosted route is chosen."),
    ]),
    ("9. Data Modeling & Semantic Layer", [
        ("Support for star and snowflake schema modeling",
         "Medium", "AtScale owns the semantic model; BI tool sees a flattened view."),
        ("Calculated columns, measures, and KPI definitions (DAX, Tableau calculations)",
         "High", "Report-local measures still needed on top of AtScale (P&L variances, ratios)."),
        ("Reusable certified dataset / semantic model concept",
         "High", "How each tool surfaces AtScale as a 'certified' model matters for governance."),
        ("Relationships and cardinality handling (many-to-many, bi-directional)",
         "Medium", "Mostly resolved at AtScale; edge cases may surface in BI tool."),
        ("Hierarchy definition and drill path management",
         "High", "P&L (#6d) needs hierarchies and drilldowns."),
        ("Time intelligence functions (YTD, MTD, rolling periods)",
         "High", "P&L reporting (#6d) depends on this — split between AtScale and BI tool to validate."),
        ("Centralized metric layer support (compatibility with dbt Semantic Layer, AtScale, Cube)",
         "High", "CRITICAL — defines how cleanly each BI tool consumes AtScale."),
        ("Ability to govern and version the semantic model independently of reports",
         "High", "Core to the AtScale architectural bet."),
        ("Support for composite models (mixing live and imported data)",
         "Medium", "Useful if some sources can't go through AtScale."),
        ("Data type handling and implicit vs. explicit conversion behavior",
         "Medium", "Routine; can cause subtle bugs in P&L calcs."),
    ]),
    ("10. Embedded Analytics", [
        ("Embedding via iFrame, JavaScript SDK, or REST API",
         "Low", "Internal BI use case; embedding not a stated requirement."),
        ("Row-level security propagation in embedded context",
         "Not needed", "No embedded scenario in scope."),
        ("White-labeling and full UI customization for embedded scenarios",
         "Not needed", "No embedded scenario in scope."),
        ("Token-based authentication for embedded sessions (no vendor login required)",
         "Not needed", "No embedded scenario in scope."),
        ("Licensing model for embedded usage (app-owns-data vs. user-owns-data)",
         "Not needed", "No embedded scenario in scope."),
        ("Interaction API for programmatic control of filters and visuals",
         "Low", "Possible future need; not required now."),
        ("Performance of embedded reports at scale",
         "Not needed", "No embedded scenario in scope."),
        ("Cross-origin (CORS) and Content Security Policy (CSP) compliance",
         "Not needed", "No embedded scenario in scope."),
        ("Support for embedding in mobile apps (iOS/Android)",
         "Not needed", "No embedded scenario in scope."),
        ("Analytics portal pattern support (multi-tenant embedded deployments)",
         "Not needed", "No embedded scenario in scope."),
    ]),
    ("11. Extensibility & Developer Ecosystem", [
        ("REST API coverage (CRUD on reports, datasets, workspaces, users)",
         "High", "Required for #8 (Robots equivalent) — automation and franchise-filtered delivery."),
        ("Webhook and event-driven integration support",
         "Low", "Not surfaced as a requirement."),
        ("Custom visual development framework and marketplace",
         "Medium", "Useful for branded executive visuals (#6c)."),
        ("CLI tooling for automation and scripting",
         "Medium", "Helps with ops/CI but not blocking."),
        ("SDK availability and language support (Python, .NET, JavaScript, REST)",
         "Medium", "Supports automation work for #8."),
        ("Third-party extension ecosystem maturity",
         "Low", "Tableau has a deeper ecosystem; informational not blocking."),
        ("Programmatic report generation (templating, parameterized reports)",
         "High", "Core to #8 — franchise-specific report delivery requires parameterization."),
        ("Integration with orchestration tools (Azure Data Factory, Airflow, dbt)",
         "Medium", "ADF already orchestrates upstream; possibly extend to BI refresh."),
        ("Community-contributed connectors and visuals quality and maintenance",
         "Low", "Marginal — internal use case."),
        ("Developer documentation quality and completeness",
         "Medium", "Affects implementation cost."),
    ]),
    ("12. Vendor Maturity & Ecosystem", [
        ("Gartner Magic Quadrant / Forrester Wave positioning and trajectory",
         "Medium", "Useful framing for stakeholders; both tools are leaders."),
        ("Product release cadence and public roadmap transparency",
         "Medium", "Power BI ships monthly; Tableau quarterly — affects long-term planning."),
        ("Vendor financial stability and long-term viability",
         "Medium", "Microsoft and Salesforce-owned Tableau both stable."),
        ("Size and activity of community (forums, Stack Overflow, GitHub)",
         "Low", "Both have strong communities; not a differentiator."),
        ("Training and certification program availability and industry recognition",
         "Medium", "Relevant for talent hiring/upskilling cost in TCO."),
        ("Partner ecosystem for implementation and support",
         "Medium", "Relevant for delivery options."),
        ("User group and conference ecosystem (Tableau Conference, Microsoft Fabric Community)",
         "Low", "Optional learning channel."),
        ("Support tier options (standard, premier, dedicated TAM)",
         "Medium", "Procurement input."),
        ("Responsiveness to enterprise support tickets",
         "Medium", "Hard to evaluate without references."),
        ("History of backward compatibility and migration support during major version changes",
         "Medium", "Matters for long-horizon TCO."),
    ]),
    ("13. Mobile & Accessibility", [
        ("Native mobile app availability (iOS and Android)",
         "Medium", "Franchise managers/field staff likely consume on mobile."),
        ("Responsive layout design vs. dedicated mobile layout authoring",
         "Medium", "Authoring overhead for mobile layouts affects effort estimates."),
        ("Offline report access on mobile",
         "Low", "Not a stated need."),
        ("Touch-optimized interaction design",
         "Low", "Standard; both tools support."),
        ("Push notification support for alerts on mobile",
         "Low", "Not a stated requirement."),
        ("WCAG 2.1 AA compliance for accessibility",
         "Low", "No accessibility mandate surfaced."),
        ("Keyboard navigation support in desktop and web interfaces",
         "Low", "Standard."),
        ("Screen reader compatibility (NVDA, JAWS, VoiceOver)",
         "Low", "No accessibility mandate surfaced."),
        ("Color contrast and color-blind safe defaults",
         "Low", "Cosmetic; configurable."),
        ("Accessibility audit tooling or built-in checker",
         "Low", "No mandate."),
    ]),
    ("14. Interoperability & Standards", [
        ("ODBC/JDBC connectivity for third-party tool access",
         "Low", "AtScale already exposes standard endpoints; not a tool-side need."),
        ("OData endpoint exposure for report data consumption",
         "Low", "Not a stated requirement."),
        ("Compatibility with enterprise metadata catalogs (Microsoft Purview, Alation, Collibra, Atlan)",
         "Low", "Catalog strategy not stated; can be deferred."),
        ("Export formats supported (PDF, Excel, CSV, PNG, PowerPoint)",
         "High", "Required for #7 (export capability) and #8 (snapshot delivery)."),
        ("Open standard support (Apache Arrow, XMLA endpoint, ANSI SQL)",
         "Medium", "XMLA matters for SSAS/AtScale connectivity (#9)."),
        ("Integration with data catalog and data dictionary platforms",
         "Low", "Not a stated requirement."),
        ("Compatibility with ETL/ELT tools (Informatica, Talend, Fivetran, dbt)",
         "Low", "ETL is upstream of AtScale, not the BI tool."),
        ("Cross-tool report migration tooling availability",
         "Medium", "Migrating existing Power BI reports to Tableau is a real cost — must scope."),
        ("Support for industry-specific data standards (FHIR for healthcare, XBRL for finance)",
         "Not needed", "Not applicable to franchise/F&B reporting."),
        ("API versioning and deprecation policy",
         "Low", "Standard hygiene."),
    ]),
    ("15. Self-Service & Authoring Experience", [
        ("Drag-and-drop report building without SQL or coding knowledge",
         "High", "Analysts and business authors expect this baseline."),
        ("Guided analytics and template-based starting points",
         "Medium", "Accelerates new-author onboarding."),
        ("Natural language query for ad-hoc exploration",
         "Low", "Not a stated requirement."),
        ("Field suggestion and smart auto-complete during authoring",
         "Medium", "Authoring quality-of-life."),
        ("In-tool data preparation and transformation (Power Query, Tableau Prep)",
         "Medium", "Most prep happens upstream; light tool-side prep still useful."),
        ("Undo/redo depth and autosave behavior",
         "Medium", "Standard expectation."),
        ("Governed self-service model — ability to expose certified datasets for user-built reports",
         "High", "AtScale-as-certified-source pattern is the architectural bet."),
        ("Sandbox or personal workspace for experimentation without impacting production",
         "Medium", "Standard governance pattern."),
        ("Complexity gradient — simple reports accessible to novices, advanced features for power users",
         "High", "Franchise viewer + central analyst personas must both be served."),
        ("In-product help, tutorials, and contextual guidance",
         "Medium", "Affects ramp-up and training cost."),
    ]),
    ("16. Report Consumption & Alerting", [
        ("Data-driven alerts on threshold breaches (KPI alerts)",
         "Medium", "Useful for ops reports; not blocking."),
        ("Scheduled email subscriptions with report snapshots",
         "High", "Direct match to client requirement #8 (scheduled emails)."),
        ("Personalized filter state persistence per user",
         "High", "Direct match to client requirement #5 (Bookmarks equivalent)."),
        ("In-report commenting and collaborative annotation",
         "Low", "Not a stated requirement."),
        ("Print and export options for consumers (not just authors)",
         "High", "Required for #7 (export capability) at consumer level."),
        ("Bookmark and personal view saving",
         "High", "Direct match to client requirement #5."),
        ("Paginated report support for operational/pixel-perfect outputs",
         "High", "Required for #7 (catering detail report) and operational printing."),
        ("Report embedding in email or intranet portals",
         "Medium", "Possible companion to #8 delivery."),
        ("In-app notification center for alerts and updates",
         "Low", "Email is the primary channel."),
        ("Performance of report load time for consumers on low-bandwidth connections",
         "Medium", "Franchise locations may have varying network quality."),
    ]),
    ("17. Data Lineage & Audit Trail", [
        ("End-to-end lineage from data source through transformation to visual",
         "Medium", "Lineage spans AtScale + BI; useful for impact analysis."),
        ("Impact analysis — which reports are affected if a dataset changes",
         "Medium", "Important during AtScale model changes."),
        ("Column-level lineage visibility",
         "Low", "Deeper than typical operational need."),
        ("Report and dataset usage analytics (who viewed, how often, last accessed)",
         "Medium", "Helps prune dead reports and right-size licenses."),
        ("Integration with enterprise data catalog for lineage federation",
         "Low", "Catalog strategy not stated."),
        ("Audit logs for user activity (logins, report access, data exports, permission changes)",
         "Medium", "Standard governance need."),
        ("Data freshness visibility — last refresh timestamp surfaced to consumers",
         "Medium", "Builds trust with franchise consumers."),
        ("Stale report detection and notification",
         "Low", "Nice-to-have."),
        ("Lineage graph visualization within the tool",
         "Low", "Visual aid; not blocking."),
        ("Retention policy for audit logs and compliance archival",
         "Low", "No specific compliance retention mandate surfaced."),
    ]),
    ("18. Version Control & CI/CD for BI", [
        ("Native Git integration for report and dataset versioning",
         "Medium", "Power BI Fabric and Tableau differ here — worth scoping."),
        ("Branching and merging support for parallel development",
         "Low", "Small author team; lightweight branching usually sufficient."),
        ("Deployment pipeline support (dev → test → prod promotion)",
         "Medium", "Standard SDLC need."),
        ("Diff and comparison tooling for BI artifact changes",
         "Low", "Useful but not blocking."),
        ("Rollback capability to previous versions",
         "Medium", "Operational safety."),
        ("Automated testing framework for reports and measures (data validation, regression)",
         "Low", "Maturity goal; not day-one requirement."),
        ("CLI or API-driven deployment for pipeline automation",
         "Medium", "Supports the automation theme in #8."),
        ("Environment variable and parameter management across environments",
         "Medium", "Standard SDLC need."),
        ("Integration with CI/CD platforms (Azure DevOps, GitHub Actions, Jenkins)",
         "Low", "Possible future; not stated."),
        ("Change approval and release gating workflows",
         "Low", "Not a stated requirement."),
    ]),
    ("19. Real-time & Streaming Analytics", [
        ("Support for streaming data ingestion (Kafka, Event Hubs, Kinesis, MQTT)",
         "Not needed", "Franchise P&L/operational reporting is not real-time."),
        ("Push dataset API for real-time dashboard updates",
         "Not needed", "Not in scope."),
        ("DirectQuery mode availability and performance across source types",
         "High", "Live connection to AtScale is the consumption pattern — must validate (req #2)."),
        ("Configurable auto-refresh intervals for near-real-time use cases",
         "Low", "Daily/scheduled refresh is the norm."),
        ("Incremental refresh support to reduce full dataset reload overhead",
         "Medium", "Relevant for any imported datasets and historical loads."),
        ("Latency SLA between data event and dashboard update",
         "Low", "Not a real-time use case."),
        ("Hybrid mode support (mix of real-time and historical data in one dashboard)",
         "Low", "Not in scope."),
        ("Alert triggering on streaming threshold breaches",
         "Not needed", "No streaming workload."),
        ("Scalability of real-time connections under concurrent viewer load",
         "Low", "Not the workload profile."),
        ("Compatibility with stream processing platforms (Spark Streaming, Flink, Azure Stream Analytics)",
         "Not needed", "Not in scope."),
    ]),
    ("20. Localization & Internationalization", [
        ("UI language support and number of available locale translations",
         "Low", "Likely single-locale internal use; confirm with client."),
        ("Locale-aware number, date, and currency formatting",
         "Medium", "Currency/date formats matter for P&L reporting (#6d)."),
        ("Right-to-left (RTL) language support (Arabic, Hebrew)",
         "Not needed", "No RTL audience indicated."),
        ("Multi-currency display and conversion handling",
         "Low", "Depends on franchise footprint — confirm."),
        ("Time zone management for scheduled refreshes and report timestamps",
         "Medium", "Affects refresh schedules (#2) and consumer-facing timestamps."),
        ("Translation workflow for report labels and field names",
         "Not needed", "No multi-language report requirement surfaced."),
        ("Regional compliance support (GDPR for EU, PDPA for Thailand, LGPD for Brazil)",
         "Low", "Depends on geography; confirm."),
        ("Support for non-Latin character sets (CJK, Devanagari, Cyrillic)",
         "Not needed", "No indication of need."),
        ("Locale-specific sorting and collation rules",
         "Low", "Edge case for franchise reporting."),
        ("Regional data residency options aligned with localization needs",
         "Low", "Depends on geography; confirm."),
    ]),
]

# Priority order for sorting / fill colors
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2, "Not needed": 3}
PRIORITY_FILLS = {
    "High": PatternFill("solid", fgColor="C6EFCE"),       # green
    "Medium": PatternFill("solid", fgColor="FFEB9C"),     # amber
    "Low": PatternFill("solid", fgColor="FFC7CE"),        # pink
    "Not needed": PatternFill("solid", fgColor="D9D9D9"), # grey
}
PRIORITY_FONT = {
    "High": Font(color="006100", bold=True),
    "Medium": Font(color="9C5700", bold=True),
    "Low": Font(color="9C0006", bold=True),
    "Not needed": Font(color="595959", bold=True),
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CATEGORY_FILL = PatternFill("solid", fgColor="D9E1F2")
CATEGORY_FONT = Font(bold=True, size=11)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    wb = Workbook()

    # --- Sheet 1: Evaluation Matrix (one row per breakdown item) ---
    ws = wb.active
    ws.title = "Evaluation Matrix"

    headers = ["S.No", "Category", "Breakdown Item", "Priority", "Reasoning (use case fit)"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    row = 2
    sno = 1
    for category, items in DATA:
        for item, priority, reasoning in items:
            ws.cell(row=row, column=1, value=sno).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=row, column=2, value=category).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row, column=3, value=item).alignment = Alignment(vertical="top", wrap_text=True)
            p_cell = ws.cell(row=row, column=4, value=priority)
            p_cell.fill = PRIORITY_FILLS[priority]
            p_cell.font = PRIORITY_FONT[priority]
            p_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=5, value=reasoning).alignment = Alignment(vertical="top", wrap_text=True)
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = BORDER
            row += 1
            sno += 1

    # Column widths
    widths = {1: 6, 2: 32, 3: 60, 4: 14, 5: 70}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{row - 1}"

    # --- Sheet 2: Priority Summary ---
    ws2 = wb.create_sheet("Priority Summary")
    summary_headers = ["Category", "High", "Medium", "Low", "Not needed", "Total"]
    ws2.append(summary_headers)
    for col_idx, _ in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    totals = {"High": 0, "Medium": 0, "Low": 0, "Not needed": 0}
    r = 2
    for category, items in DATA:
        counts = {"High": 0, "Medium": 0, "Low": 0, "Not needed": 0}
        for _, p, _ in items:
            counts[p] += 1
            totals[p] += 1
        ws2.cell(row=r, column=1, value=category).alignment = Alignment(vertical="center", wrap_text=True)
        for idx, key in enumerate(["High", "Medium", "Low", "Not needed"], start=2):
            c = ws2.cell(row=r, column=idx, value=counts[key])
            c.alignment = Alignment(horizontal="center", vertical="center")
            if counts[key] > 0:
                c.fill = PRIORITY_FILLS[key]
        ws2.cell(row=r, column=6, value=sum(counts.values())).alignment = Alignment(horizontal="center", vertical="center")
        for c_idx in range(1, 7):
            ws2.cell(row=r, column=c_idx).border = BORDER
        r += 1

    # Totals row
    ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for idx, key in enumerate(["High", "Medium", "Low", "Not needed"], start=2):
        c = ws2.cell(row=r, column=idx, value=totals[key])
        c.font = Font(bold=True)
        c.fill = PRIORITY_FILLS[key]
        c.alignment = Alignment(horizontal="center", vertical="center")
    grand = sum(totals.values())
    gc = ws2.cell(row=r, column=6, value=grand)
    gc.font = Font(bold=True)
    gc.alignment = Alignment(horizontal="center", vertical="center")
    for c_idx in range(1, 7):
        ws2.cell(row=r, column=c_idx).border = BORDER

    widths2 = {1: 36, 2: 10, 3: 10, 4: 10, 5: 14, 6: 10}
    for col_idx, width in widths2.items():
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Legend ---
    ws3 = wb.create_sheet("Legend")
    ws3.append(["Priority", "Meaning"])
    for c in range(1, 3):
        cell = ws3.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    legend = [
        ("High",       "Must validate during evaluation; directly maps to a client requirement or is critical to the AtScale-driven architecture."),
        ("Medium",     "Important supporting capability; validate at a lighter depth or rely on vendor documentation."),
        ("Low",        "Nice-to-have or already covered upstream by AtScale/Databricks/ADF; document but do not POC."),
        ("Not needed", "Out of scope for this use case (e.g., embedded analytics, streaming, RTL localization)."),
    ]
    for i, (k, v) in enumerate(legend, start=2):
        kc = ws3.cell(row=i, column=1, value=k)
        kc.fill = PRIORITY_FILLS[k]
        kc.font = PRIORITY_FONT[k]
        kc.alignment = Alignment(horizontal="center", vertical="center")
        kc.border = BORDER
        vc = ws3.cell(row=i, column=2, value=v)
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        vc.border = BORDER

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 110

    out_path = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Total rows: {sno - 1}")
    print(f"Totals: {totals}")


if __name__ == "__main__":
    build()
