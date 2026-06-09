"""Append clarifying questions to column E (Discovery Reasoning) for the 20
rows in the Client-suggestion and Talk-to-GTF buckets.

Questions are appended to whatever is already in the cell, so any user edits
to existing reasoning are preserved.
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

# Keyed by exact text of column C (Breakdown Item)
QUESTIONS = {
    # ---------- CLIENT SUGGESTION rows ----------
    "External user sharing (guest/B2B access)": [
        "Are franchisees GTF employees (internal Azure AD) or independent operators (external/B2B identities)?",
        "Expected count of external/B2B users at go-live and 3 years out?",
        "Any IT/security policy that prevents Azure AD guest invitations?",
    ],
    "On-premises vs. cloud licensing cost differential": [
        "Is on-prem (or IaaS) deployment a hard requirement, or is vendor-managed cloud acceptable?",
        "Any policy preventing reports/data from being hosted in Microsoft or Salesforce cloud?",
        "If on-prem is preferred, what is the driver — sovereignty, existing investment, or latency?",
    ],
    "Deployment modes — SaaS, self-hosted, hybrid": [
        "Is there a preferred deployment model (SaaS vs self-hosted), or is the BI tool free to choose?",
        "If self-hosted, who will own operations — in-house DevOps, an MSP, or vendor professional services?",
        "Any IT/security policy that forces a specific deployment mode?",
    ],
    "Native mobile app availability (iOS and Android)": [
        "Will franchise managers or field staff consume reports on phones/tablets?",
        "Rough split between mobile vs desktop consumers?",
        "Are devices company-issued (MDM-controlled) or BYOD?",
    ],
    "Responsive layout design vs. dedicated mobile layout authoring": [
        "If mobile is in scope, do all reports need a mobile layout, or only a small set of priority dashboards?",
        "Is a desktop layout on a tablet acceptable, or must mobile be a tailored experience?",
    ],
    "UI language support and number of available locale translations": [
        "Are all consumers expected to use English, or are non-English users in scope?",
        "If multi-language, which specific locales must be supported?",
    ],
    "Multi-currency display and conversion handling": [
        "Does GTF operate franchises in a single currency or multiple?",
        "If multi-currency, is FX conversion handled upstream (AtScale/EDW) or expected at the BI tool?",
        "Is a 'reporting currency' (e.g., USD) the standard view, with locale currency optional?",
    ],
    "Time zone management for scheduled refreshes and report timestamps": [
        "Are franchises in a single time zone or multiple?",
        "Should consumer-facing timestamps be in viewer's local time or one canonical time zone (e.g., HQ)?",
        "What time zone should scheduled refreshes and email subscriptions run on?",
    ],
    "Regional compliance support (GDPR for EU, PDPA for Thailand, LGPD for Brazil)": [
        "In which countries/regions does GTF operate franchises?",
        "Which compliance regimes apply (GDPR, PDPA, LGPD, CCPA, etc.)?",
        "Are there data-subject rights workflows (DSAR, right-to-be-forgotten) the BI tool must support?",
    ],
    "Regional data residency options aligned with localization needs": [
        "Must data remain within a specific geography (EU-only, India-only, etc.)?",
        "Are there contractual or regulatory data-residency clauses to satisfy?",
        "Is it acceptable for the BI tool to cache metadata/extracts outside the data region?",
    ],

    # ---------- TALK TO GTF rows ----------
    "Workspace or project-based collaboration with role assignments": [
        "What is GTF's franchise org structure — flat or hierarchical (region → franchise → store)?",
        "How many franchises are expected at go-live and 3 years out?",
        "Are franchises grouped by region, brand, or another dimension?",
        "Preferred isolation model — one workspace per franchise, or shared workspace with RLS?",
    ],
    "Cost at different scale points (50, 500, 5000 users)": [
        "Expected total user count at go-live and at 3 years?",
        "Split between internal GTF staff and franchisee employees?",
        "Any peak/seasonal usage patterns that affect concurrency sizing?",
    ],
    "Distinction between author/creator vs. viewer licensing costs": [
        "How many active report authors are expected (central BI team, regional analysts)?",
        "How many viewers (read-only consumers)?",
        "Will any users need explorer/ad-hoc query rights between author and viewer tiers?",
    ],
    "Contractual flexibility (annual vs. monthly, enterprise agreements)": [
        "Does GTF have an existing Microsoft Enterprise Agreement that could include Power BI?",
        "Existing Salesforce/Tableau contract that could be extended?",
        "Procurement preference — annual, 3-year, or multi-year commitment?",
    ],
    "Hidden costs — gateway infrastructure, training, professional services": [
        "Will the gateway VM run on existing Azure infrastructure or require new provisioning?",
        "Is there a budget line for vendor training/certification for the BI team?",
        "Will the implementation be in-house, vendor PS, or partner-led?",
    ],
    "Vendor discount structures for large enterprise commitments": [
        "Existing spend/relationship with Microsoft (PBI) or Salesforce (Tableau) that could unlock discounts?",
        "Procurement preference — standard pricing vs negotiated EA discount?",
    ],
    "Total 3-year TCO model including infrastructure and admin overhead": [
        "Confirm 3-year horizon, or is a different planning window preferred (e.g., 5-year)?",
        "What inputs does procurement need for the TCO model (per-user, per-capacity, infra, admin FTEs)?",
        "Are existing Power BI licenses sunk cost or recoverable if Tableau is chosen?",
    ],
    "Training and certification program availability and industry recognition": [
        "What is the current BI team's skill mix (Power BI/DAX, Tableau, both)?",
        "Headcount and ramp-up appetite for retraining if Tableau is chosen?",
        "Are there partners or contractors with Tableau skills GTF can lean on?",
    ],
    "Cross-tool report migration tooling availability": [
        "How many existing Power BI reports are in production?",
        "Of those, how many are critical (must migrate) vs disposable (can rebuild)?",
        "Are PBIX source files and semantic model documentation available?",
    ],
    "Performance of report load time for consumers on low-bandwidth connections": [
        "Typical network connectivity at franchise locations (broadband, 4G hotspot, satellite)?",
        "Any specific locations known to have poor or unreliable connectivity?",
        "Target SLA for report first-paint time on a slow connection?",
    ],
}

wb = load_workbook(PATH)

# Locate the matrix sheet by header signature
matrix_sheet_name = None
for sn in wb.sheetnames:
    s = wb[sn]
    if s.cell(row=1, column=1).value == "S.No" and s.cell(row=1, column=3).value == "Breakdown Item":
        matrix_sheet_name = sn
        break
if matrix_sheet_name is None:
    raise RuntimeError(f"Could not locate Evaluation Matrix sheet. Sheets: {wb.sheetnames}")
print(f"Using sheet: {matrix_sheet_name!r}")
ws = wb[matrix_sheet_name]

updated = 0
not_found = []

for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=3).value
    if not item or item not in QUESTIONS:
        continue
    cell = ws.cell(row=r, column=5)
    existing = cell.value or ""
    # If we've already appended (idempotency), skip
    if "Clarifying questions:" in existing:
        # Replace the old block to avoid duplication
        existing = existing.split("\n\nClarifying questions:")[0].rstrip()
    bullet_lines = "\n".join(f"  • {q}" for q in QUESTIONS[item])
    new_text = f"{existing.rstrip()}\n\nClarifying questions:\n{bullet_lines}"
    cell.value = new_text
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    updated += 1

# Track which questions never matched a row (sanity check)
matched_items = set()
for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=3).value
    if item in QUESTIONS:
        matched_items.add(item)
not_found = [k for k in QUESTIONS if k not in matched_items]

# Make column E a bit taller-friendly: increase row heights where we added content
# Excel auto-fits wrap_text rows usually, but we can set a hint.
# Leave row heights to Excel's auto behavior.

wb.save(PATH)
print(f"Rows updated: {updated}")
if not_found:
    print(f"Question keys with no matching row ({len(not_found)}):")
    for k in not_found:
        print(f"  {k}")
else:
    print("All 20 question keys matched a row.")
