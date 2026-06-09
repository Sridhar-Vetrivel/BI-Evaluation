"""Update POC steps in Category 2 (Visualization & UX) to use local Parquet
files from udi_data_gen/output/parquet/ instead of AtScale (since AtScale is
not yet available to the team).

Only the 3 POC rows in Category 2 are touched:
  - Conditional formatting and dynamic visual properties
  - Cross-filtering and drill-down/drill-through behavior
  - Dashboard layout flexibility (pixel-perfect vs. responsive grid)
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

PARQUET_DIR = r"udi_data_gen\output\parquet"

POC_UPDATES = {
    "Conditional formatting and dynamic visual properties": {
        "powerbi_proof":
            "POC steps (Parquet-based — AtScale not yet available):\n"
            f"1. In Power BI Desktop: Get Data → Parquet, point at files in {PARQUET_DIR}\\ "
            "(reported_sales.parquet, brand_targets.parquet, unit.parquet, brand.parquet, dimdate.parquet).\n"
            "2. Build the model: relationships unit↔brand, reported_sales↔unit, brand_targets↔brand, "
            "all↔dimdate (calendar).\n"
            "3. Create measures: [Actual Revenue], [Target Revenue], [Variance] = Actual − Target, "
            "[Variance %] = DIVIDE(Variance, Target).\n"
            "4. Build a matrix visual with Brand and Unit on rows, period on columns; "
            "place Actual, Target, Variance, Variance % as values.\n"
            "5. Apply conditional formatting: background color on Variance (red/green by sign), "
            "data bars on Actual, icons on Variance %. Validate cell-level granularity.\n"
            "6. Add a field parameter to toggle the displayed measure between $ and % views.\n"
            "7. Capture screenshots, authoring time, and any formatting gaps (null handling, large "
            "negatives, totals).\n"
            "Caveat: this validates Power BI's native conditional formatting, NOT the AtScale "
            "pushdown behavior. Re-run row-7 (Live vs Import) once AtScale is available.",
        "tableau_proof":
            "POC steps (Parquet-based — AtScale not yet available):\n"
            "1. In Tableau Desktop (2024.2+): use the Parquet file connector to load the same files from "
            f"{PARQUET_DIR}\\. If on an older Tableau version, convert Parquet → Hyper via Tableau Prep "
            "Builder or the Python Hyper API first.\n"
            "2. Establish relationships in the Tableau data model: unit↔brand, reported_sales↔unit, "
            "brand_targets↔brand, all↔dimdate.\n"
            "3. Create calculated fields: Actual, Target, Variance, Variance %.\n"
            "4. Build an equivalent matrix using a text table with Brand and Unit on Rows, period "
            "on Columns; Measure Names/Values for the four measures.\n"
            "5. Apply conditional color on Variance via calculated color field; use mark properties for "
            "data-bar-style and icon effects.\n"
            "6. Use a parameter to toggle $ vs % display; bind via a calc field.\n"
            "7. Record authoring time and document any cell-level formatting gaps relative to Power BI.\n"
            "Caveat: this validates Tableau's native conditional formatting, NOT AtScale pushdown. "
            "Re-run row-7 (Live vs Extract) once AtScale is available.",
    },

    "Cross-filtering and drill-down/drill-through behavior": {
        "powerbi_proof":
            "POC steps (Parquet-based — AtScale not yet available):\n"
            f"1. Load brand.parquet, unit.parquet, reported_sales.parquet, dimdate.parquet, and "
            f"unit_ranking_base.parquet from {PARQUET_DIR}\\.\n"
            "2. Define a hierarchy Brand → Unit using fields from brand and unit tables.\n"
            "3. Build a page with: a Brand slicer, KPI cards (Total Revenue, Avg Daily Sales), a bar "
            "chart of Revenue by Unit, and a line chart of Revenue over time.\n"
            "4. Validate automatic cross-filtering — selecting a brand in the slicer should propagate "
            "across all visuals on the page.\n"
            "5. Enable drill-down on the bar chart through the Brand → Unit hierarchy; verify each level "
            "fires a correctly-aggregated query.\n"
            "6. Configure a drill-through page to a Unit detail report; right-click a unit to navigate, "
            "passing context.\n"
            "7. Measure interaction latency at each level; note any unexpected re-aggregation behavior.\n"
            "Caveat: Parquet files are loaded in Import mode (no DirectQuery against files). The pushdown "
            "behavior against AtScale will differ — re-validate latency once AtScale is in play.",
        "tableau_proof":
            "POC steps (Parquet-based — AtScale not yet available):\n"
            f"1. Load the same files from {PARQUET_DIR}\\ via the Parquet connector (or Hyper extracts).\n"
            "2. Define a dimension hierarchy Brand → Unit by dragging Unit onto Brand in the data pane.\n"
            "3. Build an equivalent dashboard: Brand filter, KPI text boxes, bar chart by Unit, time-series line chart.\n"
            "4. Configure a Filter Action triggered by the Brand selector to apply across all worksheets on the dashboard.\n"
            "5. Use drilldown on the Brand → Unit hierarchy on the bar chart; verify aggregations.\n"
            "6. Build a drill-to-detail worksheet; configure a Dashboard Action (run on Select) to "
            "navigate with unit context.\n"
            "7. Use Performance Recording to measure latency per level and inspect the underlying queries.\n"
            "Caveat: Parquet is loaded as a Hyper extract (or live read of the file) — not equivalent to "
            "Live connection against AtScale. Re-validate once AtScale is available.",
    },

    "Dashboard layout flexibility (pixel-perfect vs. responsive grid)": {
        "powerbi_proof":
            "POC steps (Parquet-based — AtScale not yet available):\n"
            "1. Pick a representative target — e.g., a Brand-level executive scorecard with KPI cards, "
            "a Unit performance ranking, a revenue trend line, and a guest-satisfaction heatmap.\n"
            f"2. Load reported_sales, brand_targets, unit, brand, guest_satisfaction, dimdate from "
            f"{PARQUET_DIR}\\.\n"
            "3. Build the dashboard in Power BI Desktop using the grid + snap canvas; align visuals manually.\n"
            "4. Author a separate Mobile Layout view for the same report (required for mobile fidelity in PBI).\n"
            "5. Record authoring time, number of manual pixel-alignment adjustments, and any layout limitations.\n"
            "6. Test rendering on desktop (web + Desktop), Power BI mobile app emulator, and tablet view.\n"
            "7. Document fidelity gaps: text wrap, matrix overflow, image scaling, white-space management.\n"
            "Caveat: layout authoring effort is independent of the data source — Parquet vs AtScale should "
            "not affect this comparison.",
        "tableau_proof":
            "POC steps (Parquet-based — AtScale not yet available):\n"
            "1. Build the SAME target dashboard in Tableau Desktop, mixing Tiled and Floating layout containers.\n"
            f"2. Load the same parquet files from {PARQUET_DIR}\\ (Parquet connector or Hyper extracts).\n"
            "3. Use Device Designer to author Desktop / Tablet / Phone layouts within the same workbook.\n"
            "4. Record authoring time and pixel-alignment effort vs Power BI; compare apples-to-apples.\n"
            "5. Publish to Tableau Cloud (or Server, if available); test on web, iPad, and Tableau Mobile.\n"
            "6. Document layout fidelity gaps side-by-side with the Power BI version.\n"
            "Caveat: layout effort is data-source-independent — finding should carry over to AtScale.",
    },
}

# ---------- apply to workbook ----------
wb = load_workbook(PATH)

matrix_sheet_name = None
for sn in wb.sheetnames:
    s = wb[sn]
    if s.cell(row=1, column=1).value == "S.No" and s.cell(row=1, column=3).value == "Breakdown Item":
        matrix_sheet_name = sn
        break
if matrix_sheet_name is None:
    raise RuntimeError(f"Could not locate Evaluation Matrix sheet. Sheets: {wb.sheetnames}")
print(f"Using sheet: {matrix_sheet_name!r}")
ws = wb[matrix_sheet_name]

COL_ITEM   = 3
COL_PBI_PR = 8
COL_TAB_PR = 9

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

updated_rows = []
for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=COL_ITEM).value
    if not item or item not in POC_UPDATES:
        continue
    upd = POC_UPDATES[item]
    pbi_cell = ws.cell(row=r, column=COL_PBI_PR, value=upd["powerbi_proof"])
    tab_cell = ws.cell(row=r, column=COL_TAB_PR, value=upd["tableau_proof"])
    for c in (pbi_cell, tab_cell):
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
    updated_rows.append(r)

wb.save(PATH)
print(f"Updated rows ({len(updated_rows)}): {updated_rows}")
print(f"Expected 3 POC rows in Category 2.")
