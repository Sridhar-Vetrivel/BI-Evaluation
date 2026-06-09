"""Fill the BI-tool evaluation columns (F-I) for Category 2: Visualization & UX."""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

PINK = PatternFill(patternType="solid", fgColor="FFC7CE")

EVAL = {
    "Library of standard chart types (bar, line, scatter, pie, waterfall, funnel, bullet)": {
        "powerbi":
            "Mostly capable. Bar, line, scatter, pie, waterfall, funnel are native default visuals. "
            "BULLET chart is NOT a default visual — available via certified AppSource visuals "
            "(e.g., MAQ Software Bullet Chart, OKViz Bullet Chart) or built as a workaround on a bar visual.",
        "tableau":
            "Capable. All listed (bar, line, scatter, pie, waterfall, funnel, bullet) are natively available "
            "via the 'Show Me' panel. Bullet chart is a first-class native chart type in Tableau.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/visuals/power-bi-visualization-types-for-reports-and-q-and-a\n"
            "AppSource bullet visuals: https://appsource.microsoft.com/ (search 'bullet chart')",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/dataview_examples.htm\n"
            "Bullet chart: https://help.tableau.com/current/pro/desktop/en-us/qs_bullet_graphs.htm",
    },

    "Advanced/specialized chart types (heatmaps, treemaps, Sankey, chord, small multiples)": {
        "powerbi":
            "Mostly capable. Treemap and small multiples are native (small multiples GA since 2022). "
            "Heatmap via matrix + conditional formatting. SANKEY and CHORD are NOT native — require "
            "certified custom visuals from AppSource (MAQ Software Sankey, ZoomCharts, etc.).",
        "tableau":
            "Mostly capable. Treemap, heatmap (highlight table / density), and small multiples (via trellis "
            "layouts) are native. SANKEY and CHORD are NOT native — built via data densification techniques, "
            "community templates, or Tableau Extensions.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/visuals/power-bi-visualization-small-multiples\n"
            "https://learn.microsoft.com/en-us/power-bi/visuals/power-bi-visualization-treemaps",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/buildexamples_treemap.htm\n"
            "https://www.tableau.com/blog/build-sankey-diagram-tableau-without-any-data-prep (community pattern)",
    },

    "Custom visualization support (D3.js, Vega-Lite, marketplace extensions)": {
        "powerbi":
            "Capable. AppSource hosts hundreds of certified custom visuals. Custom Visual SDK uses TypeScript + "
            "D3.js (direct D3 support). Power BI also supports R and Python visuals for advanced charts. "
            "No native Vega-Lite, but D3 integration is first-class.",
        "tableau":
            "Capable. Tableau Exchange hosts extensions. Dashboard Extensions API (JavaScript) allows embedded "
            "extensions and external visualization libraries (including D3.js wrapped in an extension). "
            "Viz Extensions (newer, 2024) allow custom mark types. No native Vega-Lite.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/developer/visuals/\n"
            "https://appsource.microsoft.com/en-us/marketplace/apps?product=power-bi-visuals",
        "tableau_proof":
            "Official documentation:\n"
            "https://tableau.github.io/extensions-api/\n"
            "https://exchange.tableau.com/extensions",
    },

    "Conditional formatting and dynamic visual properties": {
        "powerbi":
            "Best-in-class. Conditional formatting on background color, font color, data bars, icons, and "
            "even web URLs — all drivable by DAX measures, including measures sourced from AtScale via XMLA. "
            "Field parameters allow user-toggled visual properties. Calculation groups enable format-string "
            "switching. Excellent for matrix-style P&L reports.",
        "tableau":
            "Capable but more setup. Color rules driven by calculated fields; bold/italic and shape variants "
            "via mark properties. Less granular than Power BI for matrix-style cell-by-cell conditional formatting. "
            "Parameter actions enable dynamic visual properties. Works against AtScale measures but typically "
            "requires LOD/calc field wrappers.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "POC steps:\n"
            "1. Build a P&L matrix in Power BI Desktop connected live (DirectQuery) to AtScale.\n"
            "2. Apply conditional formatting on a Variance measure (background = red/green based on sign).\n"
            "3. Add a data-bar conditional format on Actuals; verify it queries AtScale and not a local cache.\n"
            "4. Use field parameters to let users toggle between $ vs % views.\n"
            "5. Capture the DAX/MDX queries fired against AtScale (DAX Studio) and confirm pushdown.\n"
            "6. Validate format rendering for negatives, nulls, and large numbers in P&L context.",
        "tableau_proof":
            "POC steps:\n"
            "1. Build an equivalent P&L matrix in Tableau Desktop connected Live to AtScale.\n"
            "2. Create calculated fields for Variance and conditional color logic; apply to mark color.\n"
            "3. Use a parameter to toggle between $ vs % display; bind to displayed measure via calc field.\n"
            "4. Inspect generated MDX via Performance Recording — confirm pushdown to AtScale.\n"
            "5. Validate handling of negatives, nulls, and large numbers in P&L context.\n"
            "6. Compare authoring effort vs Power BI for the same target output.",
    },

    "Cross-filtering and drill-down/drill-through behavior": {
        "powerbi":
            "Capable. Automatic cross-filtering between visuals on the same page. Drill-down on dimension "
            "hierarchies (AtScale-defined hierarchies surface as drillable). Drill-through pages with "
            "context-preserving navigation. Bookmarks capture filter/visual state.",
        "tableau":
            "Capable, different paradigm. Cross-filtering via Dashboard Actions (Filter, Highlight, URL). "
            "Drilldown via dimension hierarchies and Set Actions. 'Drill-through' achieved via dashboard "
            "actions opening another worksheet with passed context. Parameter Actions for richer interactivity.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "POC steps:\n"
            "1. Build a multi-visual page in Power BI live-connected to AtScale.\n"
            "2. Place a slicer on Franchise and validate cross-filter cascade across all visuals.\n"
            "3. Configure a drill-down hierarchy (Region → Franchise → Store) using AtScale dimensions; "
            "verify each level fires a properly-aggregated query to AtScale.\n"
            "4. Build a drill-through page to a detail report; pass franchise context and validate.\n"
            "5. Measure latency at each drill level — confirm AtScale aggregations are used (not full-table scan).",
        "tableau_proof":
            "POC steps:\n"
            "1. Build an equivalent multi-visual dashboard in Tableau Desktop live-connected to AtScale.\n"
            "2. Configure a Filter Action from a franchise selector to all worksheets on the dashboard.\n"
            "3. Define dimension hierarchies (Region → Franchise → Store) — verify drilldown fires correct MDX.\n"
            "4. Build a drill-to-detail worksheet via dashboard action; pass franchise context.\n"
            "5. Measure latency at each level using Performance Recording; confirm AtScale aggregations are hit.",
    },

    "Tooltip customization and rich interactivity": {
        "powerbi":
            "Capable. Report page tooltips — author a tooltip page with multiple visuals, attach it to a "
            "host visual; hover reveals the page. Default tooltips support measure selection and formatting. "
            "Custom visuals can render HTML tooltips.",
        "tableau":
            "Capable, with a signature feature. 'Viz in Tooltip' — embed a fully rendered worksheet inside "
            "a tooltip on another worksheet (industry-leading for chart-in-tooltip). HTML formatting in "
            "tooltip text. Parameters and calculated fields can drive tooltip content dynamically.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-tooltips\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-tooltips-page",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/viz_in_tooltip.htm\n"
            "https://help.tableau.com/current/pro/desktop/en-us/tooltips.htm",
    },

    "Dashboard layout flexibility (pixel-perfect vs. responsive grid)": {
        "powerbi":
            "Capable but grid-oriented. Canvas is a fixed-size grid with snap-to-grid behavior. Responsive "
            "fit within breakpoints. Mobile layout requires SEPARATE authoring (Mobile Layout view per report). "
            "Less freeform than Tableau for pixel-perfect executive layouts.",
        "tableau":
            "Capable, more flexible. Two layout modes: TILED (snap) and FLOATING (free-form, pixel-perfect). "
            "Device Designer authors desktop/tablet/phone layouts in one workbook. Layout containers (horizontal/vertical) "
            "for responsive grouping. Edge for executive 'pretty' dashboards and P&L matrix layouts.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "POC steps:\n"
            "1. Choose a target executive 'pretty' dashboard layout (e.g., GTF executive P&L summary).\n"
            "2. Build it in Power BI Desktop using grid + snap; configure mobile layout separately.\n"
            "3. Record authoring time and number of adjustments needed for pixel alignment.\n"
            "4. Test rendering across desktop, tablet, and Power BI mobile app.\n"
            "5. Document any layout fidelity gaps (text wrap, image scaling, matrix overflow).",
        "tableau_proof":
            "POC steps:\n"
            "1. Build the SAME target dashboard in Tableau Desktop using a mix of Tiled + Floating containers.\n"
            "2. Use Device Designer to author desktop / tablet / phone layouts in the same workbook.\n"
            "3. Record authoring time and pixel-alignment effort vs Power BI.\n"
            "4. Test rendering on Tableau Cloud/Server, iPad, and Tableau Mobile app.\n"
            "5. Document any layout fidelity gaps and compare to Power BI side-by-side.",
    },

    "Theming and branding consistency across reports": {
        "powerbi":
            "Capable. JSON theme files (.json) define color palette, fonts, default visual styles, and "
            "background images. Themes apply across reports in an org. Built-in theme gallery + custom theme "
            "publishing. Strong support for org-wide brand standardization.",
        "tableau":
            "Capable, less centralized. Workbook-level formatting (fonts, colors, lines, shading). Saved "
            "workbook templates serve as starting points. No native JSON-style theme files — branding "
            "consistency typically relies on style guides + template workbooks.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-report-themes\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-report-themes#import-custom-report-themes",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/formatting_workbook.htm\n"
            "https://help.tableau.com/current/pro/desktop/en-us/save_savework_template.htm",
    },

    "Animation and transitions for storytelling": {
        "powerbi":
            "Limited. Play Axis (animated scatter / bubble over time). Visual transitions on filter/slicer "
            "changes are minimal. Bookmark Navigator + Selection Pane enable a guided storytelling flow "
            "via bookmarks, but no dedicated 'story' authoring surface.",
        "tableau":
            "Capable. Pages Shelf for animated time progression. 'Story Points' is a dedicated storytelling "
            "feature — sequence of slides with annotations, captions, and navigation. Mark Animations "
            "(since 2020.1) animate transitions between worksheet states. Strong for narrative dashboards.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/visuals/power-bi-visualization-scatter#add-play-axis\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-bookmarks",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/story_create.htm\n"
            "https://help.tableau.com/current/pro/desktop/en-us/formatting_animations.htm",
    },

    "Accessibility of visuals (screen reader compatibility, color-blind palettes)": {
        "powerbi":
            "Capable. Screen reader support (Narrator, NVDA, JAWS) for most native visuals. Keyboard "
            "navigation built in. Accessibility Insights checker integration. Color-blind safe palettes "
            "available as built-in themes. Microsoft publishes WCAG 2.1 AA conformance reports.",
        "tableau":
            "Capable. Screen reader support for published views on Tableau Server/Cloud (improved since 2019.2). "
            "Keyboard navigation. 'Color Blind 10' is one of the default color palettes. Tableau publishes "
            "VPAT (Voluntary Product Accessibility Template) and WCAG 2.1 AA conformance docs.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-accessibility-overview\n"
            "https://learn.microsoft.com/en-us/power-bi/create-reports/desktop-accessibility-creating-reports",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/accessibility_overview.htm\n"
            "https://www.tableau.com/about/legal/accessibility (VPAT)",
    },
}

# ---------- apply to workbook ----------
wb = load_workbook(PATH)

matrix_sheet_name = None
for sn in wb.sheetnames:
    s = wb[sn]
    if s.cell(row=1, column=1).value == "S.No" and s.cell(row=1, column=3).value == "Breakdown Item":
        matrix_sheet_name = sn
        break
if matrix_sheet_name is None:
    raise RuntimeError(f"Could not locate Evaluation Matrix sheet. Sheets: {wb.sheetnames}")
print(f"Using sheet: {matrix_sheet_name!r}")
ws = wb[matrix_sheet_name]

COL_ITEM   = 3
COL_PBI    = 6
COL_TAB    = 7
COL_PBI_PR = 8
COL_TAB_PR = 9

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

updated_rows = []
for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=COL_ITEM).value
    if not item or item not in EVAL:
        continue
    data = EVAL[item]
    cells = {
        COL_PBI:    (data["powerbi"],        data["powerbi_pink"]),
        COL_TAB:    (data["tableau"],        data["tableau_pink"]),
        COL_PBI_PR: (data["powerbi_proof"],  False),
        COL_TAB_PR: (data["tableau_proof"],  False),
    }
    for col, (text, pink) in cells.items():
        cell = ws.cell(row=r, column=col, value=text)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        if pink:
            cell.fill = PINK
    updated_rows.append(r)

wb.save(PATH)
print(f"Updated rows ({len(updated_rows)}): {updated_rows}")
print("Expected 10 (Visualization & UX).")
