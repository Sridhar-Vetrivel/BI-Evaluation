"""Patch Discovery.xlsx so Priority Summary uses live COUNTIFS formulas.

Preserves any in-progress edits in the Evaluation Matrix sheet — only the
Priority Summary cells (counts + totals) are rewritten as formulas.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

PRIORITY_FILLS = {
    "High": PatternFill("solid", fgColor="C6EFCE"),
    "Medium": PatternFill("solid", fgColor="FFEB9C"),
    "Low": PatternFill("solid", fgColor="FFC7CE"),
    "Not needed": PatternFill("solid", fgColor="D9D9D9"),
}

wb = load_workbook(PATH)
matrix = wb["Evaluation Matrix"]
summary = wb["Priority Summary"]

# Find the actual data range in Evaluation Matrix (row 1 is header)
last_row = matrix.max_row
matrix_range_cat = f"'Evaluation Matrix'!$B$2:$B${last_row}"
matrix_range_pri = f"'Evaluation Matrix'!$D$2:$D${last_row}"
print(f"Matrix data: rows 2..{last_row}")

# Identify summary rows (everything between header row 1 and the TOTAL row)
total_row_idx = None
for r in range(2, summary.max_row + 1):
    if summary.cell(row=r, column=1).value == "TOTAL":
        total_row_idx = r
        break

if total_row_idx is None:
    raise RuntimeError("Could not locate TOTAL row in Priority Summary")

print(f"Summary category rows: 2..{total_row_idx - 1}, TOTAL row: {total_row_idx}")

priorities = ["High", "Medium", "Low", "Not needed"]

# Per-category COUNTIFS rows
for r in range(2, total_row_idx):
    cat_cell_ref = f"$A{r}"
    for idx, p in enumerate(priorities, start=2):
        col_letter = summary.cell(row=r, column=idx).column_letter
        formula = (
            f'=COUNTIFS({matrix_range_cat},{cat_cell_ref},'
            f'{matrix_range_pri},"{p}")'
        )
        cell = summary.cell(row=r, column=idx, value=formula)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Keep priority color fill always visible (not conditional on count > 0
        # anymore, since count is now dynamic)
        cell.fill = PRIORITY_FILLS[p]
    # Total column = sum of the four count cells in this row
    b = summary.cell(row=r, column=2).coordinate
    e = summary.cell(row=r, column=5).coordinate
    total_cell = summary.cell(row=r, column=6, value=f"=SUM({b}:{e})")
    total_cell.alignment = Alignment(horizontal="center", vertical="center")

# TOTAL row — sum each priority column across all category rows
for idx, p in enumerate(priorities, start=2):
    col_letter = summary.cell(row=total_row_idx, column=idx).column_letter
    top = f"{col_letter}2"
    bot = f"{col_letter}{total_row_idx - 1}"
    cell = summary.cell(row=total_row_idx, column=idx, value=f"=SUM({top}:{bot})")
    cell.font = Font(bold=True)
    cell.fill = PRIORITY_FILLS[p]
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Grand total = sum of all four totals in the TOTAL row
b = summary.cell(row=total_row_idx, column=2).coordinate
e = summary.cell(row=total_row_idx, column=5).coordinate
gc = summary.cell(row=total_row_idx, column=6, value=f"=SUM({b}:{e})")
gc.font = Font(bold=True)
gc.alignment = Alignment(horizontal="center", vertical="center")

wb.save(PATH)
print(f"Saved: {PATH}")
print("Priority Summary now uses live COUNTIFS formulas.")
