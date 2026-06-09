"""Replace column E in Discovery.xlsx with discovery-angle reasoning.

Preserves user edits to priority and other columns. Matches rows by the
breakdown-item text in column C, so reordered rows still get the right
reasoning written back.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

# Standard phrases for "no architectural angle" rows
NO_ANGLE = "No AtScale-specific angle — evaluate as native tool capability via vendor documentation."
OUT_OF_SCOPE_EMBED = "Out of scope — internal BI use case, not embedded analytics."
OUT_OF_SCOPE_STREAM = "Out of scope — franchise/P&L reporting is not a streaming workload."
OUT_OF_SCOPE_LOC = "Out of scope unless franchise geography requires it — confirm with client."

# Discovery reasoning keyed by breakdown-item text (must match column C exactly)
REASONING = {
    # 1. Data Connectivity & Integration
    "Native connectors to structured sources (SQL Server, Oracle, Snowflake, BigQuery, Redshift, SAP HANA)":
        "Primary consumption is via AtScale, not direct DB. Validate that each tool's native DB connector exists as a fallback path for DBA/ad-hoc workflows, but is not the default pattern.",
    "Support for semi-structured and unstructured sources (JSON, XML, REST APIs, web scraping)":
        NO_ANGLE,
    "File-based ingestion (Excel, CSV, Parquet, Avro)":
        "Analysts may run ad-hoc analyses against Excel/CSV that bypass AtScale. Confirm both tools support common file formats for one-off work without breaking governance.",
    "Cloud storage connectors (S3, Azure Blob, GCS)":
        "Curated data lands via Databricks/ADF — BI tools should not connect to Blob directly. Validate that this is a guardrail discussion (policy) rather than a capability gap.",
    "ERP/CRM connectors (SAP, Salesforce, Dynamics 365, ServiceNow)":
        "Source systems land in EDW upstream; BI tool never touches them directly. No discovery action needed.",
    "Custom connector development capability (SDK support)":
        "No bespoke source planned beyond AtScale. " + NO_ANGLE,
    "Live connection vs. import mode availability per source":
        "CRITICAL — live connection to AtScale (XMLA/MDX) is the architectural consumption pattern. Validate that each tool supports live mode without silently falling back to import, and benchmark latency under realistic load.",
    "OAuth and token-based authentication support for API sources":
        "Not relevant for AtScale/SSAS connection (Kerberos/Azure AD path). " + NO_ANGLE,
    "Support for on-premises data sources via gateway":
        "Maps to client requirement #2. Compare Power BI On-Premises Data Gateway vs Tableau Bridge for connecting to AtScale (whether AtScale is on-prem or cloud-hosted) — sizing, HA, and licensing.",
    "Connector certification and update frequency by vendor":
        "Only the AtScale connector matters. Confirm whether AtScale or the BI vendor maintains it, and what the support SLA is when AtScale upgrades MDX behavior.",

    # 2. Visualization & UX
    "Library of standard chart types (bar, line, scatter, pie, waterfall, funnel, bullet)":
        NO_ANGLE,
    "Advanced/specialized chart types (heatmaps, treemaps, Sankey, chord, small multiples)":
        NO_ANGLE,
    "Custom visualization support (D3.js, Vega-Lite, marketplace extensions)":
        NO_ANGLE,
    "Conditional formatting and dynamic visual properties":
        "P&L reporting (#6d) requires variance/ratio conditional formatting in matrix layouts. Validate that conditional rules work on measures sourced from AtScale (not just on local calcs).",
    "Cross-filtering and drill-down/drill-through behavior":
        "Interactive dashboards (#6a) and P&L drilldowns (#6d) rely on AtScale-defined hierarchies. Validate that drilldowns trigger correct MDX/SQL queries to AtScale and do not over-fetch.",
    "Tooltip customization and rich interactivity":
        NO_ANGLE,
    "Dashboard layout flexibility (pixel-perfect vs. responsive grid)":
        "Executive ‘pretty’ dashboards (#6c) and P&L matrix (#6d) need pixel control. Power BI is grid-based; Tableau is freeform — compare authoring effort for the same target layout.",
    "Theming and branding consistency across reports":
        NO_ANGLE,
    "Animation and transitions for storytelling":
        NO_ANGLE,
    "Accessibility of visuals (screen reader compatibility, color-blind palettes)":
        NO_ANGLE,

    # 3. Governance, Security & Compliance
    "Role-based access control (RBAC) at report, dataset, and workspace level":
        "Maps to #3 + #4. Validate that AtScale roles + BI-tool workspace roles compose cleanly without role duplication or privilege escalation.",
    "Row-level security (RLS) and column-level security (CLS)":
        "CRITICAL — does franchise-level RLS defined at the AtScale layer propagate end-to-end to Tableau/Power BI users without re-implementation? Test the franchise-filtering scenario from #4 in both tools.",
    "Single Sign-On (SSO) integration (SAML, OIDC, Azure AD, Okta)":
        "Maps to #1. Validate the full Azure AD/Entra ID flow: user → BI tool → AtScale, with the user identity passed through (not via service principal), so RLS works.",
    "Data classification and sensitivity label enforcement":
        NO_ANGLE,
    "Compliance certifications held by vendor (SOC 2, ISO 27001, HIPAA, FedRAMP, GDPR)":
        NO_ANGLE,
    "Data residency and sovereignty controls":
        NO_ANGLE,
    "Tenant-level isolation in multi-tenant deployments":
        "Single-tenant internal deployment. " + NO_ANGLE,
    "Policy enforcement for sharing (prevent external sharing, download restrictions)":
        "Franchise data leakage is a real risk. Validate share-restriction policies and download controls in both tools — particularly for paginated/export workflows.",
    "Encryption at rest and in transit":
        NO_ANGLE,
    "Periodic access reviews and certification workflows":
        "Likely managed at Azure AD layer, not the BI tool. " + NO_ANGLE,

    # 4. Performance & Scalability
    "Query response time under concurrent user load (SLA benchmarks)":
        "Initial AtScale test was a one-shot signal. Extend to multi-user concurrent load against AtScale via each BI tool, with realistic franchise-scoped queries.",
    "Dataset size limits (row count, in-memory vs. DirectQuery thresholds)":
        "Catering detail (#7) drives this. Validate row-limit behavior of each tool when querying AtScale — does it stream, paginate, truncate, or error? Power BI DirectQuery row limits vs Tableau live connection behavior.",
    "Aggregation and caching strategies (pre-aggregation, query caching, dual mode)":
        "CRITICAL — how does each BI tool's cache interact with AtScale's pre-aggregations? Does the tool defer aggregation to AtScale or duplicate it locally (defeating AtScale's purpose)?",
    "Horizontal scaling capability for large deployments":
        "Relevant only for Tableau Server self-hosted (capacity nodes) and Power BI Premium (capacity SKUs). Sizing model differs sharply — validate against franchise concurrency profile.",
    "Performance profiling and query diagnostics tooling":
        "When a P&L is slow, can you trace from BI-tool query → AtScale query plan → underlying source? Validate observability across the AtScale boundary in both tools.",
    "Handling of high-cardinality dimensions":
        "Franchise × product × date combinations create high cardinality. Test slicer/filter UX and rendering performance at realistic cardinality.",
    "Scheduled refresh frequency limits and concurrency":
        "Maps to #2. Since AtScale is live, ‘refresh’ here means metadata refresh + any imported datasets. Validate concurrency limits per license tier.",
    "Impact of complex calculations on render time":
        "P&L variance, ratio, YoY calculations may run in AtScale or in the BI tool. Validate where each calc lives and the render-time tradeoff.",
    "CDN support for global report distribution":
        "Internal franchise audience, likely regional. " + NO_ANGLE,
    "Load testing support and documented capacity planning models":
        "Validate vendor-published sizing guides for AtScale-fronted workloads (not just generic).",

    # 5. Collaboration & Sharing
    "Workspace or project-based collaboration with role assignments":
        "Maps to #3. Compare Power BI App/Workspace hierarchy vs Tableau Project/Site/Collection — which maps cleanly to franchise grouping (one franchise = one project? one site?).",
    "Report commenting and annotation features":
        NO_ANGLE,
    "Sharing via link, embed, or direct user assignment":
        "Maps to #3. Validate share-by-link with Azure AD enforcement — links must not bypass RLS or franchise scoping.",
    "External user sharing (guest/B2B access)":
        "If franchisees are guest tenants in Azure AD, validate the B2B flow end-to-end (login + RLS) and check per-guest license cost in both tools.",
    "Subscription and snapshot sharing (static vs. live)":
        "Maps to #8. Validate franchise-filtered snapshot delivery — each tool's subscription model must support per-recipient parameter values.",
    "Co-authoring or concurrent editing support":
        NO_ANGLE,
    "Notification and change alert mechanisms":
        NO_ANGLE,
    "Integration with collaboration platforms (Teams, Slack, email)":
        "Maps to #8 (email). Compare native subscription email vs Power Automate flows vs Tableau Subscriptions vs custom REST API automation.",
    "Version history and rollback on shared content":
        NO_ANGLE,
    "Approval workflows for publishing reports":
        NO_ANGLE,

    # 6. Licensing & TCO
    "Licensing model (per user, per capacity/core, consumption-based)":
        "Power BI (per-user + Premium capacity) vs Tableau (Creator/Explorer/Viewer) are structurally different. Model each against franchise user-count and consumption pattern.",
    "Cost at different scale points (50, 500, 5000 users)":
        "Franchise viewer count is the cost driver. Build a per-user cost curve for both tools at GTF's expected scale.",
    "Distinction between author/creator vs. viewer licensing costs":
        "Most franchise users are read-only viewers — viewer SKU cost dominates 3-year TCO. Compare Power BI Pro/PPU/Premium-per-User vs Tableau Viewer.",
    "Premium or capacity-based tiers and what they unlock":
        "Power BI Premium unlocks paginated reports, large model sizes, RLS at scale, deployment pipelines. Map these unlocks to GTF's requirements to see what tier is mandatory.",
    "Cost of connectors, add-ons, and marketplace extensions":
        "Confirm whether the AtScale connector incurs additional licensing on either the BI-vendor or AtScale side. Include in TCO.",
    "On-premises vs. cloud licensing cost differential":
        "Tableau Server (on-prem/IaaS) vs Tableau Cloud is a real architectural fork. Power BI is SaaS-only for cloud; PBI Report Server exists but is feature-limited. Compare.",
    "Contractual flexibility (annual vs. monthly, enterprise agreements)":
        NO_ANGLE,
    "Hidden costs — gateway infrastructure, training, professional services":
        "Gateway VM sizing for live AtScale connection is a material cost. Include training (Tableau learning curve > Power BI for Microsoft shops) and migration services.",
    "Vendor discount structures for large enterprise commitments":
        NO_ANGLE,
    "Total 3-year TCO model including infrastructure and admin overhead":
        "Final decision artifact. Roll up license + capacity + gateway + admin + training + migration costs over 3 years for both tools.",

    # 7. AI/ML & Advanced Analytics
    "Native AI visuals (anomaly detection, forecasting, key influencers)":
        NO_ANGLE,
    "Natural language query interface (Q&A, Ask Data)":
        NO_ANGLE,
    "Integration with ML platforms (Azure ML, SageMaker, Databricks, Vertex AI)":
        "ML lives in Databricks layer, not in BI tool. " + NO_ANGLE,
    "Python and R script execution within the tool":
        NO_ANGLE,
    "AutoML or no-code ML model building":
        "Out of scope for this evaluation.",
    "Explainability features for AI-generated insights":
        "Out of scope for this evaluation.",
    "Smart narrative and automated insight generation":
        NO_ANGLE,
    "Predictive analytics and what-if scenario modeling":
        "What-if analysis is called out in #6a. Validate parameter-driven what-if in both tools (Power BI parameters/calc groups vs Tableau parameters/actions).",
    "Vector/embedding search or GenAI integration roadmap":
        NO_ANGLE,
    "Ability to surface model outputs as first-class visuals":
        NO_ANGLE,

    # 8. Deployment & IT Operations
    "Deployment modes — SaaS, self-hosted, hybrid":
        "Power BI is SaaS-only (cloud); Tableau offers SaaS (Cloud), self-hosted (Server), or hybrid. Affects the ops model significantly — decide upfront.",
    "Cloud platform support (Azure, AWS, GCP, on-premises)":
        "GTF stack is Azure (ADF, Databricks, AAS). Power BI is native Azure; Tableau runs anywhere. Validate Azure-native integration depth for both.",
    "High availability and disaster recovery architecture":
        "Standard, but document how each tool fits the existing Azure DR strategy (e.g., paired regions, RTO/RPO).",
    "Upgrade and patch management process (vendor-managed vs. self-managed)":
        "Power BI Service is fully vendor-managed; Tableau Server self-hosted = self-managed (operational overhead). Factor into TCO and ops staffing.",
    "Infrastructure-as-code support for provisioning":
        NO_ANGLE,
    "Admin portal capabilities (usage monitoring, capacity management, user management)":
        "Operating at franchise scale requires usage analytics, capacity monitoring, and bulk user management. Compare Power BI admin portal vs Tableau Server/Cloud admin.",
    "Multi-environment support (dev, staging, production)":
        NO_ANGLE,
    "Backup and restore capabilities for BI content":
        NO_ANGLE,
    "SLA commitments and incident response times":
        NO_ANGLE,
    "Container and Kubernetes deployment support (for self-hosted options)":
        "Only relevant if Tableau Server self-hosted route is chosen. " + NO_ANGLE,

    # 9. Data Modeling & Semantic Layer
    "Support for star and snowflake schema modeling":
        "AtScale owns the semantic model. Validate that each BI tool consumes AtScale's flattened model 1:1 without over-imposing its own modeling layer.",
    "Calculated columns, measures, and KPI definitions (DAX, Tableau calculations)":
        "Some report-local calcs are inevitable (variance vs. plan, dynamic ratios). Validate that local DAX/Tableau calcs don't compromise AtScale governance — i.e., don't replicate AtScale measures locally.",
    "Reusable certified dataset / semantic model concept":
        "CRITICAL — how does each BI tool surface AtScale as a ‘certified’ source? Power BI's Datasets vs Tableau's Published Data Sources vs direct AtScale connection — which gives the cleanest governance story?",
    "Relationships and cardinality handling (many-to-many, bi-directional)":
        "Resolved at AtScale. Validate edge cases (e.g., many-to-many fact tables) surface correctly in each BI tool.",
    "Hierarchy definition and drill path management":
        "P&L hierarchies (#6d) are defined in AtScale. Validate that each BI tool surfaces them as native drillable hierarchies (not flattened lists).",
    "Time intelligence functions (YTD, MTD, rolling periods)":
        "Split decision: define time intelligence in AtScale's calendar dim or in BI-tool DAX/calcs? Validate both approaches and recommend a single home for it.",
    "Centralized metric layer support (compatibility with dbt Semantic Layer, AtScale, Cube)":
        "CRITICAL — this is the architectural bet. Validate AtScale's connector maturity, MDX/XMLA fidelity, and supported features (RLS, hierarchies, calc members) for each BI tool.",
    "Ability to govern and version the semantic model independently of reports":
        "AtScale handles model versioning. Validate that BI reports do not break silently when AtScale model changes (renamed measure, removed dimension).",
    "Support for composite models (mixing live and imported data)":
        "If analyst Excel uploads must combine with AtScale data, validate composite model support. Power BI composite models (DirectQuery + Import) vs Tableau data blending.",
    "Data type handling and implicit vs. explicit conversion behavior":
        "MDX type semantics differ from SQL/DAX. Validate that decimals, dates, and currencies survive the AtScale → BI-tool boundary without silent precision loss in P&L numbers.",

    # 10. Embedded Analytics
    "Embedding via iFrame, JavaScript SDK, or REST API":
        OUT_OF_SCOPE_EMBED,
    "Row-level security propagation in embedded context":
        OUT_OF_SCOPE_EMBED,
    "White-labeling and full UI customization for embedded scenarios":
        OUT_OF_SCOPE_EMBED,
    "Token-based authentication for embedded sessions (no vendor login required)":
        OUT_OF_SCOPE_EMBED,
    "Licensing model for embedded usage (app-owns-data vs. user-owns-data)":
        OUT_OF_SCOPE_EMBED,
    "Interaction API for programmatic control of filters and visuals":
        OUT_OF_SCOPE_EMBED,
    "Performance of embedded reports at scale":
        OUT_OF_SCOPE_EMBED,
    "Cross-origin (CORS) and Content Security Policy (CSP) compliance":
        OUT_OF_SCOPE_EMBED,
    "Support for embedding in mobile apps (iOS/Android)":
        OUT_OF_SCOPE_EMBED,
    "Analytics portal pattern support (multi-tenant embedded deployments)":
        OUT_OF_SCOPE_EMBED,

    # 11. Extensibility & Developer Ecosystem
    "REST API coverage (CRUD on reports, datasets, workspaces, users)":
        "Maps to #8. Validate API surface for programmatic subscription delivery, parameterized exports, and bulk franchise user provisioning.",
    "Webhook and event-driven integration support":
        NO_ANGLE,
    "Custom visual development framework and marketplace":
        NO_ANGLE,
    "CLI tooling for automation and scripting":
        NO_ANGLE,
    "SDK availability and language support (Python, .NET, JavaScript, REST)":
        NO_ANGLE,
    "Third-party extension ecosystem maturity":
        NO_ANGLE,
    "Programmatic report generation (templating, parameterized reports)":
        "Maps to #8 — franchise-specific report delivery requires parameterized reports. Compare Power BI paginated parameters vs Tableau parameterized views.",
    "Integration with orchestration tools (Azure Data Factory, Airflow, dbt)":
        "ADF already orchestrates upstream (Databricks → AtScale). Validate whether ADF can trigger BI dataset/extract refresh too, or whether each tool's scheduler is preferred.",
    "Community-contributed connectors and visuals quality and maintenance":
        NO_ANGLE,
    "Developer documentation quality and completeness":
        NO_ANGLE,

    # 12. Vendor Maturity & Ecosystem
    "Gartner Magic Quadrant / Forrester Wave positioning and trajectory":
        NO_ANGLE,
    "Product release cadence and public roadmap transparency":
        "Power BI ships monthly; Tableau quarterly. Track whether AtScale connector improvements appear on each roadmap.",
    "Vendor financial stability and long-term viability":
        NO_ANGLE,
    "Size and activity of community (forums, Stack Overflow, GitHub)":
        NO_ANGLE,
    "Training and certification program availability and industry recognition":
        "Affects hiring/upskilling cost in TCO. GTF likely has more Power BI skills in-house given the Microsoft stack.",
    "Partner ecosystem for implementation and support":
        NO_ANGLE,
    "User group and conference ecosystem (Tableau Conference, Microsoft Fabric Community)":
        NO_ANGLE,
    "Support tier options (standard, premier, dedicated TAM)":
        NO_ANGLE,
    "Responsiveness to enterprise support tickets":
        NO_ANGLE,
    "History of backward compatibility and migration support during major version changes":
        NO_ANGLE,

    # 13. Mobile & Accessibility
    "Native mobile app availability (iOS and Android)":
        "Franchise managers/field staff likely consume on mobile. Validate that AtScale-backed reports render acceptably in each tool's mobile app with RLS intact.",
    "Responsive layout design vs. dedicated mobile layout authoring":
        "Authoring overhead — Power BI requires a separate mobile layout per report; Tableau supports device-specific layouts. Estimate effort delta for franchise dashboards.",
    "Offline report access on mobile":
        NO_ANGLE,
    "Touch-optimized interaction design":
        NO_ANGLE,
    "Push notification support for alerts on mobile":
        NO_ANGLE,
    "WCAG 2.1 AA compliance for accessibility":
        NO_ANGLE,
    "Keyboard navigation support in desktop and web interfaces":
        NO_ANGLE,
    "Screen reader compatibility (NVDA, JAWS, VoiceOver)":
        NO_ANGLE,
    "Color contrast and color-blind safe defaults":
        NO_ANGLE,
    "Accessibility audit tooling or built-in checker":
        NO_ANGLE,

    # 14. Interoperability & Standards
    "ODBC/JDBC connectivity for third-party tool access":
        "AtScale exposes standard endpoints. " + NO_ANGLE,
    "OData endpoint exposure for report data consumption":
        NO_ANGLE,
    "Compatibility with enterprise metadata catalogs (Microsoft Purview, Alation, Collibra, Atlan)":
        NO_ANGLE,
    "Export formats supported (PDF, Excel, CSV, PNG, PowerPoint)":
        "Maps to #7 (export) and #8 (snapshot delivery). Validate that exports preserve formatting for P&L matrices and large detail-row reports.",
    "Open standard support (Apache Arrow, XMLA endpoint, ANSI SQL)":
        "CRITICAL — XMLA is the protocol AtScale exposes for semantic-layer consumption. Validate XMLA fidelity in both tools (does each speak full MDX/DAX over XMLA, or is functionality reduced?).",
    "Integration with data catalog and data dictionary platforms":
        NO_ANGLE,
    "Compatibility with ETL/ELT tools (Informatica, Talend, Fivetran, dbt)":
        "ETL is upstream of AtScale. " + NO_ANGLE,
    "Cross-tool report migration tooling availability":
        "Migrating existing Power BI reports to Tableau is a real one-time cost if Tableau wins. Survey available migration tools, manual effort, and report-by-report rewrite estimates.",
    "Support for industry-specific data standards (FHIR for healthcare, XBRL for finance)":
        "Not applicable to franchise/F&B reporting.",
    "API versioning and deprecation policy":
        NO_ANGLE,

    # 15. Self-Service & Authoring Experience
    "Drag-and-drop report building without SQL or coding knowledge":
        NO_ANGLE,
    "Guided analytics and template-based starting points":
        NO_ANGLE,
    "Natural language query for ad-hoc exploration":
        NO_ANGLE,
    "Field suggestion and smart auto-complete during authoring":
        NO_ANGLE,
    "In-tool data preparation and transformation (Power Query, Tableau Prep)":
        "Most data prep lives upstream in Databricks/ADF. Light tool-side prep is still useful for ad-hoc — compare Power Query vs Tableau Prep capability and license inclusion.",
    "Undo/redo depth and autosave behavior":
        NO_ANGLE,
    "Governed self-service model — ability to expose certified datasets for user-built reports":
        "The AtScale-as-certified-source pattern is the architectural bet. Validate how each BI tool exposes AtScale as the certified base for analyst-built reports.",
    "Sandbox or personal workspace for experimentation without impacting production":
        NO_ANGLE,
    "Complexity gradient — simple reports accessible to novices, advanced features for power users":
        NO_ANGLE,
    "In-product help, tutorials, and contextual guidance":
        NO_ANGLE,

    # 16. Report Consumption & Alerting
    "Data-driven alerts on threshold breaches (KPI alerts)":
        NO_ANGLE,
    "Scheduled email subscriptions with report snapshots":
        "Maps to #8. Validate per-franchise filtered email delivery — Power BI subscriptions + Power Automate vs Tableau Subscriptions + custom logic.",
    "Personalized filter state persistence per user":
        "Maps to #5 (Bookmarks equivalent). Validate per-user filter persistence: Power BI Personal Bookmarks vs Tableau Custom Views.",
    "In-report commenting and collaborative annotation":
        NO_ANGLE,
    "Print and export options for consumers (not just authors)":
        "Maps to #7. Consumers (franchisees) need self-serve export, not just authors. Validate consumer-side export controls in both tools.",
    "Bookmark and personal view saving":
        "Direct match to client requirement #5. Compare Power BI Bookmarks (page-level state) vs Tableau Custom Views (filter/sort state) — feature-by-feature parity check.",
    "Paginated report support for operational/pixel-perfect outputs":
        "HIGH — maps to #7 (catering detail). Power BI Paginated Reports (RDL, Premium-gated) vs Tableau (no native paginated equivalent; relies on long worksheets or Tableau Prep flows). Material gap to validate.",
    "Report embedding in email or intranet portals":
        "Companion to #8. Validate inline email embed (image vs interactive link) for both tools.",
    "In-app notification center for alerts and updates":
        NO_ANGLE,
    "Performance of report load time for consumers on low-bandwidth connections":
        "Franchise locations may have varying network quality. Validate first-paint time and progressive rendering on slow connections.",

    # 17. Data Lineage & Audit Trail
    "End-to-end lineage from data source through transformation to visual":
        "Lineage spans AtScale + BI tool. Validate whether each tool exposes lineage that integrates with AtScale's lineage (or at minimum surfaces the AtScale model as a single lineage node).",
    "Impact analysis — which reports are affected if a dataset changes":
        "When AtScale model changes (renamed measure, dropped dim), which BI reports break? Validate impact-analysis tooling in both BI tools and whether it can read AtScale metadata.",
    "Column-level lineage visibility":
        NO_ANGLE,
    "Report and dataset usage analytics (who viewed, how often, last accessed)":
        "Helps right-size franchise viewer licenses. Compare Power BI usage metrics vs Tableau Server/Cloud usage analytics depth.",
    "Integration with enterprise data catalog for lineage federation":
        NO_ANGLE,
    "Audit logs for user activity (logins, report access, data exports, permission changes)":
        NO_ANGLE,
    "Data freshness visibility — last refresh timestamp surfaced to consumers":
        "AtScale refresh state must surface in the BI tool so franchise users trust the data. Validate whether each tool shows ‘as of’ timestamps sourced from AtScale (not just BI-tool extract refresh).",
    "Stale report detection and notification":
        NO_ANGLE,
    "Lineage graph visualization within the tool":
        NO_ANGLE,
    "Retention policy for audit logs and compliance archival":
        NO_ANGLE,

    # 18. Version Control & CI/CD for BI
    "Native Git integration for report and dataset versioning":
        "Power BI Fabric (PBIP + Git) is improving rapidly; Tableau requires storing TWBX/TFL files in Git with no native diff. Compare maturity and developer experience.",
    "Branching and merging support for parallel development":
        NO_ANGLE,
    "Deployment pipeline support (dev → test → prod promotion)":
        "Power BI Deployment Pipelines (Premium) is mature; Tableau relies on Server projects + manual or API-driven promotion. Significant gap to evaluate.",
    "Diff and comparison tooling for BI artifact changes":
        NO_ANGLE,
    "Rollback capability to previous versions":
        NO_ANGLE,
    "Automated testing framework for reports and measures (data validation, regression)":
        NO_ANGLE,
    "CLI or API-driven deployment for pipeline automation":
        NO_ANGLE,
    "Environment variable and parameter management across environments":
        NO_ANGLE,
    "Integration with CI/CD platforms (Azure DevOps, GitHub Actions, Jenkins)":
        NO_ANGLE,
    "Change approval and release gating workflows":
        NO_ANGLE,

    # 19. Real-time & Streaming Analytics
    "Support for streaming data ingestion (Kafka, Event Hubs, Kinesis, MQTT)":
        OUT_OF_SCOPE_STREAM,
    "Push dataset API for real-time dashboard updates":
        OUT_OF_SCOPE_STREAM,
    "DirectQuery mode availability and performance across source types":
        "HIGH — live connection to AtScale IS the consumption pattern. Validate Power BI DirectQuery vs Tableau Live Connection semantics against AtScale: query folding, latency, RLS pass-through.",
    "Configurable auto-refresh intervals for near-real-time use cases":
        NO_ANGLE,
    "Incremental refresh support to reduce full dataset reload overhead":
        "Relevant for any imported datasets (analyst extracts on top of AtScale). Validate incremental refresh patterns in both tools.",
    "Latency SLA between data event and dashboard update":
        OUT_OF_SCOPE_STREAM,
    "Hybrid mode support (mix of real-time and historical data in one dashboard)":
        NO_ANGLE,
    "Alert triggering on streaming threshold breaches":
        OUT_OF_SCOPE_STREAM,
    "Scalability of real-time connections under concurrent viewer load":
        NO_ANGLE,
    "Compatibility with stream processing platforms (Spark Streaming, Flink, Azure Stream Analytics)":
        OUT_OF_SCOPE_STREAM,

    # 20. Localization & Internationalization
    "UI language support and number of available locale translations":
        OUT_OF_SCOPE_LOC,
    "Locale-aware number, date, and currency formatting":
        "Currency/date formatting matters for P&L (#6d) even in a single-locale deployment. Validate that AtScale-sourced numeric measures format correctly in both tools.",
    "Right-to-left (RTL) language support (Arabic, Hebrew)":
        OUT_OF_SCOPE_LOC,
    "Multi-currency display and conversion handling":
        OUT_OF_SCOPE_LOC,
    "Time zone management for scheduled refreshes and report timestamps":
        "Affects refresh schedules (#2) and consumer-facing timestamps. Validate time-zone handling for cross-region franchises if applicable.",
    "Translation workflow for report labels and field names":
        OUT_OF_SCOPE_LOC,
    "Regional compliance support (GDPR for EU, PDPA for Thailand, LGPD for Brazil)":
        OUT_OF_SCOPE_LOC,
    "Support for non-Latin character sets (CJK, Devanagari, Cyrillic)":
        OUT_OF_SCOPE_LOC,
    "Locale-specific sorting and collation rules":
        OUT_OF_SCOPE_LOC,
    "Regional data residency options aligned with localization needs":
        OUT_OF_SCOPE_LOC,
}

wb = load_workbook(PATH)

# Sheet may have been renamed during review — locate the matrix sheet
# by finding a sheet whose A1 reads "S.No" (the original header signature).
matrix_sheet_name = None
for sn in wb.sheetnames:
    s = wb[sn]
    if s.cell(row=1, column=1).value == "S.No" and s.cell(row=1, column=3).value == "Breakdown Item":
        matrix_sheet_name = sn
        break
if matrix_sheet_name is None:
    raise RuntimeError(f"Could not locate Evaluation Matrix sheet. Sheets present: {wb.sheetnames}")
print(f"Using sheet: {matrix_sheet_name!r}")
ws = wb[matrix_sheet_name]

# Update header E1
ws.cell(row=1, column=5, value="Discovery Reasoning (AtScale → Tableau/Power BI)")

# Iterate data rows, match by column C (breakdown item), write column E
unmatched = []
written = 0
last_row = ws.max_row
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for r in range(2, last_row + 1):
    item = ws.cell(row=r, column=3).value
    if not item:
        continue
    text = REASONING.get(item)
    if text is None:
        unmatched.append((r, item))
        continue
    cell = ws.cell(row=r, column=5, value=text)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = border
    written += 1

# Optional: widen column E a bit since reasoning is longer now
ws.column_dimensions["E"].width = 80

wb.save(PATH)
print(f"Rows updated: {written}")
if unmatched:
    print(f"Unmatched rows (no reasoning written): {len(unmatched)}")
    for r, item in unmatched:
        print(f"  Row {r}: {item}")
else:
    print("All rows matched.")
