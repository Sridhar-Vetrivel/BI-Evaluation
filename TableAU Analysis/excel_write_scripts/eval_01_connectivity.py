"""Fill the BI-tool evaluation columns (F-I) for Category 1: Data Connectivity & Integration."""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"c:\Users\sridhar.vetrivel\OneDrive - psiog.com\Accounts\GTF\TableAU Analysis\Discovery.xlsx"

# Pink fill for "tool is incapable of this breakdown item"
PINK = PatternFill(patternType="solid", fgColor="FFC7CE")

# Per-breakdown-item evaluation content.
# Each value is a dict:
#   {
#     "powerbi": str,            # capability statement
#     "tableau": str,            # capability statement
#     "powerbi_pink": bool,      # mark F cell pink
#     "tableau_pink": bool,      # mark G cell pink
#     "powerbi_proof": str,      # H — docs link OR POC steps
#     "tableau_proof": str,      # I — docs link OR POC steps
#   }
EVAL = {
    "Native connectors to structured sources (SQL Server, Oracle, Snowflake, BigQuery, Redshift, SAP HANA)": {
        "powerbi":
            "Capable. Native certified connectors for all listed sources. "
            "SQL Server is best-in-class (Microsoft-native). Snowflake, BigQuery, Redshift, Oracle, SAP HANA "
            "all support Import + DirectQuery modes. SAP HANA has a dedicated certified connector with SSO.",
        "tableau":
            "Capable. Native connectors for all listed sources, both Live and Extract modes. "
            "SAP HANA, Snowflake, BigQuery, Redshift, Oracle, SQL Server are first-class connectors with "
            "vendor-certified drivers. Tableau publishes connector compatibility matrices per release.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-data-sources\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/exampleconnections_overview.htm",
    },

    "Support for semi-structured and unstructured sources (JSON, XML, REST APIs, web scraping)": {
        "powerbi":
            "Capable. JSON and XML are native via Power Query. REST APIs supported via the Web connector "
            "with Power Query M functions (pagination, auth, transformations). Basic web scraping via "
            "Web.Page/Web.BrowserContents. Transformation logic must live in Power Query/M.",
        "tableau":
            "Capable but lighter. Native JSON connector; XML via WDC or pre-processing. REST APIs via "
            "Web Data Connector (JavaScript-based) or Tableau Prep with custom scripts. No native generic "
            "web scraping — typically pre-processed upstream.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/web/web\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/json",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/examples_json.htm\n"
            "https://help.tableau.com/current/api/webdataconnector/en-us/",
    },

    "File-based ingestion (Excel, CSV, Parquet, Avro)": {
        "powerbi":
            "Mostly capable. Excel, CSV, Parquet (incl. Delta/Fabric) are native. "
            "Avro is NOT natively supported in Power BI Desktop — workaround is to convert upstream "
            "(Databricks/ADF) or use dataflows with Synapse.",
        "tableau":
            "Mostly capable. Excel, CSV, JSON are native. Parquet via Tableau Hyper API or Tableau Prep "
            "(also via cloud connectors like Athena/Snowflake). Avro is NOT natively supported — must "
            "convert upstream.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/excel\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/parquet",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/examples_text.htm\n"
            "https://tableau.github.io/hyper-db/docs/",
    },

    "Cloud storage connectors (S3, Azure Blob, GCS)": {
        "powerbi":
            "Capable, Azure-favored. Azure Blob Storage and ADLS Gen2 are best-in-class native connectors. "
            "AWS S3 supported via Power Query S3 connector (preview/GA depending on region) or indirectly "
            "through Athena/Snowflake. GCS via Power Query GCS connector.",
        "tableau":
            "Capable, cloud-neutral. Amazon S3 connector, Azure Data Lake Storage Gen2 connector, and "
            "Google Cloud Storage connector all native. Tableau treats the three major clouds as first-class peers.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/azureblobstorage\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/amazon-s3",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/examples_amazons3.htm\n"
            "https://help.tableau.com/current/pro/desktop/en-us/examples_azure_data_lake_gen2.htm",
    },

    "ERP/CRM connectors (SAP, Salesforce, Dynamics 365, ServiceNow)": {
        "powerbi":
            "Capable, Microsoft ecosystem advantage. Dynamics 365 + Dataverse: deepest native integration. "
            "SAP BW and SAP HANA: certified connectors. Salesforce Objects/Reports: native. "
            "ServiceNow: via REST/OData or certified partner connector.",
        "tableau":
            "Capable, Salesforce ecosystem advantage. Salesforce (Tableau is Salesforce-owned) — deepest native. "
            "SAP HANA: certified. Dynamics 365: via OData/REST. "
            "ServiceNow: via OData/REST or partner connector.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-query/connectors/ (full connector list)",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/exampleconnections_overview.htm",
    },

    "Custom connector development capability (SDK support)": {
        "powerbi":
            "Capable. Power Query SDK (Visual Studio and VS Code extensions) for building custom connectors "
            "using the M language. Connectors can be certified by Microsoft and listed in the Power BI "
            "connector gallery. OAuth flows and credential management built-in.",
        "tableau":
            "Capable. Two paths: (1) Tableau Connector SDK for native connectors over JDBC/ODBC with custom "
            "TDC dialect files; (2) Web Data Connector (WDC) framework — JavaScript-based, runs in browser. "
            "Hyper API for programmatic extract writes.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-query/install-sdk\n"
            "https://github.com/Microsoft/DataConnectors",
        "tableau_proof":
            "Official documentation:\n"
            "https://tableau.github.io/connector-plugin-sdk/\n"
            "https://help.tableau.com/current/api/webdataconnector/en-us/",
    },

    "Live connection vs. import mode availability per source": {
        "powerbi":
            "Capable. Three modes: Import (in-memory VertiPaq), DirectQuery (live pushdown), and Composite/Dual. "
            "DirectQuery supported for most relational, cloud DW, and OLAP sources (SQL Server, Snowflake, "
            "BigQuery, Redshift, SAP HANA, Synapse, SSAS/AAS, and AtScale via the Analysis Services XMLA endpoint). "
            "Mode is per-table in composite models.",
        "tableau":
            "Capable. Two modes: Live Connection (pushdown) or Extract (Hyper engine, in-memory columnar). "
            "Live connection supported for most relational, cloud DW, and OLAP sources including SAP HANA, "
            "Snowflake, BigQuery, Redshift, Microsoft Analysis Services (SSAS/AtScale via XMLA). "
            "Hybrid possible via published data sources mixing live + extract.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "POC steps:\n"
            "1. Install Power BI Desktop; connect to AtScale via the Analysis Services connector using the XMLA endpoint.\n"
            "2. Choose 'Connect live' (DirectQuery to SSAS-compatible source).\n"
            "3. Build a representative report against an AtScale cube (P&L summary).\n"
            "4. Use Performance Analyzer + DAX Studio/SQL Profiler to confirm queries are pushed to AtScale (no local cache).\n"
            "5. Benchmark Live vs Import for the same report (record load + interaction latency).\n"
            "6. Validate the connection does NOT silently switch to Import on publish to Power BI Service.",
        "tableau_proof":
            "POC steps:\n"
            "1. Install Tableau Desktop; obtain AtScale's Tableau TDC file (or use the certified AtScale connector).\n"
            "2. Connect Live to the AtScale XMLA endpoint (Microsoft Analysis Services connector).\n"
            "3. Build the same representative worksheet/dashboard as in the Power BI POC.\n"
            "4. Use Performance Recording (Help → Settings and Performance → Start Performance Recording) "
            "to confirm queries hit AtScale (MDX/SQL pushdown).\n"
            "5. Benchmark Live vs Extract for the same workbook.\n"
            "6. Publish to Tableau Server/Cloud and verify Live connection is preserved (not silently extracted).",
    },

    "OAuth and token-based authentication support for API sources": {
        "powerbi":
            "Capable. OAuth 2.0 supported natively for many cloud connectors (Microsoft Entra ID, Dataverse, "
            "Snowflake, BigQuery, Salesforce, SharePoint, etc.). Custom OAuth flows can be implemented in "
            "custom connectors via the Power Query SDK. Service Principal auth supported for "
            "Power BI Service automation.",
        "tableau":
            "Capable. OAuth 2.0 supported natively for many cloud connectors (Salesforce, Google BigQuery, "
            "Snowflake, OneDrive, etc.). PAT (Personal Access Token) and JWT supported for REST API. "
            "Custom OAuth via Web Data Connector framework.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/connect-data/service-gateway-oauth-credentials-flow\n"
            "https://learn.microsoft.com/en-us/power-query/handling-authentication",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/server/en-us/config_oauth_google.htm\n"
            "https://help.tableau.com/current/server/en-us/security_auth.htm",
    },

    "Support for on-premises data sources via gateway": {
        "powerbi":
            "Capable. On-Premises Data Gateway (Standard mode) supports DirectQuery, scheduled refresh, "
            "and dataflows for on-prem sources. HA via gateway clusters. Kerberos constrained delegation "
            "for SSO. Single gateway can serve multiple workspaces. Mature and widely deployed.",
        "tableau":
            "Capable. Tableau Bridge (for Tableau Cloud) supports both live queries and extract refresh from "
            "on-prem sources. Bridge pools provide HA. For Tableau Server (on-prem/IaaS), no gateway needed — "
            "Server itself sits next to the data. Bridge is newer than PBI gateway but production-ready.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "POC steps:\n"
            "1. Provision On-Premises Data Gateway on a Windows VM with network access to AtScale.\n"
            "2. Register the gateway in Power BI Service (Settings → Manage gateways).\n"
            "3. Add AtScale as a data source on the gateway (Analysis Services/XMLA connector + service account).\n"
            "4. Publish a Power BI DirectQuery report to the Service and bind it to the gateway.\n"
            "5. Validate connectivity, scheduled refresh, and SSO (Kerberos delegation if enabled).\n"
            "6. Test HA by adding a second gateway node to the cluster and forcibly stopping the primary.\n"
            "7. Capture sizing recommendations from the gateway performance monitor.\n"
            "Reference docs: https://learn.microsoft.com/en-us/data-integration/gateway/",
        "tableau_proof":
            "POC steps:\n"
            "1. Install Tableau Bridge on a Windows machine with network access to AtScale.\n"
            "2. Sign in to Tableau Cloud and register the Bridge client.\n"
            "3. Configure AtScale as a 'Private Network' data source via the Bridge client.\n"
            "4. Publish a workbook with a Live connection to AtScale via Bridge.\n"
            "5. Validate query latency, scheduled extract refreshes, and connection persistence.\n"
            "6. Add a second Bridge client to form a pool; test HA by stopping the primary.\n"
            "Reference docs: https://help.tableau.com/current/online/en-us/qs_refresh_local_data.htm",
    },

    "Connector certification and update frequency by vendor": {
        "powerbi":
            "Capable. Microsoft maintains the Analysis Services / XMLA connector used to reach AtScale. "
            "Power BI Desktop ships monthly with connector updates. AtScale is a Microsoft partner — "
            "compatibility is tested against Power BI Premium/Fabric XMLA endpoints. SLA-backed support "
            "through Microsoft for the connector.",
        "tableau":
            "Capable. Tableau ships connector updates with each release (~quarterly major, monthly minor). "
            "AtScale publishes an official Tableau connector and tests against Tableau Desktop/Server. "
            "Tableau's MS Analysis Services connector handles XMLA/MDX for AtScale cubes.",
        "powerbi_pink": False,
        "tableau_pink": False,
        "powerbi_proof":
            "Official documentation:\n"
            "https://learn.microsoft.com/en-us/power-bi/connect-data/desktop-analysis-services-tabular-data\n"
            "Power BI Desktop release notes (monthly): https://learn.microsoft.com/en-us/power-bi/fundamentals/desktop-latest-update",
        "tableau_proof":
            "Official documentation:\n"
            "https://help.tableau.com/current/pro/desktop/en-us/examples_olap.htm\n"
            "AtScale-Tableau integration: https://documentation.atscale.com/ (consult AtScale docs for the certified Tableau connector matrix)",
    },
}

# ---------- apply to workbook ----------
wb = load_workbook(PATH)

# Locate matrix sheet
matrix_sheet_name = None
for sn in wb.sheetnames:
    s = wb[sn]
    if s.cell(row=1, column=1).value == "S.No" and s.cell(row=1, column=3).value == "Breakdown Item":
        matrix_sheet_name = sn
        break
if matrix_sheet_name is None:
    raise RuntimeError(f"Could not locate Evaluation Matrix sheet. Sheets: {wb.sheetnames}")
print(f"Using sheet: {matrix_sheet_name!r}")
ws = wb[matrix_sheet_name]

# Column indices
COL_ITEM   = 3   # C
COL_PBI    = 6   # F
COL_TAB    = 7   # G
COL_PBI_PR = 8   # H
COL_TAB_PR = 9   # I

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

updated_rows = []
for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=COL_ITEM).value
    if not item or item not in EVAL:
        continue
    data = EVAL[item]
    cells = {
        COL_PBI:    (data["powerbi"],        data["powerbi_pink"]),
        COL_TAB:    (data["tableau"],        data["tableau_pink"]),
        COL_PBI_PR: (data["powerbi_proof"],  False),
        COL_TAB_PR: (data["tableau_proof"],  False),
    }
    for col, (text, pink) in cells.items():
        cell = ws.cell(row=r, column=col, value=text)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        if pink:
            cell.fill = PINK
    updated_rows.append(r)

# Reasonable column widths for the eval columns
ws.column_dimensions["F"].width = 55
ws.column_dimensions["G"].width = 55
ws.column_dimensions["H"].width = 55
ws.column_dimensions["I"].width = 55

wb.save(PATH)
print(f"Updated rows ({len(updated_rows)}): {updated_rows}")
print(f"Expected 10 (Data Connectivity & Integration).")
